from __future__ import annotations

from dataclasses import dataclass
import copy
import json
import os
import smtplib
import uuid
from datetime import datetime, timedelta
from email.message import EmailMessage
from functools import wraps
from io import BytesIO
import math
import re
import warnings
from typing import Any

import numpy as np
import pandas as pd
from pandas.errors import EmptyDataError
from statsmodels.tsa.statespace.sarimax import SARIMAX
from xgboost import XGBClassifier
from flask import Flask, jsonify, redirect, render_template, request, session, send_file, url_for
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

from db import get_db_connection


app = Flask(__name__)
app.secret_key = "stockwise-dev-secret-key"

@app.after_request
def add_no_cache_headers(response):
    """Prevent browser back-button caching from showing protected pages after logout."""
    try:
        if response.mimetype == "text/html":
            response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
            response.headers["Pragma"] = "no-cache"
            response.headers["Expires"] = "0"
    except Exception:
        pass
    return response



ALLOWED_EXTENSIONS = {"csv", "xlsx", "xls"}
ALLOWED_IMAGE_EXTENSIONS = {"png", "jpg", "jpeg", "gif", "webp"}
PROFILE_UPLOAD_SUBDIR = "profiles"
STORE_LOGO_UPLOAD_SUBDIR = "store_logos"
REQUIRED_UPLOAD_COLUMNS = ["date", "product_name", "quantity_sold"]
OPTIONAL_UPLOAD_COLUMNS = ["time", "category", "current_stock", "reorder_point", "unit_price", "unit_type"]
MAX_UPLOAD_SIZE_BYTES = int(os.environ.get("STOCKWISE_MAX_UPLOAD_SIZE_BYTES", str(8 * 1024 * 1024)))
MAX_SYNC_MODEL_ROWS = int(os.environ.get("STOCKWISE_MAX_SYNC_MODEL_ROWS", "3000"))
MAX_SYNC_SARIMA_PRODUCTS = int(os.environ.get("STOCKWISE_MAX_SYNC_SARIMA_PRODUCTS", "18"))
MAX_SYNC_MODEL_SECONDS = int(os.environ.get("STOCKWISE_MAX_SYNC_MODEL_SECONDS", "18"))

DATA_FORMAT_SYSTEM_FIELDS = [
    {"key": "transaction_date", "label": "Transaction Date Column", "internal": "date", "required": True, "default_column": "Transaction Date"},
    {"key": "transaction_time", "label": "Transaction Time Column", "internal": "time", "required": False, "default_column": "Transaction Time"},
    {"key": "product_id", "label": "Product ID Column", "internal": "product_id", "required": False, "default_column": "Product ID"},
    {"key": "product_name", "label": "Product Name Column", "internal": "product_name", "required": True, "default_column": "Product Name"},
    {"key": "category", "label": "Category Column", "internal": "category", "required": False, "default_column": "Category"},
    {"key": "quantity_sold", "label": "Quantity Sold Column", "internal": "quantity_sold", "required": True, "default_column": "Quantity Sold"},
    {"key": "unit_price", "label": "Unit Price Column", "internal": "unit_price", "required": False, "default_column": "Unit Price"},
    {"key": "current_stock", "label": "Current Stock Column", "internal": "current_stock", "required": False, "default_column": "Current Stock"},
    {"key": "reorder_point", "label": "Reorder Point Column", "internal": "reorder_point", "required": False, "default_column": "Reorder Point"},
    {"key": "is_payday_period", "label": "Payday Period Column", "internal": "is_payday_period", "required": False, "default_column": "Payday Period"},
]

DEFAULT_COLUMN_MAPPING = {field["key"]: field["default_column"] for field in DATA_FORMAT_SYSTEM_FIELDS}
STANDARD_TEMPLATE_COLUMNS = [field["default_column"] for field in DATA_FORMAT_SYSTEM_FIELDS]

UPLOAD_MODE_LABELS = {
    "new": "New dataset",
    "append": "Append missing records",
    "replace": "Replace previous processed data",
}

RETAIL_CATEGORY_OPTIONS = [
    "Rice & Staples",
    "Noodles, Pasta & Canned Goods",
    "Snacks, Biscuits & Sweets",
    "Drinks & Powdered Beverages",
    "Coffee & Breakfast Items",
    "Condiments, Sauces & Cooking Needs",
    "Eggs, Chilled & Frozen Items",
    "Bread & Bakery",
    "Personal Care",
    "Laundry & Cleaning",
    "Household Supplies",
    "Health & First Aid",
    "Baby Care",
    "School & Office Supplies",
    "Mobile Load & Digital Services",
    "Cigarettes & Lighters",
    "Alcoholic Drinks",
    "Pet Care",
    "Hardware & Emergency Items",
    "Seasonal / Miscellaneous Items",
]

CATEGORY_KEYWORD_MAP = [
    ("Rice & Staples", ["rice", "staple", "bigas", "grain", "sugar", "flour", "asin", "salt"]),
    ("Noodles, Pasta & Canned Goods", ["noodle", "pasta", "canned", "can goods", "sardine", "corned", "meat loaf", "mami", "instant"]),
    ("Snacks, Biscuits & Sweets", ["snack", "chips", "biscuit", "sweet", "candy", "chocolate", "cracker", "junk food", "chichirya"]),
    ("Drinks & Powdered Beverages", ["drink", "beverage", "juice", "softdrink", "soft drink", "water", "powdered", "gatas", "milk", "tang", "beverages"]),
    ("Coffee & Breakfast Items", ["coffee", "breakfast", "creamer", "cereal", "oats", "milo", "choco drink"]),
    ("Condiments, Sauces & Cooking Needs", ["condiment", "sauce", "cooking", "oil", "vinegar", "toyo", "soy", "ketchup", "seasoning", "magic sarap"]),
    ("Eggs, Chilled & Frozen Items", ["egg", "chilled", "frozen", "hotdog", "ham", "ice", "cold"]),
    ("Bread & Bakery", ["bread", "bakery", "bun", "pandesal", "cake"]),
    ("Personal Care", ["personal", "shampoo", "soap", "toothpaste", "toothbrush", "deodorant", "lotion"]),
    ("Laundry & Cleaning", ["laundry", "cleaning", "detergent", "bleach", "fabcon", "dishwashing", "cleaner"]),
    ("Household Supplies", ["household", "tissue", "napkin", "plastic", "foil", "bag", "battery"]),
    ("Health & First Aid", ["health", "first aid", "medicine", "gamot", "alcohol", "bandage", "vitamin"]),
    ("Baby Care", ["baby", "diaper", "infant", "wipes"]),
    ("School & Office Supplies", ["school", "office", "pen", "pencil", "paper", "notebook", "load paper"]),
    ("Mobile Load & Digital Services", ["mobile", "load", "digital", "gcash", "paymaya", "ewallet", "e-wallet", "service"]),
    ("Cigarettes & Lighters", ["cigarette", "lighter", "yosi", "smoke"]),
    ("Alcoholic Drinks", ["alcoholic", "beer", "gin", "liquor", "wine", "alak"]),
    ("Pet Care", ["pet", "dog", "cat", "animal"]),
    ("Hardware & Emergency Items", ["hardware", "emergency", "candle", "match", "flashlight", "rope", "nail"]),
]


def standardize_product_category(value: Any) -> str:
    """Map uploaded category text to StockWise's fixed sari-sari store category list for display and filters."""
    raw = str(value or "").strip()
    if not raw or raw.lower() in {"nan", "none", "null", "uncategorized"}:
        return "Seasonal / Miscellaneous Items"

    for category in RETAIL_CATEGORY_OPTIONS:
        if raw.casefold() == category.casefold():
            return category

    normalized = re.sub(r"[^a-z0-9]+", " ", raw.casefold()).strip()
    for category, keywords in CATEGORY_KEYWORD_MAP:
        if any(keyword in normalized for keyword in keywords):
            return category

    return "Seasonal / Miscellaneous Items"


def get_category_filter_options() -> list[str]:
    return list(RETAIL_CATEGORY_OPTIONS)

DEFAULT_SETTINGS = {
    "store_name": "Selected Retailer - Pasig City",
    "store_type": "Sari-sari Store",
    "location_area": "Pasig City",
    "currency": "PHP",
    "default_upload_mode": "new",
    "default_time_range": "30",
    "default_product_view": "needs_attention",
    "show_safe_products_dashboard": "no",
    "default_report_type": "demand_forecast_summary",
    "default_report_period": "last_30_days",
    "export_format": "csv",
    "include_filtered_rows_only": "yes",
    # Kept for compatibility with older templates/helpers.
    "default_forecast_range": "30",
    "store_logo": "",
    "data_date_format": "auto",
    "data_time_format": "auto",
    "payday_indicator_handling": "auto",
    "duplicate_handling": "remove_exact",
    "column_mapping_json": json.dumps(DEFAULT_COLUMN_MAPPING),
}

STORE_TYPE_OPTIONS = [
    "Sari-sari Store",
    "Mini Grocery",
    "Neighborhood Retail Shop",
    "Convenience Store",
]

FINAL_ROLE_OPTIONS = ["Owner", "Store Manager", "Operational Assistant"]
USER_ROLE_OPTIONS = FINAL_ROLE_OPTIONS.copy()
EMPLOYEE_ROLE_OPTIONS = ["Store Manager", "Operational Assistant"]

ROLE_ALIASES = {
    "Store Owner": "Owner",
    "System User": "Owner",
    "Owner": "Owner",

    "Manager": "Store Manager",
    "Store Manager": "Store Manager",
    "Store Supervisor": "Store Manager",
    "Supervisor": "Store Manager",

    "Inventory Staff": "Operational Assistant",
    "Store Assistant": "Operational Assistant",
    "Operations Assistant": "Operational Assistant",
    "Operational Assistant": "Operational Assistant",
    "Staff / Encoder": "Operational Assistant",
    "Staff": "Operational Assistant",
    "Encoder": "Operational Assistant",
    "Viewer": "Operational Assistant",
    "Cashier": "Operational Assistant",
}

ROLE_PAGE_ACCESS = {
    "dashboard": {"Owner", "Store Manager", "Operational Assistant"},
    "upload_data": {"Owner", "Store Manager"},
    "insights": {"Owner", "Store Manager"},
    "products": {"Owner", "Store Manager", "Operational Assistant"},
    "reports": {"Owner", "Store Manager"},
    "settings": {"Owner", "Store Manager", "Operational Assistant"},
}

CURRENCY_OPTIONS = [
    {"key": "PHP", "label": "PHP - Philippine Peso (₱)"},
    {"key": "USD", "label": "USD - US Dollar ($)"},
    {"key": "EUR", "label": "EUR - Euro (€)"},
    {"key": "GBP", "label": "GBP - British Pound (£)"},
    {"key": "JPY", "label": "JPY - Japanese Yen (¥)"},
    {"key": "CNY", "label": "CNY - Chinese Yuan (¥)"},
    {"key": "KRW", "label": "KRW - South Korean Won (₩)"},
    {"key": "SGD", "label": "SGD - Singapore Dollar (S$)"},
    {"key": "HKD", "label": "HKD - Hong Kong Dollar (HK$)"},
    {"key": "AUD", "label": "AUD - Australian Dollar (A$)"},
    {"key": "CAD", "label": "CAD - Canadian Dollar (C$)"},
    {"key": "NZD", "label": "NZD - New Zealand Dollar (NZ$)"},
    {"key": "MYR", "label": "MYR - Malaysian Ringgit (RM)"},
    {"key": "IDR", "label": "IDR - Indonesian Rupiah (Rp)"},
    {"key": "THB", "label": "THB - Thai Baht (฿)"},
    {"key": "VND", "label": "VND - Vietnamese Dong (₫)"},
    {"key": "INR", "label": "INR - Indian Rupee (₹)"},
    {"key": "AED", "label": "AED - UAE Dirham (د.إ)"},
    {"key": "SAR", "label": "SAR - Saudi Riyal (﷼)"},
    {"key": "CHF", "label": "CHF - Swiss Franc (CHF)"},
]

DEFAULT_PRODUCT_VIEW_OPTIONS = [
    {"key": "needs_attention", "label": "Needs Attention"},
    {"key": "all_products", "label": "All Products"},
]

YES_NO_OPTIONS = [
    {"key": "yes", "label": "Yes"},
    {"key": "no", "label": "No"},
]

USER_SETTINGS_SCHEMA_READY = False
USER_ACCOUNT_MEDIA_SCHEMA_READY = False
NOTIFICATION_SCHEMA_READY = False
MULTI_USER_SCHEMA_READY = False


# =========================================
# TEMPORARY IN-MEMORY STORAGE
# Replace with repository/database layer later
# =========================================
@dataclass
class AppState:
    selected_data: pd.DataFrame | None = None
    processed_data: pd.DataFrame | None = None
    selected_filename: str | None = None
    selected_file_size: int | None = None
    selected_file_type: str | None = None
    processed_filename: str | None = None
    upload_message: str | None = None
    upload_message_type: str | None = None
    selected_at: datetime | None = None
    processed_at: datetime | None = None
    last_upload_mode: str = "new"



def allowed_image_file(filename: str | None) -> bool:
    if not filename or "." not in filename:
        return False
    extension = filename.rsplit(".", 1)[1].lower()
    return extension in ALLOWED_IMAGE_EXTENSIONS


def ensure_static_upload_dir(subfolder: str) -> str:
    upload_dir = os.path.join(app.root_path, "static", "uploads", subfolder)
    os.makedirs(upload_dir, exist_ok=True)
    return upload_dir


def save_uploaded_image(file_storage, subfolder: str, prefix: str) -> str | None:
    """Save an optional uploaded image under static/uploads and return a browser-safe relative path."""
    if not file_storage or not getattr(file_storage, "filename", ""):
        return None
    if not allowed_image_file(file_storage.filename):
        raise ValueError("Please upload a PNG, JPG, JPEG, GIF, or WEBP image.")

    original_name = secure_filename(file_storage.filename)
    extension = original_name.rsplit(".", 1)[1].lower()
    safe_prefix = secure_filename(prefix or "upload") or "upload"
    filename = f"{safe_prefix}_{uuid.uuid4().hex[:12]}.{extension}"
    upload_dir = ensure_static_upload_dir(subfolder)
    file_storage.save(os.path.join(upload_dir, filename))
    return f"uploads/{subfolder}/{filename}"


def _quote_mysql_identifier(identifier: str) -> str:
    """Safely quote internal table/column names for MySQL ALTER statements."""
    cleaned = str(identifier or "")
    if not re.fullmatch(r"[A-Za-z0-9_]+", cleaned):
        raise ValueError(f"Unsafe MySQL identifier: {identifier!r}")
    return f"`{cleaned}`"


def _column_exists(cursor, table_name: str, column_name: str) -> bool:
    """Return True when the current database already has the given column."""
    cursor.execute(
        """
        SELECT COUNT(*) AS column_count
        FROM information_schema.COLUMNS
        WHERE TABLE_SCHEMA = DATABASE()
          AND TABLE_NAME = %s
          AND COLUMN_NAME = %s
        """,
        (table_name, column_name),
    )
    row = cursor.fetchone()
    if isinstance(row, dict):
        return int(row.get("column_count", 0) or 0) > 0
    return bool(row and int(row[0] or 0) > 0)


def add_column_if_missing(cursor, table_name: str, column_name: str, column_definition: str, after_column: str | None = None) -> None:
    """MySQL-compatible replacement for MariaDB's ADD COLUMN IF NOT EXISTS."""
    if _column_exists(cursor, table_name, column_name):
        return

    statement = (
        f"ALTER TABLE {_quote_mysql_identifier(table_name)} "
        f"ADD COLUMN {_quote_mysql_identifier(column_name)} {column_definition}"
    )
    if after_column:
        statement += f" AFTER {_quote_mysql_identifier(after_column)}"
    cursor.execute(statement)


def ensure_user_account_media_columns() -> None:
    """Add lightweight account media fields safely for existing MariaDB/XAMPP installs."""
    global USER_ACCOUNT_MEDIA_SCHEMA_READY
    if USER_ACCOUNT_MEDIA_SCHEMA_READY:
        return

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        add_column_if_missing(cursor, "users", "profile_image", "VARCHAR(255) NULL", after_column="position")
        conn.commit()
        USER_ACCOUNT_MEDIA_SCHEMA_READY = True
    finally:
        cursor.close()
        conn.close()


def ensure_notifications_table() -> None:
    """Create and lightly upgrade the user-facing notification table."""
    global NOTIFICATION_SCHEMA_READY
    if NOTIFICATION_SCHEMA_READY:
        return

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS notifications (
                notification_id INT(11) NOT NULL AUTO_INCREMENT,
                user_id INT(11) NOT NULL,
                store_id INT(11) NULL,
                title VARCHAR(120) NOT NULL,
                message VARCHAR(255) NOT NULL,
                event_type VARCHAR(50) NULL DEFAULT 'system',
                target_role VARCHAR(40) NULL,
                is_read TINYINT(1) NOT NULL DEFAULT 0,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (notification_id),
                INDEX idx_notifications_user_created (user_id, created_at),
                INDEX idx_notifications_user_read (user_id, is_read),
                INDEX idx_notifications_store_created (store_id, created_at)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """
        )
        add_column_if_missing(cursor, "notifications", "store_id", "INT(11) NULL", after_column="user_id")
        add_column_if_missing(cursor, "notifications", "target_role", "VARCHAR(40) NULL", after_column="event_type")
        conn.commit()
        NOTIFICATION_SCHEMA_READY = True
    finally:
        cursor.close()
        conn.close()


def normalize_role(role: Any) -> str:
    cleaned = str(role or "").strip()
    if not cleaned:
        return "Owner"
    normalized = ROLE_ALIASES.get(cleaned, cleaned)
    return normalized if normalized in FINAL_ROLE_OPTIONS else "Operational Assistant"


def get_session_role() -> str:
    return normalize_role(session.get("user_role") or session.get("user_position") or "Owner")


def is_owner_user() -> bool:
    return get_session_role() == "Owner"


def role_can_access(page_key: str, role: str | None = None) -> bool:
    normalized = normalize_role(role or get_session_role())
    if normalized == "Owner":
        return True
    return normalized in ROLE_PAGE_ACCESS.get(page_key, set())


def get_visible_nav_keys() -> set[str]:
    role = get_session_role()
    return {page_key for page_key in ROLE_PAGE_ACCESS if role_can_access(page_key, role)}


def get_settings_allowed_sections(role: str | None = None) -> set[str]:
    normalized = normalize_role(role or get_session_role())
    if normalized == "Owner":
        return {
            "profile",
            "store",
            "upload",
            "reports",
            "security",
            "data_format",
            "employees",
            "activity",
            "data_management",
        }
    return {"profile", "security"}


def normalize_settings_section_key(section_key: str | None) -> str:
    cleaned = (section_key or "profile").strip().replace("-", "_") or "profile"
    aliases = {
        "profile_account": "profile",
        "store_information": "store",
        "team_access": "employees",
        "employee_management": "employees",
        "sales_file_setup": "data_format",
        "reports_display": "reports_display",
        "data_management": "data_management",
        "activity_logs": "activity",
        "password_security": "security",
    }
    return aliases.get(cleaned, cleaned)


def can_save_settings_section(section_key: str) -> bool:
    normalized_key = normalize_settings_section_key(section_key)
    allowed = get_settings_allowed_sections()
    if normalized_key == "reports_display":
        return bool({"upload", "reports"} & allowed)
    if normalized_key == "data_management":
        return "data_management" in allowed
    return normalized_key in allowed


def role_required(page_key: str):
    def decorator(view_func):
        @wraps(view_func)
        def wrapped_view(*args, **kwargs):
            if not role_can_access(page_key):
                add_activity_log("Access denied", page_key.replace("_", " ").title(), "Blocked")
                return render_template(
                    "access_denied.html",
                    blocked_page=page_key.replace("_", " ").title(),
                    user_role=get_session_role(),
                ), 403
            return view_func(*args, **kwargs)
        return wrapped_view
    return decorator


def ensure_multi_user_schema() -> None:
    """Create simple workspace, role, and activity-log support safely."""
    global MULTI_USER_SCHEMA_READY
    if MULTI_USER_SCHEMA_READY:
        return

    ensure_user_account_media_columns()
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS stores (
                store_id INT(11) NOT NULL AUTO_INCREMENT,
                owner_user_id INT(11) NOT NULL,
                store_name VARCHAR(255) NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                PRIMARY KEY (store_id),
                INDEX idx_stores_owner (owner_user_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """
        )
        add_column_if_missing(cursor, "users", "role", "VARCHAR(50) NULL", after_column="position")
        add_column_if_missing(cursor, "users", "store_id", "INT(11) NULL", after_column="role")
        add_column_if_missing(cursor, "users", "created_by", "INT(11) NULL", after_column="store_id")
        add_column_if_missing(cursor, "users", "account_status", "VARCHAR(30) NULL", after_column="created_by")
        add_column_if_missing(cursor, "users", "last_login_at", "DATETIME NULL", after_column="is_active")
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS activity_logs (
                log_id INT(11) NOT NULL AUTO_INCREMENT,
                store_id INT(11) NULL,
                user_id INT(11) NULL,
                user_name VARCHAR(255) NULL,
                action VARCHAR(120) NOT NULL,
                module VARCHAR(80) NOT NULL,
                status VARCHAR(40) NOT NULL DEFAULT 'Success',
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (log_id),
                INDEX idx_activity_store_created (store_id, created_at),
                INDEX idx_activity_user_created (user_id, created_at)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """
        )
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS store_memberships (
                membership_id INT(11) NOT NULL AUTO_INCREMENT,
                store_id INT(11) NOT NULL,
                user_id INT(11) NOT NULL,
                role VARCHAR(50) NOT NULL,
                account_status VARCHAR(30) NOT NULL DEFAULT 'active',
                is_active TINYINT(1) NOT NULL DEFAULT 1,
                created_by INT(11) NULL,
                joined_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                removed_at DATETIME NULL,
                reactivated_at DATETIME NULL,
                last_login_at DATETIME NULL,
                PRIMARY KEY (membership_id),
                UNIQUE KEY uq_store_user_membership (store_id, user_id),
                INDEX idx_membership_store_status (store_id, account_status),
                INDEX idx_membership_user (user_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """
        )
        add_column_if_missing(cursor, "activity_logs", "membership_id", "INT(11) NULL", after_column="user_id")
        add_column_if_missing(cursor, "activity_logs", "actor_role", "VARCHAR(50) NULL", after_column="user_name")
        cursor.execute(
            """
            UPDATE users
            SET role = CASE
                    WHEN role IS NULL OR role = '' THEN
                        CASE
                            WHEN position IS NULL OR position = '' OR position IN ('Owner', 'Store Owner', 'System User') THEN 'Owner'
                            WHEN position IN ('Manager', 'Store Manager', 'Store Supervisor', 'Supervisor') THEN 'Store Manager'
                            WHEN position IN ('Inventory Staff', 'Store Assistant', 'Operations Assistant', 'Operational Assistant', 'Staff / Encoder', 'Staff', 'Encoder', 'Viewer', 'Cashier') THEN 'Operational Assistant'
                            ELSE 'Operational Assistant'
                        END
                    WHEN role IN ('Owner', 'Store Owner', 'System User') THEN 'Owner'
                    WHEN role IN ('Manager', 'Store Manager', 'Store Supervisor', 'Supervisor') THEN 'Store Manager'
                    WHEN role IN ('Inventory Staff', 'Store Assistant', 'Operations Assistant', 'Operational Assistant', 'Staff / Encoder', 'Staff', 'Encoder', 'Viewer', 'Cashier') THEN 'Operational Assistant'
                    ELSE 'Operational Assistant'
                END,
                position = CASE
                    WHEN position IS NULL OR position = '' OR position = 'System User' THEN 'Owner'
                    WHEN position IN ('Owner', 'Store Owner') THEN 'Owner'
                    WHEN position IN ('Manager', 'Store Manager', 'Store Supervisor', 'Supervisor') THEN 'Store Manager'
                    WHEN position IN ('Inventory Staff', 'Store Assistant', 'Operations Assistant', 'Operational Assistant', 'Staff / Encoder', 'Staff', 'Encoder', 'Viewer', 'Cashier') THEN 'Operational Assistant'
                    ELSE 'Operational Assistant'
                END
            """
        )
        cursor.execute(
            """
            UPDATE users
            SET account_status = CASE
                    WHEN COALESCE(account_status, '') IN ('active', 'deactivated', 'removed') THEN account_status
                    WHEN COALESCE(is_active, 1) = 1 THEN 'active'
                    ELSE 'deactivated'
                END
            """
        )
        cursor.execute("UPDATE users SET is_active = 0 WHERE account_status IN ('deactivated', 'removed')")
        cursor.execute("UPDATE users SET is_active = 1 WHERE account_status = 'active' AND COALESCE(is_active, 1) <> 1")
        cursor.execute(
            """
            INSERT IGNORE INTO store_memberships (store_id, user_id, role, account_status, is_active, created_by, joined_at, last_login_at)
            SELECT store_id,
                   user_id,
                   COALESCE(NULLIF(role, ''), NULLIF(position, ''), 'Owner'),
                   COALESCE(NULLIF(account_status, ''), CASE WHEN COALESCE(is_active, 1) = 1 THEN 'active' ELSE 'deactivated' END),
                   COALESCE(is_active, 1),
                   created_by,
                   COALESCE(created_at, NOW()),
                   last_login_at
            FROM users
            WHERE store_id IS NOT NULL
            """
        )
        cursor.execute(
            """
            UPDATE store_memberships sm
            JOIN users u ON u.user_id = sm.user_id
            SET sm.role = CASE
                    WHEN sm.role IN ('Owner', 'Store Owner', 'System User') THEN 'Owner'
                    WHEN sm.role IN ('Manager', 'Store Manager', 'Store Supervisor', 'Supervisor') THEN 'Store Manager'
                    WHEN sm.role IN ('Inventory Staff', 'Store Assistant', 'Operations Assistant', 'Operational Assistant', 'Staff / Encoder', 'Staff', 'Encoder', 'Viewer', 'Cashier') THEN 'Operational Assistant'
                    WHEN sm.role IS NULL OR sm.role = '' THEN COALESCE(NULLIF(u.role, ''), NULLIF(u.position, ''), 'Owner')
                    ELSE 'Operational Assistant'
                END
            """
        )
        cursor.execute(
            """
            UPDATE users u
            JOIN stores s ON s.owner_user_id = u.user_id
            SET u.role = 'Owner',
                u.position = 'Owner',
                u.account_status = 'active',
                u.is_active = 1
            """
        )
        cursor.execute(
            """
            UPDATE store_memberships sm
            JOIN stores s
              ON s.store_id = sm.store_id
             AND s.owner_user_id = sm.user_id
            SET sm.role = 'Owner',
                sm.account_status = 'active',
                sm.is_active = 1
            """
        )
        cursor.execute(
            """
            UPDATE activity_logs l
            JOIN store_memberships sm ON sm.store_id = l.store_id AND sm.user_id = l.user_id
            SET l.membership_id = sm.membership_id,
                l.actor_role = COALESCE(l.actor_role, sm.role)
            WHERE l.membership_id IS NULL
            """
        )
        conn.commit()
        MULTI_USER_SCHEMA_READY = True
    finally:
        cursor.close()
        conn.close()


def ensure_user_workspace(user_id: int | None) -> int | None:
    """Assign an existing or new store workspace to a user without locking out old accounts."""
    if not user_id:
        return None
    try:
        ensure_multi_user_schema()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        try:
            cursor.execute("SELECT user_id, full_name, role, position, store_id, created_by, account_status, is_active FROM users WHERE user_id = %s LIMIT 1", (user_id,))
            user = cursor.fetchone()
            if not user:
                return None
            role = normalize_role(user.get("role") or user.get("position") or "Owner")
            if user.get("store_id"):
                store_id = int(user["store_id"])
                cursor2 = conn.cursor()
                try:
                    cursor2.execute(
                        """
                        INSERT IGNORE INTO store_memberships (store_id, user_id, role, account_status, is_active, created_by, joined_at, last_login_at)
                        SELECT store_id, user_id,
                               COALESCE(NULLIF(role, ''), NULLIF(position, ''), 'Owner'),
                               COALESCE(NULLIF(account_status, ''), CASE WHEN COALESCE(is_active, 1) = 1 THEN 'active' ELSE 'deactivated' END),
                               COALESCE(is_active, 1), created_by, COALESCE(created_at, NOW()), last_login_at
                        FROM users
                        WHERE user_id = %s AND store_id IS NOT NULL
                        """,
                        (user_id,),
                    )
                    conn.commit()
                finally:
                    cursor2.close()
                return store_id

            store_name = None
            try:
                prefs = get_user_settings_from_db(user_id)
                store_name = prefs.get("store_name")
            except Exception:
                store_name = None
            store_name = store_name or f"{user.get('full_name') or 'Owner'}'s Store"
            cursor2 = conn.cursor()
            try:
                cursor2.execute(
                    """
                    INSERT INTO stores (owner_user_id, store_name)
                    VALUES (%s, %s)
                    """,
                    (user_id, store_name),
                )
                store_id = cursor2.lastrowid
                cursor2.execute(
                    """
                    UPDATE users
                    SET store_id = %s,
                        role = %s,
                        position = CASE WHEN position IS NULL OR position = '' THEN %s ELSE position END,
                        account_status = COALESCE(NULLIF(account_status, ''), 'active'),
                        is_active = 1
                    WHERE user_id = %s
                    """,
                    (store_id, role, role, user_id),
                )
                cursor2.execute(
                    """
                    INSERT IGNORE INTO store_memberships (store_id, user_id, role, account_status, is_active, created_by, joined_at)
                    VALUES (%s, %s, %s, 'active', 1, %s, NOW())
                    """,
                    (store_id, user_id, role, user_id),
                )
                conn.commit()
                return int(store_id)
            finally:
                cursor2.close()
        finally:
            cursor.close()
            conn.close()
    except Exception:
        return None


def refresh_session_user(user_id: int | None = None) -> None:
    if not user_id:
        user_id = get_current_user_id()
    if not user_id:
        return
    user = get_user_by_id(user_id)
    if not user:
        return
    session["user_name"] = user.get("full_name") or session.get("user_name", "User")
    session["user_email"] = user.get("email") or session.get("user_email", "")
    preferred_store_id = session.get("user_store_id") or user.get("store_id")
    membership = get_active_membership_for_user(user_id, preferred_store_id)
    if not membership and normalize_role(user.get("role") or user.get("position")) == "Owner":
        ensured_store_id = ensure_user_workspace(user_id)
        membership = get_active_membership_for_user(user_id, ensured_store_id)
    if membership:
        role = normalize_role(membership.get("role"))
        session["user_role"] = role
        session["user_position"] = role
        session["user_store_id"] = membership.get("store_id")
        session["user_membership_id"] = membership.get("membership_id")
        sync_user_legacy_access_from_membership(user_id, membership)
    else:
        role = normalize_role(user.get("role") or user.get("position"))
        session["user_role"] = role
        session["user_position"] = role
        session["user_store_id"] = user.get("store_id") or ensure_user_workspace(user_id)
        session["user_membership_id"] = get_current_membership_id(user_id, session.get("user_store_id"))
    session["user_profile_image"] = user.get("profile_image") or ""


def update_last_login(user_id: int | None) -> None:
    if not user_id:
        return
    try:
        ensure_multi_user_schema()
        store_id = get_current_store_id()
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("UPDATE users SET last_login_at = NOW() WHERE user_id = %s", (user_id,))
            if store_id:
                cursor.execute(
                    """
                    UPDATE store_memberships
                    SET last_login_at = NOW()
                    WHERE user_id = %s AND store_id = %s
                    """,
                    (user_id, store_id),
                )
            conn.commit()
        finally:
            cursor.close()
            conn.close()
    except Exception:
        return


def get_current_store_id() -> int | None:
    store_id = session.get("user_store_id")
    try:
        return int(store_id) if store_id is not None else ensure_user_workspace(get_current_user_id())
    except (TypeError, ValueError):
        return ensure_user_workspace(get_current_user_id())


def get_membership_for_store_user(store_id: int | None, user_id: int | None) -> dict[str, Any] | None:
    if not store_id or not user_id:
        return None
    try:
        ensure_multi_user_schema()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        try:
            cursor.execute(
                """
                SELECT sm.membership_id, sm.store_id, sm.user_id, sm.role,
                       sm.account_status, sm.is_active, sm.created_by, sm.joined_at,
                       sm.removed_at, sm.reactivated_at, sm.last_login_at,
                       u.full_name, u.email, u.username, u.profile_image
                FROM store_memberships sm
                JOIN users u ON u.user_id = sm.user_id
                WHERE sm.store_id = %s AND sm.user_id = %s
                LIMIT 1
                """,
                (store_id, user_id),
            )
            return cursor.fetchone()
        finally:
            cursor.close()
            conn.close()
    except Exception:
        return None


def get_active_membership_for_user(user_id: int | None, preferred_store_id: int | None = None) -> dict[str, Any] | None:
    if not user_id:
        return None
    try:
        ensure_multi_user_schema()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        try:
            preferred = int(preferred_store_id or 0)
            cursor.execute(
                """
                SELECT sm.membership_id, sm.store_id, sm.user_id, sm.role,
                       sm.account_status, sm.is_active, sm.created_by, sm.joined_at,
                       sm.removed_at, sm.reactivated_at, sm.last_login_at,
                       u.full_name, u.email, u.username, u.profile_image
                FROM store_memberships sm
                JOIN users u ON u.user_id = sm.user_id
                WHERE sm.user_id = %s
                  AND sm.account_status = 'active'
                  AND sm.is_active = 1
                ORDER BY CASE WHEN sm.store_id = %s THEN 0 ELSE 1 END,
                         sm.reactivated_at DESC, sm.joined_at DESC, sm.membership_id DESC
                LIMIT 1
                """,
                (user_id, preferred),
            )
            return cursor.fetchone()
        finally:
            cursor.close()
            conn.close()
    except Exception:
        return None


def sync_user_legacy_access_from_membership(user_id: int | None, membership: dict[str, Any] | None) -> None:
    if not user_id or not membership:
        return
    try:
        ensure_multi_user_schema()
        role = normalize_role(membership.get("role"))
        status_key = normalize_employee_status(membership.get("account_status"), membership.get("is_active", 1))
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute(
                """
                UPDATE users
                SET store_id = %s,
                    role = %s,
                    position = %s,
                    account_status = %s,
                    is_active = %s
                WHERE user_id = %s
                """,
                (
                    membership.get("store_id"),
                    role,
                    role,
                    status_key,
                    1 if status_key == "active" else 0,
                    user_id,
                ),
            )
            conn.commit()
        finally:
            cursor.close()
            conn.close()
    except Exception:
        return


def sync_user_after_membership_change(user_id: int | None, changed_store_id: int | None = None) -> None:
    if not user_id:
        return
    active_membership = get_active_membership_for_user(user_id, changed_store_id)
    if active_membership:
        sync_user_legacy_access_from_membership(user_id, active_membership)
        return
    try:
        ensure_multi_user_schema()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        try:
            cursor.execute(
                """
                SELECT sm.store_id, sm.role, sm.account_status, sm.is_active
                FROM store_memberships sm
                WHERE sm.user_id = %s
                ORDER BY sm.joined_at DESC, sm.membership_id DESC
                LIMIT 1
                """,
                (user_id,),
            )
            membership = cursor.fetchone()
            if membership:
                sync_user_legacy_access_from_membership(user_id, membership)
        finally:
            cursor.close()
            conn.close()
    except Exception:
        return


def get_current_membership_id(user_id: int | None = None, store_id: int | None = None) -> int | None:
    user_id = user_id or get_current_user_id()
    store_id = store_id or get_current_store_id()
    membership = get_membership_for_store_user(store_id, user_id)
    try:
        return int(membership["membership_id"]) if membership and membership.get("membership_id") is not None else None
    except Exception:
        return None


def get_store_id_for_user(user_id: int | None) -> int | None:
    if not user_id:
        return None
    if user_id == get_current_user_id():
        return get_current_store_id()
    membership = get_active_membership_for_user(user_id)
    if membership and membership.get("store_id"):
        try:
            return int(membership["store_id"])
        except Exception:
            pass
    try:
        ensure_multi_user_schema()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        try:
            cursor.execute("SELECT store_id FROM users WHERE user_id = %s LIMIT 1", (user_id,))
            row = cursor.fetchone()
            if row and row.get("store_id"):
                return int(row["store_id"])
        finally:
            cursor.close()
            conn.close()
    except Exception:
        return None
    return None


def get_store_user_ids(user_id: int | None = None) -> list[int]:
    """Return active user ids in the same store workspace for store-aware data access."""
    target_user_id = user_id or get_current_user_id()
    if not target_user_id:
        return []
    store_id = get_store_id_for_user(target_user_id)
    if not store_id:
        return [int(target_user_id)]
    try:
        ensure_multi_user_schema()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        try:
            cursor.execute(
                """
                SELECT user_id
                FROM store_memberships
                WHERE store_id = %s
                  AND account_status = 'active'
                  AND is_active = 1
                ORDER BY user_id
                """,
                (store_id,),
            )
            rows = cursor.fetchall() or []
        finally:
            cursor.close()
            conn.close()
        ids = [int(row["user_id"]) for row in rows if row.get("user_id") is not None]
        return ids or [int(target_user_id)]
    except Exception:
        return [int(target_user_id)]


def make_in_clause(values: list[int]) -> tuple[str, tuple[int, ...]]:
    cleaned = [int(value) for value in values if value is not None]
    if not cleaned:
        cleaned = [int(get_current_user_id() or 0)]
    placeholders = ", ".join(["%s"] * len(cleaned))
    return placeholders, tuple(cleaned)


def add_activity_log(action: str, module: str, status: str = "Success", user_id: int | None = None, store_id: int | None = None) -> None:
    try:
        ensure_multi_user_schema()
        if user_id is None:
            user_id = get_current_user_id()
        if store_id is None:
            store_id = get_current_store_id()
        membership = get_membership_for_store_user(store_id, user_id) if user_id and store_id else None
        membership_id = membership.get("membership_id") if membership else None
        actor_role = normalize_role(membership.get("role")) if membership else (get_session_role() if user_id == get_current_user_id() else None)
        user_name = session.get("user_name", "User") if user_id == get_current_user_id() else None
        if user_id and not user_name:
            user = get_user_by_id(user_id)
            user_name = (user or {}).get("full_name")
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute(
                """
                INSERT INTO activity_logs (store_id, user_id, membership_id, user_name, actor_role, action, module, status)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (store_id, user_id, membership_id, (user_name or "User")[:255], (actor_role or "")[:50], action[:120], module[:80], status[:40]),
            )
            conn.commit()
        finally:
            cursor.close()
            conn.close()
    except Exception:
        return


def get_activity_display_label(value: Any, label_type: str = "action") -> str:
    text = str(value or "").strip()
    key = text.casefold()
    action_labels = {
        "generate results": "Results generated",
        "generated results": "Results generated",
        "file upload": "File uploaded",
        "file uploaded": "File uploaded",
        "login": "Login",
        "logout": "Logout",
        "authentication": "Account access",
        "settings update": "Settings updated",
        "data format settings changed": "Sales file setup changed",
        "employee added": "Employee added",
        "employee role updated": "Employee role updated",
        "employee deactivated": "Employee deactivated",
        "employee removed": "Employee removed",
        "employee rehired": "Employee rehired",
        "report exported": "Report exported",
        "report export failed": "Report export failed",
    }
    module_labels = {
        "authentication": "Account access",
        "upload sales data": "Upload Sales Data",
        "upload": "Upload Sales Data",
        "settings": "Settings",
        "employee": "Team & Access",
        "activity logs": "Activity Logs",
        "data management": "Data Management",
        "reports": "Reports",
    }
    status_labels = {
        "success": "Success",
        "failed": "Unsuccessful",
        "blocked": "Blocked",
        "warning": "Warning",
        "pending": "Pending",
        "error": "Unsuccessful",
    }
    if label_type == "module":
        return module_labels.get(key, text or "System")
    if label_type == "status":
        return status_labels.get(key, text or "Success")
    return action_labels.get(key, text or "Activity")


def get_activity_filter_label_options(filter_options: dict[str, list[str]]) -> dict[str, dict[str, str]]:
    return {
        "modules": {value: get_activity_display_label(value, "module") for value in filter_options.get("modules", [])},
        "actions": {value: get_activity_display_label(value, "action") for value in filter_options.get("actions", [])},
        "statuses": {value: get_activity_display_label(value, "status") for value in filter_options.get("statuses", [])},
    }


def get_user_display_name(row: dict[str, Any] | None, fallback: str = "User") -> str:
    if not row:
        return fallback
    return (
        row.get("full_name")
        or row.get("display_name")
        or row.get("email")
        or row.get("username")
        or fallback
    )


def get_activity_user_groups_for_current_store() -> list[dict[str, Any]]:
    store_id = get_current_store_id()
    if not store_id:
        return [
            {"key": "active", "label": "Active", "users": []},
            {"key": "deactivated", "label": "Deactivated", "users": []},
            {"key": "removed", "label": "Removed", "users": []},
        ]
    try:
        ensure_multi_user_schema()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        try:
            cursor.execute(
                """
                SELECT sm.user_id, u.full_name, u.email, u.username, sm.role, sm.account_status, sm.is_active
                FROM store_memberships sm
                JOIN users u ON u.user_id = sm.user_id
                WHERE sm.store_id = %s
                ORDER BY u.full_name ASC, u.email ASC
                """,
                (store_id,),
            )
            rows = cursor.fetchall() or []
        finally:
            cursor.close()
            conn.close()
    except Exception:
        rows = []

    grouped = {"active": [], "deactivated": [], "removed": []}
    for row in rows:
        status_key = normalize_employee_status(row.get("account_status"), row.get("is_active", 1))
        if status_key not in grouped:
            status_key = "active"
        name = get_user_display_name(row, "User")
        grouped[status_key].append({
            "user_id": row.get("user_id"),
            "name": name,
            "role": normalize_role(row.get("role")),
        })

    for users in grouped.values():
        users.sort(key=lambda item: str(item.get("name") or "").casefold())

    return [
        {"key": "active", "label": "Active", "users": grouped["active"]},
        {"key": "deactivated", "label": "Deactivated", "users": grouped["deactivated"]},
        {"key": "removed", "label": "Removed", "users": grouped["removed"]},
    ]


def get_activity_selected_user_label(employee_id: str, user_name: str = "") -> str:
    cleaned_employee_id = str(employee_id or "").strip()
    if cleaned_employee_id.isdigit():
        try:
            membership = get_membership_for_store_user(get_current_store_id(), int(cleaned_employee_id))
            if membership:
                return get_user_display_name(membership, "Selected user")
        except Exception:
            pass
    return (user_name or "").strip() or "All users"


def get_activity_logs_for_current_store(limit: int = 40) -> list[dict[str, Any]]:
    store_id = get_current_store_id()
    if not store_id:
        return []
    try:
        ensure_multi_user_schema()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        try:
            cursor.execute(
                """
                SELECT l.user_id,
                       COALESCE(NULLIF(u.full_name, ''), NULLIF(l.user_name, ''), 'User') AS user_name,
                       COALESCE(NULLIF(l.actor_role, ''), NULLIF(sm.role, ''), NULLIF(u.role, ''), NULLIF(u.position, ''), '') AS user_role,
                       l.action, l.module, l.status, l.created_at
                FROM activity_logs l
                LEFT JOIN users u ON u.user_id = l.user_id
                LEFT JOIN store_memberships sm
                       ON sm.store_id = l.store_id
                      AND sm.user_id = l.user_id
                      AND (l.membership_id IS NULL OR sm.membership_id = l.membership_id)
                WHERE l.store_id = %s
                ORDER BY l.created_at DESC, l.log_id DESC
                LIMIT %s
                """,
                (store_id, limit),
            )
            rows = cursor.fetchall() or []
        finally:
            cursor.close()
            conn.close()
        logs = []
        for row in rows:
            created_at = row.get("created_at")
            try:
                created_label = created_at.strftime("%b %d, %Y %I:%M %p") if created_at else ""
            except Exception:
                created_label = str(created_at or "")
            logs.append({
                "user_id": row.get("user_id"),
                "user_name": row.get("user_name") or "User",
                "user_role": normalize_role(row.get("user_role")),
                "action": row.get("action") or "Activity",
                "module": row.get("module") or "System",
                "status": row.get("status") or "Success",
                "action_display": get_activity_display_label(row.get("action"), "action"),
                "module_display": get_activity_display_label(row.get("module"), "module"),
                "status_display": get_activity_display_label(row.get("status"), "status"),
                "created_label": created_label,
            })
        return logs
    except Exception:
        return []


def get_recent_team_activity_for_dashboard(limit: int = 5) -> list[dict[str, Any]]:
    """Return a compact Owner dashboard preview of recent employee/team activity."""
    store_id = get_current_store_id()
    if not store_id:
        return []

    owner_user_id = None
    try:
        ensure_multi_user_schema()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        try:
            cursor.execute(
                """
                SELECT s.owner_user_id
                FROM stores s
                WHERE s.store_id = %s
                LIMIT 1
                """,
                (store_id,),
            )
            owner_row = cursor.fetchone() or {}
            owner_user_id = owner_row.get("owner_user_id")
        finally:
            cursor.close()
            conn.close()
    except Exception:
        owner_user_id = get_current_user_id()

    try:
        logs = get_activity_logs_for_current_store(limit=40)
    except Exception:
        return []

    employee_logs: list[dict[str, Any]] = []
    for log in logs:
        try:
            if owner_user_id and int(log.get("user_id") or 0) == int(owner_user_id):
                continue
        except Exception:
            pass

        if normalize_role(log.get("user_role")) == "Owner":
            continue

        employee_logs.append(log)
        if len(employee_logs) >= limit:
            break

    return employee_logs


def store_has_team_members() -> bool:
    """Return True when the current store has at least one non-owner membership."""
    store_id = get_current_store_id()
    if not store_id:
        return False
    try:
        ensure_multi_user_schema()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        try:
            cursor.execute(
                """
                SELECT COUNT(*) AS total
                FROM store_memberships sm
                JOIN stores s ON s.store_id = sm.store_id
                WHERE sm.store_id = %s
                  AND sm.user_id <> s.owner_user_id
                  AND COALESCE(sm.role, '') NOT IN ('Owner', 'Store Owner', 'System User')
                """,
                (store_id,),
            )
            row = cursor.fetchone() or {}
            return int(row.get("total") or 0) > 0
        finally:
            cursor.close()
            conn.close()
    except Exception:
        return False


def normalize_employee_status(status: Any, is_active: Any = None) -> str:
    cleaned = str(status or "").strip().lower()
    if cleaned in {"active", "deactivated", "removed"}:
        return cleaned
    try:
        return "active" if int(is_active) == 1 else "deactivated"
    except Exception:
        return "active"


def get_employee_status_options() -> list[dict[str, str]]:
    return [
        {"key": "active", "label": "Active"},
        {"key": "deactivated", "label": "Deactivated"},
        {"key": "removed", "label": "Removed"},
    ]


def get_full_activity_logs_for_current_store(filters: dict[str, str] | None = None, limit: int = 250) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    store_id = get_current_store_id()
    base_statuses = ["Success", "Pending", "Failed", "Warning"]
    empty_options = {"users": [], "user_groups": get_activity_user_groups_for_current_store(), "roles": [], "modules": [], "actions": [], "statuses": base_statuses}
    if not store_id:
        return [], empty_options

    filters = filters or {}
    where = ["l.store_id = %s"]
    params: list[Any] = [store_id]
    employee_id = str(filters.get("employee_id") or "").strip()

    if employee_id.isdigit():
        where.append("l.user_id = %s")
        params.append(int(employee_id))
    else:
        user_value = (filters.get("user") or "").strip()
        if user_value:
            where.append("COALESCE(NULLIF(u.full_name, ''), NULLIF(l.user_name, ''), 'User') = %s")
            params.append(user_value)

    for key, sql in {
        "role": "COALESCE(NULLIF(l.actor_role, ''), NULLIF(sm.role, ''), NULLIF(u.role, ''), NULLIF(u.position, ''), '') = %s",
        "module": "l.module = %s",
        "action": "l.action = %s",
        "status": "l.status = %s",
    }.items():
        value = (filters.get(key) or "").strip()
        if value:
            where.append(sql)
            params.append(value)

    date_from = (filters.get("date_from") or "").strip()
    date_to = (filters.get("date_to") or "").strip()
    if date_from:
        where.append("DATE(l.created_at) >= %s")
        params.append(date_from)
    if date_to:
        where.append("DATE(l.created_at) <= %s")
        params.append(date_to)

    try:
        ensure_multi_user_schema()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        join_sql = """
                LEFT JOIN users u ON u.user_id = l.user_id
                LEFT JOIN store_memberships sm
                       ON sm.store_id = l.store_id
                      AND sm.user_id = l.user_id
                      AND (l.membership_id IS NULL OR sm.membership_id = l.membership_id)
        """
        try:
            where_sql = " AND ".join(where)
            cursor.execute(
                f"""
                SELECT l.user_id,
                       COALESCE(NULLIF(u.full_name, ''), NULLIF(l.user_name, ''), 'User') AS user_name,
                       COALESCE(NULLIF(l.actor_role, ''), NULLIF(sm.role, ''), NULLIF(u.role, ''), NULLIF(u.position, ''), '') AS user_role,
                       l.action, l.module, l.status, l.created_at
                FROM activity_logs l
                {join_sql}
                WHERE {where_sql}
                ORDER BY l.created_at DESC, l.log_id DESC
                LIMIT %s
                """,
                (*params, limit),
            )
            rows = cursor.fetchall() or []
            cursor.execute(
                f"""
                SELECT DISTINCT
                       COALESCE(NULLIF(u.full_name, ''), NULLIF(l.user_name, ''), 'User') AS user_name,
                       COALESCE(NULLIF(l.actor_role, ''), NULLIF(sm.role, ''), NULLIF(u.role, ''), NULLIF(u.position, ''), '') AS user_role,
                       l.action, l.module, l.status
                FROM activity_logs l
                {join_sql}
                WHERE l.store_id = %s
                """,
                (store_id,),
            )
            option_rows = cursor.fetchall() or []
        finally:
            cursor.close()
            conn.close()

        logs = []
        for row in rows:
            created_at = row.get("created_at")
            try:
                created_label = created_at.strftime("%b %d, %Y %I:%M %p") if created_at else ""
            except Exception:
                created_label = str(created_at or "")
            logs.append({
                "user_name": row.get("user_name") or "User",
                "user_role": normalize_role(row.get("user_role")),
                "action": row.get("action") or "Activity",
                "module": row.get("module") or "System",
                "status": row.get("status") or "Success",
                "action_display": get_activity_display_label(row.get("action"), "action"),
                "module_display": get_activity_display_label(row.get("module"), "module"),
                "status_display": get_activity_display_label(row.get("status"), "status"),
                "created_label": created_label,
            })

        existing_statuses = {str(row.get("status") or "").strip() for row in option_rows if row.get("status")}
        status_order = {value: index for index, value in enumerate(base_statuses)}
        statuses = sorted(
            set(base_statuses) | existing_statuses,
            key=lambda value: (status_order.get(value, 99), str(value).casefold()),
        )

        options = {
            "users": sorted({row.get("user_name") for row in option_rows if row.get("user_name")}, key=str.casefold),
            "user_groups": get_activity_user_groups_for_current_store(),
            "roles": sorted({normalize_role(row.get("user_role")) for row in option_rows if row.get("user_role")}, key=str.casefold),
            "modules": sorted({row.get("module") for row in option_rows if row.get("module")}, key=str.casefold),
            "actions": sorted({row.get("action") for row in option_rows if row.get("action")}, key=str.casefold),
            "statuses": statuses,
        }
        return logs, options
    except Exception:
        return [], empty_options


def get_employees_for_current_store(include_removed: bool = False) -> list[dict[str, Any]]:
    store_id = get_current_store_id()
    if not store_id:
        return []
    try:
        ensure_multi_user_schema()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        try:
            cursor.execute(
                """
                SELECT sm.membership_id, sm.user_id, u.full_name, u.email, u.username,
                       sm.role, sm.is_active, sm.account_status,
                       COALESCE(sm.last_login_at, u.last_login_at) AS last_login_at
                FROM store_memberships sm
                JOIN users u ON u.user_id = sm.user_id
                WHERE sm.store_id = %s AND sm.user_id <> %s
                  AND (%s = 1 OR sm.account_status <> 'removed')
                ORDER BY FIELD(sm.account_status, 'active', 'deactivated', 'removed'), u.full_name ASC, u.email ASC
                """,
                (store_id, get_current_user_id() or 0, 1 if include_removed else 0),
            )
            rows = cursor.fetchall() or []
        finally:
            cursor.close()
            conn.close()
        employees = []
        for row in rows:
            last_login = row.get("last_login_at")
            try:
                last_login_label = last_login.strftime("%b %d, %Y %I:%M %p") if last_login else "Not yet"
            except Exception:
                last_login_label = str(last_login or "Not yet")
            status_key = normalize_employee_status(row.get("account_status"), row.get("is_active", 1))
            employees.append({
                "membership_id": row.get("membership_id"),
                "user_id": row.get("user_id"),
                "name": row.get("full_name") or "Employee",
                "email": row.get("email") or row.get("username") or "",
                "username": row.get("username") or "",
                "role": normalize_role(row.get("role")),
                "status": status_key.replace("_", " ").title(),
                "status_key": status_key,
                "is_active": status_key == "active",
                "is_removed": status_key == "removed",
                "last_login": last_login_label,
            })
        return employees
    except Exception:
        return []


def read_settings_notice() -> tuple[str | None, str | None]:
    message = session.pop("settings_flash_message", None)
    message_type = session.pop("settings_flash_type", None)
    return message, message_type


def set_settings_notice(message: str, message_type: str = "success") -> None:
    session["settings_flash_message"] = message
    session["settings_flash_type"] = message_type


def build_employee_temp_password(first_name: str, last_name: str, role: str, created_at: datetime | None = None) -> str:
    created_at = created_at or datetime.now()
    first_clean = re.sub(r"[^A-Za-z]", "", str(first_name or ""))
    last_clean = re.sub(r"[^A-Za-z]", "", str(last_name or ""))
    first_part = (first_clean[:2] or "St").title()
    last_part = (last_clean[:2] or "Wi").title()
    date_part = created_at.strftime("%m%d")
    return f"{first_part}{last_part}@{date_part}!"


def get_email_configuration_status() -> tuple[bool, list[str]]:
    required = {
        "STOCKWISE_SMTP_HOST": os.getenv("STOCKWISE_SMTP_HOST", "").strip(),
        "STOCKWISE_SMTP_PORT": os.getenv("STOCKWISE_SMTP_PORT", "").strip(),
        "STOCKWISE_SMTP_USER": os.getenv("STOCKWISE_SMTP_USER", "").strip(),
        "STOCKWISE_SMTP_PASSWORD": os.getenv("STOCKWISE_SMTP_PASSWORD", "").strip(),
        "STOCKWISE_MAIL_FROM": os.getenv("STOCKWISE_MAIL_FROM", os.getenv("STOCKWISE_SMTP_USER", "")).strip(),
    }
    missing = [name for name, value in required.items() if not value]
    return len(missing) == 0, missing


def send_employee_invite_email(
    to_email: str,
    employee_name: str,
    role: str,
    login_email: str,
    temporary_password: str,
    login_url: str,
) -> bool:
    email_ready, _missing = get_email_configuration_status()
    if not email_ready:
        print("StockWise email not sent: missing SMTP configuration.")
        return False

    host = os.getenv("STOCKWISE_SMTP_HOST", "").strip()
    port_value = os.getenv("STOCKWISE_SMTP_PORT", "587").strip()
    smtp_user = os.getenv("STOCKWISE_SMTP_USER", "").strip()
    smtp_password = os.getenv("STOCKWISE_SMTP_PASSWORD", "").strip()
    mail_from = os.getenv("STOCKWISE_MAIL_FROM", smtp_user).strip()

    try:
        port = int(port_value or 587)
    except ValueError:
        print("StockWise email not sent: invalid SMTP port configuration.")
        return False

    message = EmailMessage()
    message["Subject"] = "Your StockWise account is ready"
    message["From"] = mail_from
    message["To"] = to_email
    message.set_content(
        "Hello {name},\n\n"
        "You have been added to StockWise as {role}.\n\n"
        "Login email: {email}\n"
        "Temporary password: {password}\n"
        "Login page: {login_url}\n\n"
        "Please log in using the temporary password. You may update your password after signing in.\n\n"
        "Thank you,\n"
        "StockWise"
        .format(
            name=employee_name,
            role=role,
            email=login_email,
            password=temporary_password,
            login_url=login_url,
        )
    )

    try:
        if port == 465:
            with smtplib.SMTP_SSL(host, port, timeout=12) as smtp:
                smtp.login(smtp_user, smtp_password)
                smtp.send_message(message)
        else:
            with smtplib.SMTP(host, port, timeout=12) as smtp:
                smtp.starttls()
                smtp.login(smtp_user, smtp_password)
                smtp.send_message(message)
        return True
    except Exception:
        print("StockWise email not sent: SMTP delivery failed.")
        return False


def send_existing_employee_access_email(
    to_email: str,
    employee_name: str,
    role: str,
    login_url: str,
) -> bool:
    email_ready, _missing = get_email_configuration_status()
    if not email_ready:
        print("StockWise email not sent: missing SMTP configuration.")
        return False

    host = os.getenv("STOCKWISE_SMTP_HOST", "").strip()
    port_value = os.getenv("STOCKWISE_SMTP_PORT", "587").strip()
    smtp_user = os.getenv("STOCKWISE_SMTP_USER", "").strip()
    smtp_password = os.getenv("STOCKWISE_SMTP_PASSWORD", "").strip()
    mail_from = os.getenv("STOCKWISE_MAIL_FROM", smtp_user).strip()

    try:
        port = int(port_value or 587)
    except ValueError:
        print("StockWise email not sent: invalid SMTP port configuration.")
        return False

    message = EmailMessage()
    message["Subject"] = "You were added to a StockWise store"
    message["From"] = mail_from
    message["To"] = to_email
    message.set_content(
        "Hello {name},\n\n"
        "You have been added to another StockWise store workspace as {role}.\n\n"
        "Use your existing StockWise email and password to sign in.\n"
        "Login page: {login_url}\n\n"
        "Thank you,\n"
        "StockWise"
        .format(name=employee_name, role=role, login_url=login_url)
    )

    try:
        if port == 465:
            with smtplib.SMTP_SSL(host, port, timeout=12) as smtp:
                smtp.login(smtp_user, smtp_password)
                smtp.send_message(message)
        else:
            with smtplib.SMTP(host, port, timeout=12) as smtp:
                smtp.starttls()
                smtp.login(smtp_user, smtp_password)
                smtp.send_message(message)
        return True
    except Exception:
        print("StockWise email not sent: SMTP delivery failed.")
        return False


def create_employee_account(first_name: str, last_name: str, email: str, role: str) -> tuple[bool, str, str | None]:
    if not is_owner_user():
        return False, "Only the Owner can add employees.", None

    cleaned_first = " ".join(str(first_name or "").split())
    cleaned_last = " ".join(str(last_name or "").split())
    cleaned_name = f"{cleaned_first} {cleaned_last}".strip()
    cleaned_email = normalize_email(email or "")
    raw_role = str(role or "").strip()
    cleaned_role = normalize_role(raw_role)

    if not cleaned_first:
        return False, "Please enter the employee first name.", None
    if not cleaned_last:
        return False, "Please enter the employee last name.", None
    if not cleaned_email:
        return False, "Please enter the employee email address.", None
    if not is_valid_email_format(cleaned_email):
        return False, "Please enter a valid employee email address.", None
    if not raw_role:
        return False, "Please select a role.", None
    if raw_role not in ROLE_ALIASES and raw_role not in FINAL_ROLE_OPTIONS:
        return False, "Please choose a valid employee role.", None
    if cleaned_role not in EMPLOYEE_ROLE_OPTIONS:
        return False, "Please choose a valid employee role.", None

    store_id = get_current_store_id()
    owner_id = get_current_user_id()
    if not store_id or not owner_id:
        return False, "Your store workspace could not be verified.", None

    created_at = datetime.now()
    existing_user = get_user_by_email(cleaned_email)

    try:
        ensure_multi_user_schema()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        try:
            if existing_user:
                employee_id = int(existing_user["user_id"])
                if employee_id == owner_id:
                    return False, "The Owner account is already part of this store.", None

                cursor.execute(
                    """
                    SELECT membership_id, account_status, is_active
                    FROM store_memberships
                    WHERE store_id = %s AND user_id = %s
                    LIMIT 1
                    """,
                    (store_id, employee_id),
                )
                membership = cursor.fetchone()

                if membership:
                    status_key = normalize_employee_status(membership.get("account_status"), membership.get("is_active", 1))
                    if status_key == "active":
                        return False, "This employee is already active in this store.", None
                    if status_key == "deactivated":
                        return False, "This employee already exists but is deactivated. You can activate their access from the Employee List.", None
                    cursor.execute(
                        """
                        UPDATE store_memberships
                        SET role = %s,
                            account_status = 'active',
                            is_active = 1,
                            reactivated_at = NOW(),
                            removed_at = NULL
                        WHERE membership_id = %s
                        """,
                        (cleaned_role, membership.get("membership_id")),
                    )
                    conn.commit()
                    sync_user_legacy_access_from_membership(employee_id, {
                        "store_id": store_id,
                        "role": cleaned_role,
                        "account_status": "active",
                        "is_active": 1,
                    })
                    add_activity_log("Employee rehired", "Settings", "Success")
                    notify_store_roles({"Owner"}, "Employee added back", f"{cleaned_name or existing_user.get('full_name') or 'An employee'} was added back to this store.", "employee", actor_user_id=owner_id, include_actor=True, store_id=store_id)
                    add_notification(employee_id, "Access restored", f"You were added back as {cleaned_role} for this store.", "employee", store_id=store_id, target_role=cleaned_role)
                    return True, "Employee added back to this store.", None

                cursor.execute(
                    """
                    INSERT INTO store_memberships (store_id, user_id, role, account_status, is_active, created_by, joined_at)
                    VALUES (%s, %s, %s, 'active', 1, %s, NOW())
                    """,
                    (store_id, employee_id, cleaned_role, owner_id),
                )
                conn.commit()
                sync_user_legacy_access_from_membership(employee_id, {
                    "store_id": store_id,
                    "role": cleaned_role,
                    "account_status": "active",
                    "is_active": 1,
                })
                login_url = url_for("auth", _external=True)
                email_sent = send_existing_employee_access_email(
                    cleaned_email,
                    cleaned_name or existing_user.get("full_name") or "Employee",
                    cleaned_role,
                    login_url,
                )
                add_activity_log("Employee added", "Settings", "Success")
                notify_store_roles({"Owner"}, "Employee added", f"{cleaned_name or existing_user.get('full_name') or 'An employee'} was added as {cleaned_role}.", "employee", actor_user_id=owner_id, include_actor=True, store_id=store_id)
                add_notification(employee_id, "You were added to this store", f"Use your existing StockWise account to access this store as {cleaned_role}.", "employee", store_id=store_id, target_role=cleaned_role)
                if email_sent:
                    return True, "Employee added to this store. Login instructions were sent to their email.", None
                return True, "Employee added to this store, but email could not be sent. Please share their existing login details directly.", None

            temp_password = build_employee_temp_password(cleaned_first, cleaned_last, cleaned_role, created_at)
            username_base = cleaned_email.split("@")[0]
            username = username_base
            counter = 1
            while get_user_by_username(username) is not None:
                username = f"{username_base}{counter}"
                counter += 1

            cursor.execute(
                """
                INSERT INTO users (full_name, email, position, role, store_id, created_by, username, password_hash, is_active, account_status)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 1, 'active')
                """,
                (cleaned_name, cleaned_email, cleaned_role, cleaned_role, store_id, owner_id, username, generate_password_hash(temp_password)),
            )
            employee_id = cursor.lastrowid
            cursor.execute(
                """
                INSERT IGNORE INTO store_memberships (store_id, user_id, role, account_status, is_active, created_by, joined_at)
                VALUES (%s, %s, %s, 'active', 1, %s, NOW())
                """,
                (store_id, employee_id, cleaned_role, owner_id),
            )
            conn.commit()
        finally:
            cursor.close()
            conn.close()

        login_url = url_for("auth", _external=True)
        email_sent = send_employee_invite_email(
            cleaned_email,
            cleaned_name,
            cleaned_role,
            cleaned_email,
            temp_password,
            login_url,
        )

        add_activity_log("Employee added", "Settings", "Success")
        notify_store_roles({"Owner"}, "Employee added", f"{cleaned_name} was added as {cleaned_role}.", "employee", actor_user_id=owner_id, include_actor=True, store_id=store_id)
        add_notification(employee_id, "Your StockWise account is ready", f"You were added as {cleaned_role} for this store.", "employee", store_id=store_id, target_role=cleaned_role)

        if email_sent:
            return True, "Employee added. Login instructions were sent to the employee’s email.", None
        return True, f"Employee added, but email could not be sent. Temporary password: {temp_password}", temp_password
    except Exception:
        return False, "The employee account could not be created right now.", None


def update_employee_role(employee_id: int, role: str) -> tuple[bool, str]:
    if not is_owner_user():
        return False, "Only the Owner can change employee roles."
    raw_role = str(role or "").strip()
    cleaned_role = normalize_role(raw_role)
    if not raw_role or (raw_role not in ROLE_ALIASES and raw_role not in FINAL_ROLE_OPTIONS):
        return False, "Please choose a valid employee role."
    if cleaned_role not in EMPLOYEE_ROLE_OPTIONS:
        return False, "Please choose a valid employee role."
    store_id = get_current_store_id()
    owner_id = get_current_user_id()
    try:
        ensure_multi_user_schema()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        try:
            cursor.execute(
                """
                SELECT membership_id, account_status, is_active
                FROM store_memberships
                WHERE user_id = %s AND store_id = %s AND user_id <> %s
                LIMIT 1
                """,
                (employee_id, store_id, owner_id),
            )
            membership = cursor.fetchone()
            if not membership:
                return False, "Employee could not be found in this store."
            previous_status = normalize_employee_status(membership.get("account_status"), membership.get("is_active", 1))
            if previous_status == "removed":
                return False, "Removed employee roles can be changed when they are added back."
            cursor.execute(
                """
                UPDATE store_memberships
                SET role = %s
                WHERE membership_id = %s
                """,
                (cleaned_role, membership.get("membership_id")),
            )
            if get_store_id_for_user(employee_id) == store_id:
                cursor.execute(
                    """
                    UPDATE users
                    SET role = %s, position = %s
                    WHERE user_id = %s
                    """,
                    (cleaned_role, cleaned_role, employee_id),
                )
            changed = cursor.rowcount >= 0
            conn.commit()
        finally:
            cursor.close()
            conn.close()
        if changed:
            add_activity_log("Role changed", "Settings", "Success")
            notify_store_roles({"Owner"}, "Employee role updated", f"An employee role was changed to {cleaned_role}.", "employee", actor_user_id=owner_id, include_actor=True, store_id=store_id)
            add_notification(employee_id, "Your role was updated", f"Your StockWise role is now {cleaned_role}.", "employee", store_id=store_id, target_role=cleaned_role)
            return True, "Employee role updated."
        return False, "Employee could not be found in this store."
    except Exception:
        return False, "Employee role could not be updated right now."


def update_employee_status(employee_id: int, status: str) -> tuple[bool, str]:
    if not is_owner_user():
        return False, "Only the Owner can update employee status."
    status_key = normalize_employee_status(status)
    store_id = get_current_store_id()
    owner_id = get_current_user_id()
    if employee_id == owner_id:
        return False, "The Owner account cannot be changed from the employee list."
    try:
        ensure_multi_user_schema()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        try:
            cursor.execute(
                """
                SELECT sm.membership_id, sm.user_id, sm.role, sm.account_status, sm.is_active, u.full_name
                FROM store_memberships sm
                JOIN users u ON u.user_id = sm.user_id
                WHERE sm.user_id = %s AND sm.store_id = %s AND sm.user_id <> %s
                LIMIT 1
                """,
                (employee_id, store_id, owner_id),
            )
            row = cursor.fetchone()
            if not row:
                return False, "Employee account was not found in this store."
            previous_status = normalize_employee_status(row.get("account_status"), row.get("is_active", 1))
            if previous_status == "removed" and status_key in {"active", "deactivated"}:
                return False, "Removed employee accounts must be added back using Rehire."
            if status_key == "active" and previous_status != "deactivated":
                return False, "Only deactivated employee accounts can be activated."
            if status_key == "deactivated" and previous_status != "active":
                return False, "Only active employee accounts can be deactivated."
            if status_key == "removed" and previous_status == "removed":
                return False, "This employee is already removed from the store."
            cursor.execute(
                """
                UPDATE store_memberships
                SET account_status = %s,
                    is_active = %s,
                    removed_at = CASE WHEN %s = 'removed' THEN NOW() ELSE removed_at END,
                    reactivated_at = CASE WHEN %s = 'active' THEN NOW() ELSE reactivated_at END
                WHERE membership_id = %s
                """,
                (status_key, 1 if status_key == "active" else 0, status_key, status_key, row.get("membership_id")),
            )
            conn.commit()
        finally:
            cursor.close()
            conn.close()
        sync_user_after_membership_change(employee_id, store_id)
        if previous_status != status_key:
            action = "Employee status updated"
            if status_key == "removed":
                action = "Employee removed"
            elif status_key == "deactivated":
                action = "Employee deactivated"
            elif status_key == "active":
                action = "Employee reactivated"
            add_activity_log(action, "Settings", "Success")
            notify_store_roles({"Owner"}, action, f"{row.get('full_name') or 'An employee'} is now {status_key.replace('_', ' ')}.", "employee", actor_user_id=owner_id, include_actor=True, store_id=store_id)
        return True, f"Employee status updated to {status_key.replace('_', ' ').title()}."
    except Exception:
        return False, "The employee status could not be updated right now."


def rehire_employee(employee_id: int) -> tuple[bool, str]:
    if not is_owner_user():
        return False, "Only the Owner can add back employees."
    store_id = get_current_store_id()
    owner_id = get_current_user_id()
    if employee_id == owner_id:
        return False, "The Owner account cannot be changed from the employee list."
    try:
        ensure_multi_user_schema()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        try:
            cursor.execute(
                """
                SELECT sm.membership_id, sm.role, sm.account_status, sm.is_active, u.full_name
                FROM store_memberships sm
                JOIN users u ON u.user_id = sm.user_id
                WHERE sm.user_id = %s AND sm.store_id = %s AND sm.user_id <> %s
                LIMIT 1
                """,
                (employee_id, store_id, owner_id),
            )
            row = cursor.fetchone()
            if not row:
                return False, "Employee account was not found in this store."
            previous_status = normalize_employee_status(row.get("account_status"), row.get("is_active", 1))
            if previous_status != "removed":
                return False, "Only removed employees can be added back."
            cursor.execute(
                """
                UPDATE store_memberships
                SET account_status = 'active',
                    is_active = 1,
                    reactivated_at = NOW(),
                    removed_at = NULL
                WHERE membership_id = %s
                """,
                (row.get("membership_id"),),
            )
            conn.commit()
        finally:
            cursor.close()
            conn.close()
        sync_user_legacy_access_from_membership(employee_id, {
            "store_id": store_id,
            "role": row.get("role"),
            "account_status": "active",
            "is_active": 1,
        })
        add_activity_log("Employee rehired", "Settings", "Success")
        notify_store_roles({"Owner"}, "Employee added back", f"{row.get('full_name') or 'An employee'} was added back to this store.", "employee", actor_user_id=owner_id, include_actor=True, store_id=store_id)
        add_notification(employee_id, "Access restored", "You were added back to this store workspace.", "employee", store_id=store_id, target_role=row.get("role"))
        return True, "Employee added back to this store."
    except Exception:
        return False, "Employee could not be added back right now."


def deactivate_employee(employee_id: int) -> tuple[bool, str]:
    return update_employee_status(employee_id, "deactivated")


def reset_employee_password(employee_id: int) -> tuple[bool, str]:
    if not is_owner_user():
        return False, "Only the Owner can reset employee passwords."
    store_id = get_current_store_id()
    temp_password = f"Sw{uuid.uuid4().hex[:6]}!1"
    try:
        ensure_multi_user_schema()
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute(
                """
                UPDATE users
                SET password_hash = %s, is_active = 1, account_status = 'active' 
                WHERE user_id = %s AND store_id = %s AND user_id <> %s
                """,
                (generate_password_hash(temp_password), employee_id, store_id, get_current_user_id()),
            )
            changed = cursor.rowcount > 0
            conn.commit()
        finally:
            cursor.close()
            conn.close()
        if changed:
            add_activity_log("Temporary password reset", "Settings", "Success")
            notify_store_roles({"Owner"}, "Temporary password reset", "An employee temporary password was reset.", "employee", actor_user_id=get_current_user_id(), include_actor=True, store_id=store_id)
            add_notification(employee_id, "Temporary password reset", "Your temporary password was reset by the Owner.", "employee", store_id=store_id, target_role=None)
            return True, "Employee temporary password was reset."
        return False, "Employee could not be found in this store."
    except Exception:
        return False, "Temporary password could not be reset right now."


def add_notification(
    user_id: int | None,
    title: str,
    message: str,
    event_type: str = "system",
    is_read: bool = False,
    store_id: int | None = None,
    target_role: str | None = None,
    dedupe_minutes: int = 30,
) -> None:
    """Add one lightweight role-aware notification without blocking the workflow."""
    if not user_id:
        return
    try:
        ensure_notifications_table()
        if store_id is None:
            store_id = get_store_id_for_user(user_id)
        clean_role = normalize_role(target_role) if target_role else None
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        try:
            if dedupe_minutes > 0:
                cursor.execute(
                    """
                    SELECT notification_id
                    FROM notifications
                    WHERE user_id = %s
                      AND title = %s
                      AND message = %s
                      AND event_type = %s
                      AND created_at >= DATE_SUB(NOW(), INTERVAL %s MINUTE)
                    LIMIT 1
                    """,
                    (user_id, title[:120], message[:255], event_type[:50], int(dedupe_minutes)),
                )
                if cursor.fetchone():
                    return
            cursor.execute(
                """
                INSERT INTO notifications (user_id, store_id, title, message, event_type, target_role, is_read)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                """,
                (user_id, store_id, title[:120], message[:255], event_type[:50], clean_role, 1 if is_read else 0),
            )
            conn.commit()
        finally:
            cursor.close()
            conn.close()
    except Exception:
        # Notifications should never break upload, settings, onboarding, or navigation.
        return


def get_active_store_users_for_roles(roles: list[str] | set[str], store_id: int | None = None) -> list[dict[str, Any]]:
    """Return active users in the current store with one of the requested roles."""
    normalized_roles = sorted({normalize_role(role) for role in roles if normalize_role(role) in FINAL_ROLE_OPTIONS})
    if not normalized_roles:
        return []
    if store_id is None:
        store_id = get_current_store_id()
    if not store_id:
        return []
    try:
        ensure_multi_user_schema()
        placeholders = ", ".join(["%s"] * len(normalized_roles))
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        try:
            cursor.execute(
                f"""
                SELECT u.user_id, u.full_name, sm.role, sm.role AS position
                FROM store_memberships sm
                JOIN users u ON u.user_id = sm.user_id
                WHERE sm.store_id = %s
                  AND sm.account_status = 'active'
                  AND sm.is_active = 1
                  AND sm.role IN ({placeholders})
                """,
                [store_id, *normalized_roles],
            )
            rows = cursor.fetchall() or []
        finally:
            cursor.close()
            conn.close()
        return rows
    except Exception:
        return []


def notify_store_roles(
    roles: list[str] | set[str],
    title: str,
    message: str,
    event_type: str = "system",
    actor_user_id: int | None = None,
    include_actor: bool = True,
    store_id: int | None = None,
) -> None:
    """Notify active users in the same store whose roles should see this alert."""
    if store_id is None:
        store_id = get_current_store_id()
    notified: set[int] = set()
    for user in get_active_store_users_for_roles(roles, store_id=store_id):
        target_user_id = user.get("user_id")
        if not target_user_id:
            continue
        if not include_actor and actor_user_id and int(target_user_id) == int(actor_user_id):
            continue
        if int(target_user_id) in notified:
            continue
        notified.add(int(target_user_id))
        add_notification(
            int(target_user_id),
            title,
            message,
            event_type,
            store_id=store_id,
            target_role=normalize_role(user.get("role") or user.get("position")),
        )


def notify_current_user(title: str, message: str, event_type: str = "system") -> None:
    current_user_id = get_current_user_id()
    if current_user_id:
        add_notification(
            current_user_id,
            title,
            message,
            event_type,
            store_id=get_current_store_id(),
            target_role=get_session_role(),
        )


def notify_upload_event(title: str, message: str, event_type: str = "upload", success: bool = True) -> None:
    """Send only major upload notifications. Minor file-selection updates stay as page feedback."""
    current_user_id = get_current_user_id()
    if not current_user_id:
        return

    # Successful file selection is intentionally not a notification. Results generation has its own notification.
    if success and event_type == "upload":
        return

    notify_current_user(title, message, event_type)
    target_roles = {"Owner", "Store Manager"}
    notify_store_roles(target_roles, title, message, event_type, actor_user_id=current_user_id, include_actor=False)


def get_recent_notifications(user_id: int | None, limit: int = 8) -> dict[str, Any]:
    empty = {"items": [], "unread_count": 0, "total_count": 0}
    if not user_id:
        return empty
    try:
        ensure_notifications_table()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        try:
            cursor.execute(
                """
                SELECT notification_id, title, message, event_type, target_role, is_read, created_at
                FROM notifications
                WHERE user_id = %s
                ORDER BY created_at DESC, notification_id DESC
                LIMIT %s
                """,
                (user_id, limit),
            )
            rows = cursor.fetchall() or []
            cursor.execute(
                """
                SELECT COUNT(*) AS unread_count
                FROM notifications
                WHERE user_id = %s AND is_read = 0
                """,
                (user_id,),
            )
            unread_row = cursor.fetchone() or {}
            notifications = []
            for row in rows:
                created_at = row.get("created_at")
                created_label = ""
                if created_at:
                    try:
                        created_label = created_at.strftime("%b %d, %Y %I:%M %p")
                    except Exception:
                        created_label = str(created_at)
                notifications.append({
                    "id": row.get("notification_id"),
                    "title": row.get("title") or "Notification",
                    "message": row.get("message") or "",
                    "event_type": row.get("event_type") or "system",
                    "target_role": normalize_role(row.get("target_role")) if row.get("target_role") else "",
                    "is_read": bool(row.get("is_read")),
                    "created_label": created_label,
                })
            return {
                "items": notifications,
                "unread_count": int(unread_row.get("unread_count") or 0),
                "total_count": len(notifications),
            }
        finally:
            cursor.close()
            conn.close()
    except Exception:
        return empty


def add_result_notifications(user_id: int | None, processed_df: pd.DataFrame | None) -> None:
    actor_user = get_user_by_id(user_id) if user_id else None
    actor_name = (actor_user or {}).get("full_name") or session.get("user_name", "A store user")
    store_id = get_store_id_for_user(user_id) if user_id else get_current_store_id()
    success_message = f"{actor_name} generated new results. Dashboard, Insights, Products, and Reports now reflect the latest processed sales data."
    notify_store_roles(
        {"Owner", "Store Manager"},
        "Results generated successfully",
        success_message,
        "success",
        actor_user_id=user_id,
        include_actor=True,
        store_id=store_id,
    )
    if user_id and normalize_role((actor_user or {}).get("role") or (actor_user or {}).get("position")) not in {"Owner", "Store Manager"}:
        add_notification(
            user_id,
            "Results generated successfully",
            "The latest sales records were processed successfully.",
            "success",
            store_id=store_id,
            target_role=normalize_role((actor_user or {}).get("role") or (actor_user or {}).get("position")),
        )
    try:
        metrics = get_dashboard_metrics(processed_df)
        high_risk = int(metrics.get("high_risk") or 0)
    except Exception:
        high_risk = 0
    if high_risk > 0:
        product_word = "product" if high_risk == 1 else "products"
        notify_store_roles(
            {"Owner", "Store Manager", "Operational Assistant"},
            "High-risk products detected",
            f"{high_risk} {product_word} may need inventory review based on the latest generated results.",
            "warning",
            actor_user_id=user_id,
            include_actor=True,
            store_id=store_id,
        )


@app.route("/notifications/mark_read", methods=["POST"])
def mark_notifications_read():
    """Mark the current user's visible notifications as read."""
    user_id = get_current_user_id()
    if not user_id:
        return jsonify({"success": False}), 401
    try:
        ensure_notifications_table()
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute(
                """
                UPDATE notifications
                SET is_read = 1
                WHERE user_id = %s AND is_read = 0
                """,
                (user_id,),
            )
            conn.commit()
        finally:
            cursor.close()
            conn.close()
        return jsonify({"success": True})
    except Exception:
        return jsonify({"success": False}), 200


@app.route("/notifications/<int:notification_id>/mark_read", methods=["POST"])
def mark_single_notification_read(notification_id: int):
    """Mark only one notification as read for the currently signed-in user."""
    user_id = get_current_user_id()
    if not user_id:
        return jsonify({"success": False}), 401
    try:
        ensure_notifications_table()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        try:
            cursor.execute(
                """
                UPDATE notifications
                SET is_read = 1
                WHERE notification_id = %s AND user_id = %s
                """,
                (notification_id, user_id),
            )
            changed_count = cursor.rowcount
            cursor.execute(
                """
                SELECT COUNT(*) AS unread_count
                FROM notifications
                WHERE user_id = %s AND is_read = 0
                """,
                (user_id,),
            )
            unread_row = cursor.fetchone() or {}
            conn.commit()
        finally:
            cursor.close()
            conn.close()
        return jsonify({"success": True, "updated": changed_count > 0, "unread_count": int(unread_row.get("unread_count") or 0)})
    except Exception:
        return jsonify({"success": False}), 200

# =========================================
# AUTH HELPERS
# =========================================
def login_required(view_func):
    @wraps(view_func)
    def wrapped_view(*args, **kwargs):
        if not session.get("user_logged_in"):
            return redirect(url_for("auth"))
        return view_func(*args, **kwargs)

    return wrapped_view


@app.context_processor
def inject_global_template_context():
    user_id = get_current_user_id() if session.get("user_logged_in") else None
    notification_context = get_recent_notifications(user_id) if user_id else {"items": [], "unread_count": 0, "total_count": 0}
    return {
        "current_user": get_current_user(),
        "current_path": request.path,
        "notification_context": notification_context,
        "current_role": get_session_role() if session.get("user_logged_in") else None,
        "visible_nav_keys": get_visible_nav_keys() if session.get("user_logged_in") else set(),
        "is_owner_account": is_owner_user() if session.get("user_logged_in") else False,
    }


def get_current_user() -> dict[str, str] | None:
    if session.get("user_logged_in"):
        return {
            "name": session.get("user_name", "Admin"),
            "email": session.get("user_email", ""),
            "position": get_session_role(),
            "role": get_session_role(),
            "profile_image": session.get("user_profile_image", ""),
        }
    return None


@app.before_request
def require_first_time_store_setup():
    """Keep new users in the first-time setup flow until their store details are saved."""
    if not session.get("user_logged_in"):
        return None

    endpoint = request.endpoint or ""
    allowed_endpoints = {"auth", "logout", "static", "first_time_setup", "preparing_dashboard"}
    if endpoint in allowed_endpoints:
        return None

    user_id = get_current_user_id()
    if user_id and not session.get("user_store_id"):
        try:
            ensure_user_workspace(user_id)
            refresh_session_user(user_id)
        except Exception:
            pass
    if user_needs_onboarding(user_id):
        return redirect(url_for("first_time_setup"))

    return None


def normalize_email(email: str) -> str:
    return email.strip().lower()


def is_valid_email_format(email: str) -> bool:
    return bool(re.match(r"^[^\s@]+@[^\s@]+\.[^\s@]+$", normalize_email(email or "")))


def get_user_by_email(email: str) -> dict[str, Any] | None:
    normalized_email = normalize_email(email)
    ensure_multi_user_schema()
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute(
            """
            SELECT user_id, full_name, email, position, role, store_id, profile_image, username, password_hash, is_active, account_status, last_login_at
            FROM users
            WHERE LOWER(email) = %s
            LIMIT 1
            """,
            (normalized_email,)
        )
        return cursor.fetchone()
    finally:
        cursor.close()
        conn.close()


def get_user_by_username(username: str) -> dict[str, Any] | None:
    normalized_username = username.strip().lower()
    ensure_multi_user_schema()
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute(
            """
            SELECT user_id, full_name, email, position, role, store_id, username, password_hash, is_active, account_status, last_login_at
            FROM users
            WHERE LOWER(username) = %s
            LIMIT 1
            """,
            (normalized_username,)
        )
        return cursor.fetchone()
    finally:
        cursor.close()
        conn.close()


def create_user(full_name: str, email: str, password: str, position: str = "Owner") -> int:
    ensure_multi_user_schema()
    normalized_email = normalize_email(email)
    base_username = normalized_email.split("@")[0]
    username = base_username
    counter = 1

    while get_user_by_username(username) is not None:
        username = f"{base_username}{counter}"
        counter += 1

    password_hash = generate_password_hash(password)

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            """
            INSERT INTO users (full_name, email, position, role, username, password_hash, account_status)
            VALUES (%s, %s, %s, %s, %s, %s, 'active')
            """,
            (full_name.strip(), normalized_email, normalize_role(position), normalize_role(position), username, password_hash)
        )
        conn.commit()
        return cursor.lastrowid
    finally:
        cursor.close()
        conn.close()


def get_user_by_id(user_id: int) -> dict[str, Any] | None:
    ensure_multi_user_schema()
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute(
            """
            SELECT user_id, full_name, email, position, role, store_id, profile_image, username, password_hash, is_active, account_status, last_login_at
            FROM users
            WHERE user_id = %s
            LIMIT 1
            """,
            (user_id,)
        )
        return cursor.fetchone()
    finally:
        cursor.close()
        conn.close()


def update_user_profile(user_id: int, full_name: str, email: str, new_password: str | None = None, position: str | None = None, profile_image: str | None = None) -> bool:
    ensure_multi_user_schema()
    normalized_email = normalize_email(email)
    cleaned_position = normalize_role(position or "Owner")

    existing_user = get_user_by_email(normalized_email)
    if existing_user and existing_user["user_id"] != user_id:
        return False

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        if new_password:
            cursor.execute(
                """
                UPDATE users
                SET full_name = %s,
                    email = %s,
                    position = %s,
                    role = %s,
                    profile_image = COALESCE(%s, profile_image),
                    password_hash = %s
                WHERE user_id = %s
                """,
                (
                    full_name.strip(),
                    normalized_email,
                    cleaned_position,
                    normalize_role(cleaned_position),
                    profile_image,
                    generate_password_hash(new_password),
                    user_id,
                )
            )
        else:
            cursor.execute(
                """
                UPDATE users
                SET full_name = %s,
                    email = %s,
                    position = %s,
                    role = %s,
                    profile_image = COALESCE(%s, profile_image)
                WHERE user_id = %s
                """,
                (
                    full_name.strip(),
                    normalized_email,
                    cleaned_position,
                    normalize_role(cleaned_position),
                    profile_image,
                    user_id,
                )
            )

        conn.commit()
        return True
    finally:
        cursor.close()
        conn.close()


# =========================================
# REPOSITORY / STATE ACCESS HELPERS
# Replace these with real database repository calls later
# =========================================
user_states: dict[int, AppState] = {}
guest_state = AppState()


def get_current_user_id() -> int | None:
    raw_user_id = session.get("user_id")
    try:
        return int(raw_user_id) if raw_user_id is not None else None
    except (TypeError, ValueError):
        return None


def get_app_state() -> AppState:
    user_id = get_current_user_id()
    if user_id is None:
        return guest_state

    if user_id not in user_states:
        user_states[user_id] = AppState()

    return user_states[user_id]


def get_selected_dataset() -> pd.DataFrame | None:
    return get_app_state().selected_data




def get_selected_filename() -> str | None:
    return get_app_state().selected_filename




def set_upload_feedback(message: str | None, message_type: str | None = None) -> None:
    state = get_app_state()
    state.upload_message = message
    state.upload_message_type = message_type


def get_upload_feedback() -> tuple[str | None, str | None]:
    state = get_app_state()
    return state.upload_message, state.upload_message_type


def clear_selected_dataset() -> None:
    state = get_app_state()
    state.selected_data = None
    state.selected_filename = None
    state.selected_file_size = None
    state.selected_file_type = None
    state.selected_at = None
    state.last_upload_mode = "new"




def set_selected_upload_mode(upload_mode: str | None) -> str:
    normalized = str(upload_mode or "new").strip().lower() or "new"
    if normalized not in {"new", "append", "replace"}:
        normalized = "new"
    get_app_state().last_upload_mode = normalized
    return normalized


def get_upload_mode_label(upload_mode: str | None) -> str:
    normalized = str(upload_mode or "new").strip().lower() or "new"
    return UPLOAD_MODE_LABELS.get(normalized, UPLOAD_MODE_LABELS["new"])


def get_file_type_label(filename: str | None) -> str:
    if not filename or "." not in filename:
        return "Unknown file"
    extension = filename.rsplit(".", 1)[1].upper()
    if extension == "CSV":
        return "CSV file"
    if extension in {"XLS", "XLSX"}:
        return f"Excel {extension} file"
    if extension in {"PNG", "JPG", "JPEG", "GIF", "WEBP"}:
        return f"Image {extension} file"
    return f"{extension} file"


def format_file_size(size_bytes: int | None) -> str:
    if not size_bytes or size_bytes <= 0:
        return "Size unavailable"
    units = ["B", "KB", "MB", "GB"]
    value = float(size_bytes)
    unit_index = 0
    while value >= 1024 and unit_index < len(units) - 1:
        value /= 1024
        unit_index += 1
    if unit_index == 0:
        return f"{int(value)} {units[unit_index]}"
    return f"{value:.1f} {units[unit_index]}"


def get_upload_file_size(file_storage) -> int | None:
    try:
        stream = file_storage.stream
        current_position = stream.tell()
        stream.seek(0, 2)
        size = stream.tell()
        stream.seek(current_position)
        return int(size)
    except Exception:
        try:
            return int(getattr(file_storage, "content_length", 0) or 0) or None
        except Exception:
            return None


def get_selected_file_info() -> dict[str, str]:
    state = get_app_state()
    return {
        "name": state.selected_filename or "",
        "type": state.selected_file_type or get_file_type_label(state.selected_filename),
        "size": format_file_size(state.selected_file_size),
        "status": "Selected and ready for review" if state.selected_filename else "No file selected",
    }


def store_selected_dataset(df: pd.DataFrame, filename: str, upload_mode: str = "new", file_size: int | None = None, file_type: str | None = None) -> None:
    state = get_app_state()
    previous_size = state.selected_file_size
    previous_type = state.selected_file_type
    state.selected_data = df.copy()
    state.selected_filename = filename
    state.selected_file_size = file_size if file_size is not None else previous_size
    state.selected_file_type = file_type or previous_type or get_file_type_label(filename)
    state.selected_at = datetime.now()
    state.last_upload_mode = set_selected_upload_mode(upload_mode)




# =========================================
# FILE + NORMALIZATION HELPERS
# =========================================
def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def normalize_column_name(name: str) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9]+", "_", str(name).strip().lower()).strip("_")
    aliases = {
        "product": "product_name",
        "item": "product_name",
        "item_name": "product_name",
        "product_name": "product_name",
        "productname": "product_name",
        "product_name_": "product_name",
        "product_id": "product_id",
        "sku": "product_id",
        "item_id": "product_id",
        "category": "category",
        "product_category": "category",
        "current_stock": "current_stock",
        "latest_recorded_stock": "current_stock",
        "latest_stock": "current_stock",
        "stock": "current_stock",
        "stock_on_hand": "current_stock",
        "available_stock": "current_stock",
        "reorder_point": "reorder_point",
        "reorder_level": "reorder_point",
        "minimum_stock": "reorder_point",
        "unit_price": "unit_price",
        "price": "unit_price",
        "selling_price": "unit_price",
        "quantity": "quantity_sold",
        "quantity_sold": "quantity_sold",
        "qty": "quantity_sold",
        "qty_sold": "quantity_sold",
        "units_sold": "quantity_sold",
        "date": "date",
        "sales_date": "date",
        "transaction_date": "date",
        "purchase_date": "date",
        "time": "time",
        "transaction_time": "time",
        "purchase_time": "time",
        "is_payday": "is_payday_period",
        "is_payday_period": "is_payday_period",
        "payday_indicator": "is_payday_period",
        "payday_period": "is_payday_period",
        "payday": "is_payday_period",
        "unit_type": "unit_type",
        "uom": "unit_type",
    }
    return aliases.get(cleaned, cleaned)


def get_default_column_mapping() -> dict[str, str]:
    return dict(DEFAULT_COLUMN_MAPPING)


def friendly_column_default(value: str) -> str:
    special_values = {
        "product_id": "Product ID",
        "is_payday_period": "Payday Period",
    }
    text = str(value or "").strip()
    if not text:
        return ""
    lowered = text.lower()
    if lowered in special_values:
        return special_values[lowered]
    if "_" in text:
        return text.replace("_", " ").title()
    return text


def get_display_column_mapping(mapping: dict[str, str]) -> dict[str, str]:
    display_mapping: dict[str, str] = {}
    for field in DATA_FORMAT_SYSTEM_FIELDS:
        key = field["key"]
        display_mapping[key] = friendly_column_default(mapping.get(key, field["default_column"]))
    return display_mapping


def parse_column_mapping_json(raw_value: Any) -> dict[str, str]:
    mapping = get_default_column_mapping()
    if isinstance(raw_value, dict):
        incoming = raw_value
    else:
        try:
            incoming = json.loads(raw_value or "{}")
        except Exception:
            incoming = {}
    for field in DATA_FORMAT_SYSTEM_FIELDS:
        key = field["key"]
        value = str(incoming.get(key, "")).strip()
        mapping[key] = value or field["default_column"]
    return mapping


def get_data_format_preferences(source: dict[str, Any] | None = None) -> dict[str, Any]:
    prefs = source or get_settings_preferences()
    mapping = parse_column_mapping_json(prefs.get("column_mapping_json"))
    return {
        "data_date_format": prefs.get("data_date_format", DEFAULT_SETTINGS["data_date_format"]),
        "data_time_format": prefs.get("data_time_format", DEFAULT_SETTINGS["data_time_format"]),
        "payday_indicator_handling": prefs.get("payday_indicator_handling", DEFAULT_SETTINGS["payday_indicator_handling"]),
        "duplicate_handling": prefs.get("duplicate_handling", DEFAULT_SETTINGS["duplicate_handling"]),
        "column_mapping": mapping,
        "column_mapping_display": get_display_column_mapping(mapping),
        "column_mapping_json": json.dumps(mapping),
    }


def get_data_date_format_options() -> list[dict[str, str]]:
    return [
        {"key": "auto", "label": "Auto-detect"},
        {"key": "%Y-%m-%d", "label": "YYYY-MM-DD"},
        {"key": "%m/%d/%Y", "label": "MM/DD/YYYY"},
        {"key": "%d/%m/%Y", "label": "DD/MM/YYYY"},
        {"key": "%b %d, %Y", "label": "Month DD, YYYY"},
    ]


def get_data_time_format_options() -> list[dict[str, str]]:
    return [
        {"key": "auto", "label": "Auto-detect"},
        {"key": "%H:%M", "label": "24-hour time, HH:MM"},
        {"key": "%H:%M:%S", "label": "24-hour time, HH:MM:SS"},
        {"key": "%I:%M %p", "label": "12-hour time, HH:MM AM/PM"},
    ]


def get_payday_handling_options() -> list[dict[str, str]]:
    return [
        {"key": "auto", "label": "Auto-detect from date (15th and 30th)"},
        {"key": "use_column", "label": "Use uploaded payday indicator column when available"},
        {"key": "ignore", "label": "Do not mark payday periods"},
    ]


def get_duplicate_handling_options() -> list[dict[str, str]]:
    return [
        {"key": "remove_exact", "label": "Remove exact duplicate records"},
        {"key": "keep_all", "label": "Keep all records and only warn me"},
    ]


def normalize_boolean_flag(value: Any) -> int:
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "y", "payday", "during payday"}:
        return 1
    if text in {"0", "false", "no", "n", "regular", "not payday"}:
        return 0
    try:
        return 1 if float(text) > 0 else 0
    except Exception:
        return 0


def apply_column_mapping(df: pd.DataFrame, preferences: dict[str, Any] | None = None) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    working = df.copy()
    prefs = preferences or get_data_format_preferences()
    mapping = prefs.get("column_mapping") or get_default_column_mapping()
    normalized_lookup = {normalize_column_name(column): column for column in working.columns}
    raw_lookup = {str(column).strip().casefold(): column for column in working.columns}
    rename_map: dict[Any, str] = {}
    used_targets: set[str] = set()

    for field in DATA_FORMAT_SYSTEM_FIELDS:
        key = field["key"]
        internal_name = field["internal"]
        preferred_name = str(mapping.get(key, field["default_column"]) or "").strip()
        if not preferred_name or internal_name in used_targets:
            continue
        source_column = None
        if preferred_name.casefold() in raw_lookup:
            source_column = raw_lookup[preferred_name.casefold()]
        else:
            source_column = normalized_lookup.get(normalize_column_name(preferred_name))
        if source_column is not None:
            rename_map[source_column] = internal_name
            used_targets.add(internal_name)

    if rename_map:
        working = working.rename(columns=rename_map)
    return working


def parse_upload_dates(values, date_format: str | None = None) -> pd.Series:
    """Parse uploaded dates safely, including saved date format preferences."""
    raw = pd.Series(values).copy()
    parsed = pd.Series(pd.NaT, index=raw.index, dtype="datetime64[ns]")

    text_values = raw.astype(str).str.strip()
    numeric_values = pd.to_numeric(raw, errors="coerce")

    selected_format = (date_format or get_settings_preferences().get("data_date_format", "auto")).strip()
    if selected_format and selected_format != "auto":
        formatted_mask = text_values.ne("") & ~text_values.str.lower().isin({"nan", "none", "null"})
        if formatted_mask.any():
            parsed.loc[formatted_mask] = pd.to_datetime(
                text_values.loc[formatted_mask],
                format=selected_format,
                errors="coerce",
            )

    excel_serial_mask = numeric_values.between(20000, 60000, inclusive="both")
    if excel_serial_mask.any():
        parsed.loc[excel_serial_mask] = pd.to_datetime(
            numeric_values.loc[excel_serial_mask],
            unit="D",
            origin="1899-12-30",
            errors="coerce",
        )

    yyyymmdd_mask = parsed.isna() & text_values.str.fullmatch(r"\d{8}", na=False)
    if yyyymmdd_mask.any():
        parsed.loc[yyyymmdd_mask] = pd.to_datetime(
            text_values.loc[yyyymmdd_mask],
            format="%Y%m%d",
            errors="coerce",
        )

    remaining_mask = parsed.isna()
    if remaining_mask.any():
        parsed.loc[remaining_mask] = pd.to_datetime(raw.loc[remaining_mask], errors="coerce")

    return parsed


def normalize_dataframe(df: pd.DataFrame, preferences: dict[str, Any] | None = None) -> pd.DataFrame:
    prefs = preferences or get_data_format_preferences()
    working = apply_column_mapping(df, prefs)
    rename_map: dict[str, str] = {}
    used_names: set[str] = set()

    for original in working.columns:
        normalized = normalize_column_name(original)
        candidate = normalized
        suffix = 2
        while candidate in used_names:
            candidate = f"{normalized}_{suffix}"
            suffix += 1
        used_names.add(candidate)
        rename_map[original] = candidate

    working = working.rename(columns=rename_map)

    if "date" in working.columns:
        working["date"] = parse_upload_dates(working["date"], prefs.get("data_date_format"))

    for numeric_col in ["quantity_sold", "unit_price", "current_stock", "reorder_point"]:
        if numeric_col in working.columns:
            working[numeric_col] = pd.to_numeric(working[numeric_col], errors="coerce")

    if "product_name" in working.columns:
        working["product_name"] = working["product_name"].astype(str).str.strip()
        working.loc[working["product_name"].isin(["", "nan", "None"]), "product_name"] = pd.NA

    if "time" in working.columns:
        working["time"] = working["time"].astype(str).str.strip()
        working.loc[working["time"].isin(["", "nan", "None"]), "time"] = pd.NA

    return working.reset_index(drop=True)


def read_uploaded_file(file_storage):
    original_filename = file_storage.filename or ""
    filename = secure_filename(original_filename)

    if not filename or "." not in filename:
        raise ValueError("Please choose a CSV or Excel file before continuing.")

    extension = filename.rsplit(".", 1)[1].lower()
    if extension not in ALLOWED_EXTENSIONS:
        raise ValueError("Invalid file type. Please upload a CSV, XLSX, or XLS file.")

    file_bytes = file_storage.read()
    if not file_bytes:
        raise ValueError("The selected file is empty. Please upload a file with sales records.")
    if len(file_bytes) > MAX_UPLOAD_SIZE_BYTES:
        max_mb = MAX_UPLOAD_SIZE_BYTES / (1024 * 1024)
        raise ValueError(f"The selected file is too large for this deployment. Please upload a file under {max_mb:.0f} MB or split it into smaller batches.")

    buffer = BytesIO(file_bytes)

    try:
        if extension == "csv":
            df = pd.read_csv(buffer, encoding="utf-8-sig")
        elif extension in {"xlsx", "xls"}:
            df = pd.read_excel(buffer)
        else:
            raise ValueError("Invalid file type. Please upload a CSV, XLSX, or XLS file.")
    except EmptyDataError:
        raise ValueError("The selected file is empty. Please upload a file with sales records.")
    except ValueError:
        raise ValueError("The selected file could not be read. Please check the file format and try again.")
    except Exception:
        raise ValueError("The selected file could not be read. Please check the file and try again.")

    if df is None or df.empty or len(df.columns) == 0:
        raise ValueError("The selected file has no sales records to check. Please upload a file with transaction rows.")

    return normalize_dataframe(df, get_data_format_preferences()), filename


def format_datetime(value: datetime | None) -> str:
    if not value:
        return "No uploaded sales data yet"
    return value.strftime("%b %d, %Y %I:%M %p")


def format_date(value: pd.Timestamp | datetime | None) -> str:
    if value is None or pd.isna(value):
        return "No date available"
    return pd.to_datetime(value).strftime("%b %d, %Y")


def infer_coverage_period(df: pd.DataFrame | None) -> str:
    if df is None or "date" not in df.columns:
        return "Coverage date unavailable"

    valid_dates = pd.to_datetime(df["date"], errors="coerce").dropna()
    if valid_dates.empty:
        return "Coverage date unavailable"

    return f"{format_date(valid_dates.min())} to {format_date(valid_dates.max())}"


def infer_upload_freshness(processed_at: datetime | None) -> str:
    if not processed_at:
        return "No uploaded sales data yet"

    delta_days = (datetime.now().date() - processed_at.date()).days
    if delta_days <= 0:
        return "Updated today"
    if delta_days == 1:
        return "Updated 1 day ago"
    return f"Updated {delta_days} days ago"


def get_duplicate_subset(df: pd.DataFrame) -> list[str] | None:
    subset = [column for column in ["date", "product_name", "quantity_sold", "time"] if column in df.columns]
    return subset or None


def find_missing_date_gaps(df: pd.DataFrame | None, limit: int = 6) -> list[str]:
    if df is None or "date" not in df.columns:
        return []

    valid_dates = pd.to_datetime(df["date"], errors="coerce").dropna().dt.normalize().sort_values().unique()
    if len(valid_dates) < 2:
        return []

    full_range = pd.date_range(valid_dates[0], valid_dates[-1], freq="D")
    missing = [date for date in full_range if date not in valid_dates]
    if not missing:
        return []

    formatted = [date.strftime("%b %d, %Y") for date in missing[:limit]]
    if len(missing) > limit:
        formatted.append(f"+{len(missing) - limit} more")
    return formatted


def coverage_overlap(selected_df: pd.DataFrame | None, processed_df: pd.DataFrame | None) -> bool:
    if selected_df is None or processed_df is None:
        return False
    if "date" not in selected_df.columns or "date" not in processed_df.columns:
        return False

    selected_dates = set(pd.to_datetime(selected_df["date"], errors="coerce").dropna().dt.normalize().tolist())
    processed_dates = set(pd.to_datetime(processed_df["date"], errors="coerce").dropna().dt.normalize().tolist())
    return len(selected_dates.intersection(processed_dates)) > 0


def refresh_upload_mode_feedback(selected_df: pd.DataFrame | None, upload_mode: str) -> tuple[str | None, str | None]:
    """Keep upload feedback accurate when the selected upload mode changes."""
    if selected_df is None:
        set_upload_feedback(None, None)
        return get_upload_feedback()

    validation = analyze_upload_dataset(selected_df)
    if validation.get("missing_required_columns"):
        missing_columns = ", ".join(validation.get("missing_required_columns") or [])
        set_upload_feedback(
            f"File selected, but these required columns are missing: {missing_columns}. Please update the file or review Data Format Settings column mapping before generating results.",
            "error",
        )
    elif validation.get("valid_rows", 0) <= 0:
        set_upload_feedback(
            "File selected, but no valid sales records were found. Please check the dates, product names, and quantities.",
            "error",
        )
    else:
        has_overlap = coverage_overlap(selected_df, get_processed_dataset())
        if has_overlap and upload_mode == "new":
            set_upload_feedback(
                "Coverage overlap detected. You may append missing records or replace the previous processed sales data.",
                "warning",
            )
        elif has_overlap and upload_mode == "append":
            set_upload_feedback(
                "Append mode selected. StockWise will add missing records while avoiding records already covered by the latest processed data.",
                "info",
            )
        elif has_overlap and upload_mode == "replace":
            set_upload_feedback(
                "Replace mode selected. The selected file will replace the previous processed sales data when results are generated.",
                "info",
            )
        elif validation.get("total_invalid_rows", 0) > 0 or validation.get("duplicate_rows", 0) > 0:
            set_upload_feedback(
                "File selected with review notes. Some rows may be cleaned or excluded before analysis.",
                "warning",
            )
        else:
            set_upload_feedback(None, None)

    return get_upload_feedback()


def analyze_upload_dataset(df: pd.DataFrame | None) -> dict[str, Any]:
    if df is None:
        return {
            "total_rows": 0,
            "total_columns": 0,
            "missing_required_columns": [],
            "invalid_date_rows": 0,
            "invalid_product_rows": 0,
            "invalid_quantity_rows": 0,
            "missing_value_rows": 0,
            "suspicious_value_rows": 0,
            "duplicate_rows": 0,
            "total_invalid_rows": 0,
            "valid_rows": 0,
            "is_valid": False,
            "can_process": False,
            "status_label": "Waiting for file selection",
            "status_class": "neutral",
            "warning_messages": [],
        }

    missing_required = [column for column in REQUIRED_UPLOAD_COLUMNS if column not in df.columns]
    total_rows = len(df)
    total_columns = len(df.columns)

    invalid_date_mask = pd.Series(False, index=df.index)
    invalid_product_mask = pd.Series(False, index=df.index)
    invalid_quantity_mask = pd.Series(False, index=df.index)
    suspicious_value_mask = pd.Series(False, index=df.index)

    if "date" in df.columns:
        invalid_date_mask = df["date"].isna()
    if "product_name" in df.columns:
        invalid_product_mask = df["product_name"].isna() | (df["product_name"].astype(str).str.strip() == "")
    if "quantity_sold" in df.columns:
        invalid_quantity_mask = df["quantity_sold"].isna() | (df["quantity_sold"] <= 0)

    for numeric_column in ["current_stock", "reorder_point", "unit_price"]:
        if numeric_column in df.columns:
            numeric_values = pd.to_numeric(df[numeric_column], errors="coerce")
            suspicious_value_mask |= numeric_values.lt(0).fillna(False)

    duplicate_rows = int(df.duplicated(subset=get_duplicate_subset(df)).sum())
    invalid_union = invalid_date_mask | invalid_product_mask | invalid_quantity_mask
    total_invalid_rows = int(invalid_union.sum())
    valid_rows = max(total_rows - total_invalid_rows, 0)
    missing_value_rows = int(df.isna().any(axis=1).sum()) if total_rows else 0
    suspicious_value_rows = int(suspicious_value_mask.sum())

    can_process = len(missing_required) == 0 and valid_rows > 0
    if can_process and total_invalid_rows == 0 and duplicate_rows == 0 and suspicious_value_rows == 0:
        status_label = "Ready for processing"
        status_class = "success"
    elif can_process:
        status_label = "Ready with warnings"
        status_class = "warning"
    else:
        status_label = "Needs review"
        status_class = "danger"

    warning_messages: list[str] = []
    if missing_required:
        warning_messages.append("Missing required columns must be fixed before processing.")
    if duplicate_rows > 0:
        warning_messages.append("Some duplicate records were detected. They will be reviewed during cleaning.")
    if invalid_quantity_mask.any():
        warning_messages.append("Some rows have invalid quantity values and may be excluded.")
    if invalid_date_mask.any():
        warning_messages.append("Some rows have missing or invalid dates and may be excluded.")
    if invalid_product_mask.any():
        warning_messages.append("Some rows have missing product names and may be excluded.")
    if suspicious_value_rows > 0:
        warning_messages.append("Some stock, price, or reorder values look unusual and should be checked.")
    if missing_value_rows > total_invalid_rows and not missing_required:
        warning_messages.append("Some optional values are missing. StockWise will use safe defaults where possible.")

    return {
        "total_rows": total_rows,
        "total_columns": total_columns,
        "missing_required_columns": missing_required,
        "invalid_date_rows": int(invalid_date_mask.sum()),
        "invalid_product_rows": int(invalid_product_mask.sum()),
        "invalid_quantity_rows": int(invalid_quantity_mask.sum()),
        "missing_value_rows": missing_value_rows,
        "suspicious_value_rows": suspicious_value_rows,
        "duplicate_rows": duplicate_rows,
        "total_invalid_rows": total_invalid_rows,
        "valid_rows": valid_rows,
        "is_valid": len(missing_required) == 0,
        "can_process": can_process,
        "status_label": status_label,
        "status_class": status_class,
        "warning_messages": warning_messages[:4],
    }


def preprocess_dataset(df: pd.DataFrame | None, base_df: pd.DataFrame | None = None, upload_mode: str = "new") -> tuple[pd.DataFrame | None, dict[str, Any]]:
    empty_summary = {
        "rows_before": 0,
        "rows_after": 0,
        "excluded_rows": 0,
        "duplicates_removed": 0,
        "missing_values_handled": 0,
        "normalized_dates": 0,
        "generated_day_features": 0,
        "generated_time_features": 0,
        "generated_payday_flags": 0,
        "mode_label": get_upload_mode_label(upload_mode),
        "status_label": "Waiting for file selection",
        "duplicate_note": "Duplicate records will be checked after a file is selected.",
    }

    if df is None:
        return None, empty_summary

    prefs = get_data_format_preferences()
    duplicate_handling = prefs.get("duplicate_handling", "remove_exact")
    payday_handling = prefs.get("payday_indicator_handling", "auto")
    time_format = prefs.get("data_time_format", "auto")

    working = df.copy()
    validation = analyze_upload_dataset(working)
    if not validation["can_process"]:
        summary = empty_summary.copy()
        summary["rows_before"] = len(working)
        summary["status_label"] = "Cannot process until required issues are fixed"
        return None, summary

    rows_before = len(working)
    duplicate_subset = get_duplicate_subset(working)
    duplicate_rows_initial = int(working.duplicated(subset=duplicate_subset).sum())

    missing_values_handled = 0
    for column, default_value in {"category": "Uncategorized", "unit_type": "Unit"}.items():
        if column not in working.columns:
            working[column] = default_value
            missing_values_handled += len(working)
        else:
            missing_mask = working[column].isna() | (working[column].astype(str).str.strip() == "")
            missing_values_handled += int(missing_mask.sum())
            working.loc[missing_mask, column] = default_value

    for numeric_column in ["current_stock", "reorder_point", "unit_price"]:
        if numeric_column not in working.columns:
            working[numeric_column] = pd.NA

    normalized_dates = 0
    if "date" in working.columns:
        working["date"] = parse_upload_dates(working["date"], prefs.get("data_date_format"))
        normalized_dates = int(working["date"].notna().sum())

    if "product_name" in working.columns:
        working["product_name"] = working["product_name"].astype(str).str.strip()
        working.loc[working["product_name"].isin(["", "nan", "None"]), "product_name"] = pd.NA

    valid_mask = pd.Series(True, index=working.index)
    if "date" in working.columns:
        valid_mask &= working["date"].notna()
    if "product_name" in working.columns:
        valid_mask &= working["product_name"].notna()
    if "quantity_sold" in working.columns:
        valid_mask &= working["quantity_sold"].notna() & (working["quantity_sold"] > 0)

    working = working.loc[valid_mask].copy()

    if "date" in working.columns:
        working["day_of_week"] = working["date"].dt.dayofweek
        working["day_name"] = working["date"].dt.day_name()
        working["is_weekend"] = working["day_of_week"].isin([5, 6]).astype(int)
        if payday_handling == "ignore":
            working["is_payday"] = 0
            generated_payday_flags = 0
        elif payday_handling == "use_column" and "is_payday_period" in working.columns:
            working["is_payday"] = working["is_payday_period"].apply(normalize_boolean_flag).astype(int)
            generated_payday_flags = int(working["is_payday"].sum())
        else:
            working["is_payday"] = working["date"].dt.day.isin([15, 30]).astype(int)
            generated_payday_flags = len(working)
        working["month_period"] = working["date"].dt.to_period("M").astype(str)
        generated_day_features = len(working)
    else:
        generated_day_features = 0
        generated_payday_flags = 0

    if "time" in working.columns:
        if time_format and time_format != "auto":
            parsed_time = pd.to_datetime(working["time"].astype(str), format=time_format, errors="coerce")
        else:
            parsed_time = pd.to_datetime(working["time"].astype(str), errors="coerce")
        working["hour_of_day"] = parsed_time.dt.hour
        generated_time_features = int(parsed_time.notna().sum())
    else:
        working["hour_of_day"] = pd.NA
        generated_time_features = 0

    pre_drop_len = len(working)
    if duplicate_handling == "keep_all":
        duplicates_removed = 0
    else:
        working = working.drop_duplicates(subset=duplicate_subset, keep="last")
        duplicates_removed = pre_drop_len - len(working)

    if upload_mode == "append" and base_df is not None and not base_df.empty:
        combined = pd.concat([base_df.copy(), working], ignore_index=True)
        before_merge = len(combined)
        if duplicate_handling == "keep_all":
            duplicates_removed += 0
        else:
            combined = combined.drop_duplicates(subset=get_duplicate_subset(combined), keep="last")
            duplicates_removed += before_merge - len(combined)
        working = combined.reset_index(drop=True)
        mode_label = get_upload_mode_label("append")
    elif upload_mode == "replace":
        mode_label = get_upload_mode_label("replace")
    else:
        mode_label = get_upload_mode_label("new")

    rows_after = len(working)
    excluded_rows = max(rows_before - rows_after, 0)
    summary = {
        "rows_before": rows_before,
        "rows_after": rows_after,
        "excluded_rows": excluded_rows,
        "duplicates_removed": max(duplicates_removed, 0),
        "missing_values_handled": missing_values_handled,
        "normalized_dates": normalized_dates,
        "generated_day_features": generated_day_features,
        "generated_time_features": generated_time_features,
        "generated_payday_flags": generated_payday_flags,
        "mode_label": mode_label,
        "status_label": "Ready for analysis",
        "duplicate_note": "Duplicate records are removed after invalid rows are excluded, so this number may be lower than the validation count when duplicate rows also contain invalid values." if duplicate_handling != "keep_all" else "Duplicate records were kept because your Data Format Settings are set to keep all records.",
    }
    return working.reset_index(drop=True), summary






def get_upload_wizard_state(selected_df: pd.DataFrame | None, validation_summary: dict[str, Any], processed_filename: str | None, upload_message_type: str | None = None, last_action: str = "") -> dict[str, Any]:
    """Return simple upload wizard step access hints.

    The working Phase 9 behavior used current_step and max_step as the
    source of truth.  Extra flags are still returned for compatibility, but
    they no longer control step navigation in JavaScript.
    """
    file_ready = selected_df is not None
    can_process = bool(validation_summary.get("can_process"))
    results_success = bool(processed_filename and upload_message_type == "success" and (last_action or "").strip().lower() == "process")

    max_step = 1
    if file_ready:
        max_step = 2
    if can_process:
        max_step = 4

    action = (last_action or "").strip().lower()
    current_step = 1
    if action == "process":
        current_step = 4
    elif action in {"finish_upload", "clear_selected"}:
        current_step = 1

    current_step = max(1, min(current_step, max_step))
    return {
        "current_step": current_step,
        "max_step": max_step,
        "file_ready": file_ready,
        "can_process": can_process,
        "results_success": results_success,
    }


def to_int(value, default: int = 0) -> int:
    if value is None or pd.isna(value):
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def to_float(value, default: float = 0.0) -> float:
    if value is None or pd.isna(value):
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def parse_time_for_sql(value, time_format: str | None = None) -> str | None:
    if value is None or pd.isna(value):
        return None

    text = str(value).strip()
    if not text or text.lower() in {"nan", "none"}:
        return None

    selected_time_format = (time_format or get_settings_preferences().get("data_time_format", "auto")).strip()
    if selected_time_format and selected_time_format != "auto":
        parsed = pd.to_datetime(text, format=selected_time_format, errors="coerce")
    else:
        parsed = pd.to_datetime(text, errors="coerce")
    if pd.isna(parsed):
        return None

    return parsed.strftime("%H:%M:%S")


def infer_time_of_day_label(time_text: str | None) -> str | None:
    if not time_text:
        return None

    try:
        hour = int(time_text.split(":")[0])
    except (TypeError, ValueError, IndexError):
        return None

    if 5 <= hour < 12:
        return "morning"
    if 12 <= hour < 17:
        return "afternoon"
    if 17 <= hour < 21:
        return "evening"
    return "night"


def get_or_create_category_id(cursor, category_name: str) -> int:
    safe_name = (category_name or "Uncategorized").strip() or "Uncategorized"

    cursor.execute(
        "SELECT category_id FROM categories WHERE category_name = %s LIMIT 1",
        (safe_name,)
    )
    row = cursor.fetchone()
    if row:
        return row[0]

    cursor.execute(
        "INSERT INTO categories (category_name) VALUES (%s)",
        (safe_name,)
    )
    return cursor.lastrowid




def upsert_inventory(cursor, product_id: int, current_stock: int) -> None:
    cursor.execute(
        "SELECT inventory_id FROM inventory WHERE product_id = %s LIMIT 1",
        (product_id,)
    )
    row = cursor.fetchone()

    if row:
        cursor.execute(
            """
            UPDATE inventory
            SET current_stock = %s,
                last_updated = CURRENT_TIMESTAMP
            WHERE product_id = %s
            """,
            (current_stock, product_id)
        )
    else:
        cursor.execute(
            """
            INSERT INTO inventory (product_id, current_stock)
            VALUES (%s, %s)
            """,
            (product_id, current_stock)
        )


        

# =========================================
# SETTINGS / USER PREFERENCE HELPERS
# =========================================
def get_store_type_options() -> list[str]:
    return list(STORE_TYPE_OPTIONS)


def get_user_role_options() -> list[str]:
    return list(USER_ROLE_OPTIONS)


def get_currency_options() -> list[dict[str, str]]:
    return list(CURRENCY_OPTIONS)


def get_upload_mode_options() -> list[dict[str, str]]:
    return [
        {"key": "new", "label": "New Dataset"},
        {"key": "append", "label": "Append Missing Records"},
        {"key": "replace", "label": "Replace Previous Processed Data"},
    ]


def get_default_product_view_options() -> list[dict[str, str]]:
    return list(DEFAULT_PRODUCT_VIEW_OPTIONS)


def get_yes_no_options() -> list[dict[str, str]]:
    return list(YES_NO_OPTIONS)


def ensure_user_settings_table() -> None:
    """Create the lightweight settings table if it is not present yet."""
    global USER_SETTINGS_SCHEMA_READY
    if USER_SETTINGS_SCHEMA_READY:
        return

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS user_settings (
                setting_id INT(11) NOT NULL AUTO_INCREMENT,
                user_id INT(11) NOT NULL,
                store_name VARCHAR(255) NULL,
                store_type VARCHAR(100) NULL,
                location_area VARCHAR(255) NULL,
                store_logo VARCHAR(255) NULL,
                currency VARCHAR(20) NULL DEFAULT 'PHP',
                default_upload_mode VARCHAR(32) NULL DEFAULT 'new',
                default_time_range VARCHAR(32) NULL DEFAULT '30',
                default_product_view VARCHAR(32) NULL DEFAULT 'needs_attention',
                show_safe_products_dashboard VARCHAR(8) NULL DEFAULT 'no',
                default_report_type VARCHAR(100) NULL DEFAULT 'demand_forecast_summary',
                default_report_period VARCHAR(32) NULL DEFAULT 'last_30_days',
                export_format VARCHAR(20) NULL DEFAULT 'csv',
                include_filtered_rows_only VARCHAR(8) NULL DEFAULT 'yes',
                data_date_format VARCHAR(32) NULL DEFAULT 'auto',
                data_time_format VARCHAR(32) NULL DEFAULT 'auto',
                payday_indicator_handling VARCHAR(32) NULL DEFAULT 'auto',
                duplicate_handling VARCHAR(32) NULL DEFAULT 'remove_exact',
                column_mapping_json TEXT NULL,
                onboarding_completed TINYINT(1) NOT NULL DEFAULT 1,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                PRIMARY KEY (setting_id),
                UNIQUE KEY uq_user_settings_user (user_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """
        )
        add_column_if_missing(cursor, "user_settings", "onboarding_completed", "TINYINT(1) NOT NULL DEFAULT 1", after_column="include_filtered_rows_only")
        add_column_if_missing(cursor, "user_settings", "store_logo", "VARCHAR(255) NULL", after_column="location_area")
        for column_name, column_definition, after_column in [
            ("data_date_format", "VARCHAR(32) NULL DEFAULT 'auto'", "include_filtered_rows_only"),
            ("data_time_format", "VARCHAR(32) NULL DEFAULT 'auto'", "data_date_format"),
            ("payday_indicator_handling", "VARCHAR(32) NULL DEFAULT 'auto'", "data_time_format"),
            ("duplicate_handling", "VARCHAR(32) NULL DEFAULT 'remove_exact'", "payday_indicator_handling"),
            ("column_mapping_json", "TEXT NULL", "duplicate_handling"),
        ]:
            add_column_if_missing(cursor, "user_settings", column_name, column_definition, after_column=after_column)
        conn.commit()
        USER_SETTINGS_SCHEMA_READY = True
    finally:
        cursor.close()
        conn.close()


def get_user_settings_from_db(user_id: int | None) -> dict[str, Any]:
    if not user_id:
        return {}

    try:
        ensure_user_settings_table()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        try:
            cursor.execute(
                """
                SELECT store_name, store_type, location_area, store_logo, currency,
                       default_upload_mode, default_time_range, default_product_view,
                       show_safe_products_dashboard, default_report_type,
                       default_report_period, export_format, include_filtered_rows_only,
                       data_date_format, data_time_format, payday_indicator_handling,
                       duplicate_handling, column_mapping_json, onboarding_completed
                FROM user_settings
                WHERE user_id = %s
                LIMIT 1
                """,
                (user_id,),
            )
            row = cursor.fetchone() or {}
            return {key: value for key, value in row.items() if value is not None}
        finally:
            cursor.close()
            conn.close()
    except Exception:
        return {}


def save_user_settings_to_db(user_id: int, prefs: dict[str, Any]) -> bool:
    if not user_id:
        return False

    try:
        ensure_user_settings_table()
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute(
                """
                INSERT INTO user_settings (
                    user_id, store_name, store_type, location_area, store_logo, currency,
                    default_upload_mode, default_time_range, default_product_view,
                    show_safe_products_dashboard, default_report_type,
                    default_report_period, export_format, include_filtered_rows_only,
                    data_date_format, data_time_format, payday_indicator_handling,
                    duplicate_handling, column_mapping_json
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    store_name = VALUES(store_name),
                    store_type = VALUES(store_type),
                    location_area = VALUES(location_area),
                    store_logo = VALUES(store_logo),
                    currency = VALUES(currency),
                    default_upload_mode = VALUES(default_upload_mode),
                    default_time_range = VALUES(default_time_range),
                    default_product_view = VALUES(default_product_view),
                    show_safe_products_dashboard = VALUES(show_safe_products_dashboard),
                    default_report_type = VALUES(default_report_type),
                    default_report_period = VALUES(default_report_period),
                    export_format = VALUES(export_format),
                    include_filtered_rows_only = VALUES(include_filtered_rows_only),
                    data_date_format = VALUES(data_date_format),
                    data_time_format = VALUES(data_time_format),
                    payday_indicator_handling = VALUES(payday_indicator_handling),
                    duplicate_handling = VALUES(duplicate_handling),
                    column_mapping_json = VALUES(column_mapping_json)
                """,
                (
                    user_id,
                    prefs.get("store_name"),
                    prefs.get("store_type"),
                    prefs.get("location_area"),
                    prefs.get("store_logo"),
                    prefs.get("currency"),
                    prefs.get("default_upload_mode"),
                    prefs.get("default_time_range"),
                    prefs.get("default_product_view"),
                    prefs.get("show_safe_products_dashboard"),
                    prefs.get("default_report_type"),
                    prefs.get("default_report_period"),
                    prefs.get("export_format"),
                    prefs.get("include_filtered_rows_only"),
                    prefs.get("data_date_format"),
                    prefs.get("data_time_format"),
                    prefs.get("payday_indicator_handling"),
                    prefs.get("duplicate_handling"),
                    prefs.get("column_mapping_json"),
                ),
            )
            conn.commit()
            return True
        finally:
            cursor.close()
            conn.close()
    except Exception:
        return False



def create_initial_user_settings(user_id: int) -> bool:
    """Create an incomplete setup row for a newly registered user."""
    if not user_id:
        return False

    try:
        ensure_user_settings_table()
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute(
                """
                INSERT INTO user_settings (
                    user_id, store_name, store_type, location_area, store_logo, currency,
                    default_upload_mode, default_time_range, default_product_view,
                    show_safe_products_dashboard, default_report_type,
                    default_report_period, export_format, include_filtered_rows_only,
                    onboarding_completed
                )
                VALUES (%s, NULL, NULL, NULL, NULL, %s, %s, %s, %s, %s, %s, %s, %s, %s, 0)
                ON DUPLICATE KEY UPDATE
                    onboarding_completed = 0
                """,
                (
                    user_id,
                    DEFAULT_SETTINGS["currency"],
                    DEFAULT_SETTINGS["default_upload_mode"],
                    DEFAULT_SETTINGS["default_time_range"],
                    DEFAULT_SETTINGS["default_product_view"],
                    DEFAULT_SETTINGS["show_safe_products_dashboard"],
                    DEFAULT_SETTINGS["default_report_type"],
                    DEFAULT_SETTINGS["default_report_period"],
                    DEFAULT_SETTINGS["export_format"],
                    DEFAULT_SETTINGS["include_filtered_rows_only"],
                ),
            )
            conn.commit()
            return True
        finally:
            cursor.close()
            conn.close()
    except Exception:
        return False


def user_needs_onboarding(user_id: int | None) -> bool:
    """Return True only for newly created users who still need to finish store setup."""
    if not user_id:
        return False
    if session.get("force_onboarding"):
        return True

    try:
        ensure_user_settings_table()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        try:
            cursor.execute(
                """
                SELECT onboarding_completed
                FROM user_settings
                WHERE user_id = %s
                LIMIT 1
                """,
                (user_id,),
            )
            row = cursor.fetchone()
            if row is None:
                return False
            return int(row.get("onboarding_completed", 1) or 0) == 0
        finally:
            cursor.close()
            conn.close()
    except Exception:
        return False


def validate_personal_setup_form(form) -> tuple[dict[str, str], list[str]]:
    requested_role = normalize_role(form.get("user_role", "Owner"))
    values = {
        "display_name": form.get("display_name", "").strip(),
        "user_role": requested_role if requested_role in USER_ROLE_OPTIONS else "Owner",
    }
    errors: list[str] = []

    if not values["display_name"]:
        errors.append("Please enter your full name.")
    if values["user_role"] not in USER_ROLE_OPTIONS:
        errors.append("Please choose Owner, Store Manager, or Operational Assistant.")

    return values, errors


def save_personal_setup_from_form(user_id: int, form, files=None) -> tuple[bool, str, dict[str, str]]:
    values, errors = validate_personal_setup_form(form)
    uploaded_profile_image = None

    if errors:
        return False, " ".join(errors), values

    try:
        if files and files.get("profile_image") and files.get("profile_image").filename:
            uploaded_profile_image = save_uploaded_image(files.get("profile_image"), PROFILE_UPLOAD_SUBDIR, f"profile_{user_id}")
    except ValueError as exc:
        return False, str(exc), values

    try:
        current_user = get_user_by_id(user_id)
        current_email = (current_user or {}).get("email") or session.get("user_email", "")
        if not current_email:
            return False, "Your session could not be verified. Please log in again.", values

        updated = update_user_profile(
            user_id,
            values["display_name"],
            current_email,
            None,
            values["user_role"],
            profile_image=uploaded_profile_image,
        )
        if not updated:
            return False, "Personal information could not be saved right now.", values

        refreshed_user = get_user_by_id(user_id)
        if refreshed_user:
            session["user_name"] = refreshed_user.get("full_name") or values["display_name"]
            session["user_email"] = refreshed_user.get("email") or current_email
            session["user_position"] = refreshed_user.get("position") or values["user_role"]
            session["user_role"] = normalize_role(refreshed_user.get("role") or refreshed_user.get("position") or values["user_role"])
            session["user_profile_image"] = refreshed_user.get("profile_image") or ""

        return True, "Personal information saved.", values
    except Exception:
        return False, "Personal information could not be saved. Please check the database connection and try again.", values


def validate_store_setup_form(form) -> tuple[dict[str, str], list[str]]:
    values = {
        "store_name": form.get("store_name", "").strip(),
        "store_type": form.get("store_type", "").strip(),
        "location_area": form.get("location_area", "").strip(),
        "currency": form.get("currency", DEFAULT_SETTINGS["currency"]).strip().upper(),
        "default_report_period": form.get("default_report_period", DEFAULT_SETTINGS["default_report_period"]).strip(),
        "default_upload_mode": form.get("default_upload_mode", DEFAULT_SETTINGS["default_upload_mode"]).strip(),
        "has_store_logo": form.get("has_store_logo", "").strip().lower(),
    }
    errors: list[str] = []

    if not values["store_name"]:
        errors.append("Please enter your store name.")
    if values["store_type"] not in STORE_TYPE_OPTIONS:
        errors.append("Please choose a valid store type.")
    if not values["location_area"]:
        errors.append("Please enter your store location or area.")
    if values["currency"] not in {option["key"] for option in CURRENCY_OPTIONS}:
        errors.append("Please choose a valid currency.")

    allowed_report_periods = {option["key"] for option in get_report_period_options()}
    if values["default_report_period"] not in allowed_report_periods:
        values["default_report_period"] = DEFAULT_SETTINGS["default_report_period"]
    if values["default_upload_mode"] not in {"new", "append", "replace"}:
        values["default_upload_mode"] = DEFAULT_SETTINGS["default_upload_mode"]
    if values["has_store_logo"] not in {"yes", "no"}:
        values["has_store_logo"] = ""

    return values, errors


def save_store_setup_from_form(user_id: int, form, files=None) -> tuple[bool, str, dict[str, str]]:
    values, errors = validate_store_setup_form(form)
    store_logo_path = None

    if errors:
        return False, " ".join(errors), values

    try:
        if values.get("has_store_logo") == "yes":
            logo_file = files.get("store_logo") if files else None
            if logo_file and logo_file.filename:
                store_logo_path = save_uploaded_image(logo_file, STORE_LOGO_UPLOAD_SUBDIR, f"store_logo_{user_id}")
            else:
                existing_logo = get_user_settings_from_db(user_id).get("store_logo", "")
                if existing_logo:
                    store_logo_path = existing_logo
                else:
                    return False, "Please upload your store logo or choose No for now.", values
    except ValueError as exc:
        return False, str(exc), values

    try:
        ensure_user_settings_table()
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute(
                """
                INSERT INTO user_settings (
                    user_id, store_name, store_type, location_area, store_logo, currency,
                    default_upload_mode, default_time_range, default_product_view,
                    show_safe_products_dashboard, default_report_type,
                    default_report_period, export_format, include_filtered_rows_only,
                    onboarding_completed
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 1)
                ON DUPLICATE KEY UPDATE
                    store_name = VALUES(store_name),
                    store_type = VALUES(store_type),
                    location_area = VALUES(location_area),
                    store_logo = VALUES(store_logo),
                    currency = VALUES(currency),
                    default_upload_mode = VALUES(default_upload_mode),
                    default_report_period = VALUES(default_report_period),
                    onboarding_completed = 1
                """,
                (
                    user_id,
                    values["store_name"],
                    values["store_type"],
                    values["location_area"],
                    store_logo_path,
                    values["currency"],
                    values["default_upload_mode"],
                    DEFAULT_SETTINGS["default_time_range"],
                    DEFAULT_SETTINGS["default_product_view"],
                    DEFAULT_SETTINGS["show_safe_products_dashboard"],
                    DEFAULT_SETTINGS["default_report_type"],
                    values["default_report_period"],
                    DEFAULT_SETTINGS["export_format"],
                    DEFAULT_SETTINGS["include_filtered_rows_only"],
                ),
            )
            conn.commit()
        finally:
            cursor.close()
            conn.close()

        session.pop("force_onboarding", None)
        session["settings_preferences"] = {
            **session.get("settings_preferences", {}),
            "store_name": values["store_name"],
            "store_type": values["store_type"],
            "location_area": values["location_area"],
            "store_logo": store_logo_path or "",
            "currency": values["currency"],
            "default_upload_mode": values["default_upload_mode"],
            "default_report_period": values["default_report_period"],
            "onboarding_completed": 1,
        }
        add_notification(user_id, "Store setup completed", "Your store workspace is ready. You can now upload sales data or review your dashboard.", "setup", store_id=get_store_id_for_user(user_id), target_role="Owner")
        return True, "Store information saved.", values
    except Exception:
        return False, "Store information could not be saved right now. Please check the database connection and try again.", values


def save_store_logo_setup_from_form(user_id: int, form, files=None) -> tuple[bool, str, str]:
    logo_choice = form.get("has_store_logo", "no").strip().lower()
    if logo_choice not in {"yes", "no"}:
        logo_choice = "no"

    store_logo_path = None
    if logo_choice == "yes":
        try:
            logo_file = files.get("store_logo") if files else None
            if logo_file and logo_file.filename:
                store_logo_path = save_uploaded_image(logo_file, STORE_LOGO_UPLOAD_SUBDIR, f"store_logo_{user_id}")
            else:
                existing_logo = get_user_settings_from_db(user_id).get("store_logo", "")
                if existing_logo:
                    store_logo_path = existing_logo
                else:
                    return False, "Please upload your store logo or choose No.", logo_choice
        except ValueError as exc:
            return False, str(exc), logo_choice

    try:
        ensure_user_settings_table()
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute(
                """
                UPDATE user_settings
                SET store_logo = %s,
                    onboarding_completed = 1
                WHERE user_id = %s
                """,
                (store_logo_path, user_id),
            )
            if cursor.rowcount == 0:
                cursor.execute(
                    """
                    INSERT INTO user_settings (
                        user_id, store_name, store_type, location_area, currency,
                        default_upload_mode, default_time_range, default_product_view,
                        show_safe_products_dashboard, default_report_type,
                        default_report_period, export_format, include_filtered_rows_only,
                        store_logo, onboarding_completed
                    )
                    VALUES (%s, NULL, NULL, NULL, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 1)
                    """,
                    (
                        user_id,
                        DEFAULT_SETTINGS["currency"],
                        DEFAULT_SETTINGS["default_upload_mode"],
                        DEFAULT_SETTINGS["default_time_range"],
                        DEFAULT_SETTINGS["default_product_view"],
                        DEFAULT_SETTINGS["show_safe_products_dashboard"],
                        DEFAULT_SETTINGS["default_report_type"],
                        DEFAULT_SETTINGS["default_report_period"],
                        DEFAULT_SETTINGS["export_format"],
                        DEFAULT_SETTINGS["include_filtered_rows_only"],
                        store_logo_path,
                    ),
                )
            conn.commit()
        finally:
            cursor.close()
            conn.close()

        session.pop("force_onboarding", None)
        session["settings_preferences"] = {
            **session.get("settings_preferences", {}),
            "store_logo": store_logo_path or "",
            "onboarding_completed": 1,
        }
        add_notification(user_id, "Store setup completed", "Your store workspace is ready. You can now upload sales data or review your dashboard.", "setup", store_id=get_store_id_for_user(user_id), target_role="Owner")
        return True, "Store logo preference saved.", logo_choice
    except Exception:
        return False, "Store logo preference could not be saved right now. Please check the database connection and try again.", logo_choice

def normalize_yes_no(value: str | None, default: str = "no") -> str:
    normalized = str(value or default).strip().lower()
    return normalized if normalized in {"yes", "no"} else default


def get_settings_preferences() -> dict[str, Any]:
    stored = session.get("settings_preferences", {}) if session else {}
    db_preferences = get_user_settings_from_db(get_current_user_id()) if session else {}
    merged = {**DEFAULT_SETTINGS, **stored, **db_preferences}
    merged["display_name"] = session.get("user_name", "Admin") if session else "Admin"
    merged["email"] = session.get("user_email", "admin@stockwise.local") if session else "admin@stockwise.local"
    merged["role"] = get_session_role() if session else "Owner"
    merged["profile_image"] = session.get("user_profile_image", "") if session else ""
    merged["store_logo"] = merged.get("store_logo", "") or ""
    data_format_preferences = get_data_format_preferences(merged)
    merged.update(data_format_preferences)
    merged["default_forecast_range"] = merged.get("default_time_range") or merged.get("default_forecast_range") or "30"
    return merged


def update_current_user_profile(name: str, email: str, new_password: str | None = None, position: str | None = None, profile_image: str | None = None) -> bool:
    user_id = session.get("user_id")
    if not user_id:
        return False

    cleaned_position = normalize_role(position or session.get("user_position") or "Owner")
    success = update_user_profile(user_id, name, email, new_password, cleaned_position, profile_image=profile_image)
    if not success:
        return False

    updated_user = get_user_by_id(user_id)
    if updated_user:
        session["user_name"] = updated_user["full_name"]
        session["user_email"] = updated_user["email"]
        session["user_position"] = normalize_role(updated_user.get("position") or "Owner")
        session["user_profile_image"] = updated_user.get("profile_image") or ""

    return True


def save_settings_from_form(form, files=None) -> tuple[str, str]:
    user_id = get_current_user_id()
    if not user_id:
        return "Your session could not be verified. Please log in again.", "error"

    settings_section = normalize_settings_section_key(form.get("settings_section", "profile"))
    if not can_save_settings_section(settings_section):
        add_activity_log("Access denied", "Settings", "Blocked")
        return "Your account role cannot update that settings section.", "error"

    current_preferences = get_settings_preferences()
    display_name = form.get("display_name", current_preferences.get("display_name", "")).strip() or session.get("user_name", "Admin")
    email = form.get("email", current_preferences.get("email", "")).strip() or session.get("user_email", "admin@stockwise.local")
    store_name = form.get("store_name", current_preferences.get("store_name", DEFAULT_SETTINGS["store_name"])).strip() or DEFAULT_SETTINGS["store_name"]
    store_type = form.get("store_type", current_preferences.get("store_type", DEFAULT_SETTINGS["store_type"])).strip()
    location_area = form.get("location_area", current_preferences.get("location_area", DEFAULT_SETTINGS["location_area"])).strip() or DEFAULT_SETTINGS["location_area"]
    user_role = get_session_role()
    currency = form.get("currency", current_preferences.get("currency", DEFAULT_SETTINGS["currency"])).strip().upper()
    default_upload_mode = form.get("default_upload_mode", current_preferences.get("default_upload_mode", DEFAULT_SETTINGS["default_upload_mode"])).strip().lower()
    default_time_range = form.get("default_time_range", current_preferences.get("default_time_range", DEFAULT_SETTINGS["default_time_range"])).strip()
    default_product_view = form.get("default_product_view", current_preferences.get("default_product_view", DEFAULT_SETTINGS["default_product_view"])).strip().lower()
    show_safe_products_dashboard = normalize_yes_no(form.get("show_safe_products_dashboard", current_preferences.get("show_safe_products_dashboard")), DEFAULT_SETTINGS["show_safe_products_dashboard"])
    default_report_type = form.get("default_report_type", current_preferences.get("default_report_type", DEFAULT_SETTINGS["default_report_type"])).strip()
    default_report_period = form.get("default_report_period", current_preferences.get("default_report_period", DEFAULT_SETTINGS["default_report_period"])).strip()
    export_format = form.get("export_format", current_preferences.get("export_format", DEFAULT_SETTINGS["export_format"])).strip().lower()
    include_filtered_rows_only = normalize_yes_no(form.get("include_filtered_rows_only", current_preferences.get("include_filtered_rows_only")), DEFAULT_SETTINGS["include_filtered_rows_only"])
    data_date_format = form.get("data_date_format", current_preferences.get("data_date_format", DEFAULT_SETTINGS["data_date_format"])).strip()
    data_time_format = form.get("data_time_format", current_preferences.get("data_time_format", DEFAULT_SETTINGS["data_time_format"])).strip()
    payday_indicator_handling = form.get("payday_indicator_handling", current_preferences.get("payday_indicator_handling", DEFAULT_SETTINGS["payday_indicator_handling"])).strip()
    duplicate_handling = form.get("duplicate_handling", current_preferences.get("duplicate_handling", DEFAULT_SETTINGS["duplicate_handling"])).strip()
    submitted_mapping = parse_column_mapping_json(current_preferences.get("column_mapping_json"))
    for field in DATA_FORMAT_SYSTEM_FIELDS:
        form_key = f"column_map_{field['key']}"
        if form_key in form:
            submitted_mapping[field["key"]] = form.get(form_key, "").strip() or field["default_column"]
    column_mapping_json = json.dumps(submitted_mapping)
    current_password = form.get("current_password", "").strip()
    new_password = form.get("new_password", "").strip()
    confirm_password = form.get("confirm_password", "").strip()

    uploaded_profile_image = None
    uploaded_store_logo = None
    try:
        if files and files.get("profile_image") and files.get("profile_image").filename:
            uploaded_profile_image = save_uploaded_image(files.get("profile_image"), PROFILE_UPLOAD_SUBDIR, f"profile_{user_id}")
        if files and files.get("store_logo") and files.get("store_logo").filename:
            uploaded_store_logo = save_uploaded_image(files.get("store_logo"), STORE_LOGO_UPLOAD_SUBDIR, f"store_logo_{user_id}")
    except ValueError as exc:
        return str(exc), "error"

    if "@" not in email or "." not in email.split("@")[-1]:
        return "Please enter a valid email address.", "error"

    if store_type not in STORE_TYPE_OPTIONS:
        store_type = DEFAULT_SETTINGS["store_type"]
    user_role = normalize_role(user_role)
    if currency not in {option["key"] for option in CURRENCY_OPTIONS}:
        currency = DEFAULT_SETTINGS["currency"]
    if default_upload_mode not in {"new", "append", "replace"}:
        default_upload_mode = DEFAULT_SETTINGS["default_upload_mode"]
    if default_time_range not in {"7", "14", "30", "monthly"}:
        default_time_range = DEFAULT_SETTINGS["default_time_range"]
    if default_product_view not in {"needs_attention", "all_products"}:
        default_product_view = DEFAULT_SETTINGS["default_product_view"]
    if default_report_type not in {option["key"] for option in get_report_type_options()}:
        default_report_type = DEFAULT_SETTINGS["default_report_type"]
    allowed_report_periods = {option["key"] for option in get_report_period_options()}
    if default_report_period not in allowed_report_periods:
        default_report_period = DEFAULT_SETTINGS["default_report_period"]
    if export_format != "csv":
        export_format = "csv"
    if data_date_format not in {option["key"] for option in get_data_date_format_options()}:
        data_date_format = DEFAULT_SETTINGS["data_date_format"]
    if data_time_format not in {option["key"] for option in get_data_time_format_options()}:
        data_time_format = DEFAULT_SETTINGS["data_time_format"]
    if payday_indicator_handling not in {option["key"] for option in get_payday_handling_options()}:
        payday_indicator_handling = DEFAULT_SETTINGS["payday_indicator_handling"]
    if duplicate_handling not in {option["key"] for option in get_duplicate_handling_options()}:
        duplicate_handling = DEFAULT_SETTINGS["duplicate_handling"]

    password_to_save: str | None = None
    password_change_requested = settings_section == "security" or current_password or new_password or confirm_password
    if password_change_requested:
        if not current_password:
            return "Please enter your current password before setting a new password.", "error"
        if not new_password:
            return "Please enter your new password.", "error"
        if not confirm_password:
            return "Please confirm your new password.", "error"

        current_user = get_user_by_id(user_id)
        if not current_user or not check_password_hash(current_user.get("password_hash", ""), current_password):
            return "Current password is incorrect.", "error"

        if new_password != confirm_password:
            return "New password and confirmation do not match.", "error"
        password_ok, password_error = validate_strong_password(new_password)
        if not password_ok:
            return password_error, "error"
        password_to_save = new_password

    prefs = {
        "store_name": store_name,
        "store_type": store_type,
        "location_area": location_area,
        "store_logo": uploaded_store_logo or current_preferences.get("store_logo", ""),
        "currency": currency,
        "default_upload_mode": default_upload_mode,
        "default_time_range": default_time_range,
        "default_product_view": default_product_view,
        "show_safe_products_dashboard": show_safe_products_dashboard,
        "default_report_type": default_report_type,
        "default_report_period": default_report_period,
        "export_format": "csv",
        "include_filtered_rows_only": include_filtered_rows_only,
        "data_date_format": data_date_format,
        "data_time_format": data_time_format,
        "payday_indicator_handling": payday_indicator_handling,
        "duplicate_handling": duplicate_handling,
        "column_mapping_json": column_mapping_json,
        "default_forecast_range": default_time_range,
    }

    profile_updated = update_current_user_profile(display_name, email, password_to_save, position=user_role, profile_image=uploaded_profile_image)
    if not profile_updated:
        return "That email is already being used by another account, or your session has expired.", "error"

    session["settings_preferences"] = prefs
    preferences_saved = save_user_settings_to_db(user_id, prefs)
    if not preferences_saved:
        return "Profile saved, but preferences could not be stored permanently. Please check the database connection and try again.", "error"

    if settings_section == "data_format":
        add_activity_log("Data format settings changed", "Settings", "Success")
    else:
        add_activity_log("Settings update", "Settings", "Success")
    return "Settings saved successfully.", "success"


def get_report_period_options() -> list[dict[str, str]]:
    return [
        {"key": "last_7_days", "label": "Last 7 Days"},
        {"key": "last_30_days", "label": "Last 30 Days"},
        {"key": "last_90_days", "label": "Last 90 Days"},
        {"key": "all_data", "label": "All Processed Records"},
    ]


def get_report_type_options() -> list[dict[str, str]]:
    return [
        {"key": "demand_forecast_summary", "label": "Predicted Demand Summary"},
        {"key": "stockout_risk_summary", "label": "Stockout Risk Summary"},
        {"key": "product_risk_report", "label": "Product Risk Report"},
        {"key": "monthly_sales_and_demand_summary", "label": "Sales and Demand Summary"},
    ]


def filter_dataset_for_report(
    df: pd.DataFrame | None,
    period_key: str = "last_30_days",
    product_name: str = "",
    product_names: list[str] | None = None,
    category_filter: str = "",
) -> pd.DataFrame | None:
    if df is None:
        return None
    working = df.copy()
    if "date" in working.columns and not working["date"].dropna().empty:
        valid_dates = pd.to_datetime(working["date"], errors="coerce")
        max_date = valid_dates.dropna().max()
        if pd.notna(max_date):
            if period_key == "last_7_days":
                cutoff = max_date.normalize() - pd.Timedelta(days=6)
                working = working[valid_dates >= cutoff]
            elif period_key == "last_30_days":
                cutoff = max_date.normalize() - pd.Timedelta(days=29)
                working = working[valid_dates >= cutoff]
            elif period_key == "last_90_days":
                cutoff = max_date.normalize() - pd.Timedelta(days=89)
                working = working[valid_dates >= cutoff]

    if "category" not in working.columns:
        working["category"] = "Seasonal / Miscellaneous Items"
    working["category"] = working["category"].fillna("Seasonal / Miscellaneous Items").apply(standardize_product_category)

    if category_filter:
        safe_category = standardize_product_category(category_filter)
        working = working[working["category"] == safe_category]

    selected_products = [str(name).strip() for name in (product_names or []) if str(name).strip()]
    if product_name and not selected_products:
        selected_products = [product_name]
    if selected_products and "product_name" in working.columns:
        working = working[working["product_name"].isin(selected_products)]
    return working.reset_index(drop=True)
def _build_ranked_bar_colors(values: list[float], palette: str = "forecast") -> list[str]:
    """Return soft alternating StockWise yellows for report bar charts.

    The report overview should not imply ranking through darker colors.  The
    values already rank the bars, so the colors stay visually calm and alternate
    only by position.
    """
    if not values:
        return []
    colors = ['#FFE58A', '#FFF1B8']
    return [colors[index % len(colors)] for index, _ in enumerate(values)]


def _normalize_report_risk_level(value: Any) -> str:
    text = str(value or "Unavailable").strip().lower()
    if "high" in text:
        return "High"
    if "moderate" in text:
        return "Moderate"
    if "low" in text:
        return "Low"
    if "safe" in text:
        return "Safe"
    if "review" in text:
        return "Needs review"
    return "Unavailable"


def _build_risk_bar_colors(rows: list[dict[str, Any]]) -> list[str]:
    palette = {
        "High": "#FFA1A1",
        "Moderate": "#FFEA92",
        "Low": "#BCFF9F",
        "Safe": "#BCFF9F",
        "Needs review": "#EAEAEA",
        "Unavailable": "#EAEAEA",
    }
    return [palette.get(_normalize_report_risk_level(row.get("risk_level")), "#EAEAEA") for row in rows]


def _build_risk_bar_borders(rows: list[dict[str, Any]]) -> list[str]:
    palette = {
        "High": "#F28B8B",
        "Moderate": "#E5C75D",
        "Low": "#9DD982",
        "Safe": "#9DD982",
        "Needs review": "#D6D6D6",
        "Unavailable": "#D6D6D6",
    }
    return [palette.get(_normalize_report_risk_level(row.get("risk_level")), "#D6D6D6") for row in rows]


def get_report_chart_data(report_rows: list[dict[str, Any]], report_type: str) -> dict[str, Any]:
    base = {
        'chart_type': 'bar',
        'chart_labels': [],
        'chart_datasets': [],
        'chart_options': {},
        'chart_message': None,
        'chart_explanation': '',
        'chart_summary': None,
    }
    if not report_rows:
        base['chart_message'] = 'No chart data available yet.'
        base['chart_explanation'] = ''
        return base

    def _to_number(value: Any) -> float | None:
        try:
            if value is None:
                return None
            if isinstance(value, str) and not value.strip():
                return None
            return float(value)
        except Exception:
            return None

    def _short_name(value: Any, max_chars: int = 22) -> str:
        text = str(value or 'Unnamed product').strip() or 'Unnamed product'
        return text if len(text) <= max_chars else f"{text[:max_chars - 1]}…"

    def _top_rows(rows: list[dict[str, Any]], key_func, limit: int = 10) -> list[dict[str, Any]]:
        return sorted(rows, key=key_func, reverse=True)[:limit]

    common_options = {
        'plugins': {
            'legend': {'position': 'bottom', 'labels': {'boxWidth': 10, 'usePointStyle': True}},
            'tooltip': {'displayColors': True},
        },
        'scales': {
            'y': {'beginAtZero': True, 'grid': {'color': '#f0ebe1'}},
            'x': {
                'grid': {'display': False},
                'ticks': {'maxRotation': 0, 'minRotation': 0, 'autoSkip': False},
            },
        },
    }

    if report_type in {'stockout_risk_summary', 'product_risk_report'}:
        risk_score = {'High': 3, 'Moderate': 2, 'Low': 1, 'Safe': 1, 'Needs review': 0, 'Unavailable': 0}
        rows = _top_rows(
            report_rows,
            lambda row: (
                risk_score.get(_normalize_report_risk_level(row.get('risk_level')), 0),
                _to_number(row.get('forecast_demand')) or 0,
                _to_number(row.get('total_monthly_sales_qty')) or 0,
            ),
        )
        labels = [_short_name(row.get('product_name')) for row in rows]
        values = [risk_score.get(_normalize_report_risk_level(row.get('risk_level')), 0) for row in rows]
        if not any(values):
            base['chart_message'] = 'No chart data available yet.'
            base['chart_explanation'] = 'This chart compares products by stockout risk when report data is available.'
            return base
        colors = _build_risk_bar_colors(rows)
        borders = _build_risk_bar_borders(rows)
        return {
            'chart_type': 'bar',
            'chart_labels': labels,
            'chart_datasets': [{
                'label': 'Stockout Risk',
                'data': values,
                'backgroundColor': colors,
                'borderColor': borders,
                'borderWidth': 0,
                'borderRadius': 10,
                'barThickness': 18,
            }],
            'chart_options': {
                'indexAxis': 'y',
                'plugins': {'legend': {'display': False}, 'tooltip': {'displayColors': True}},
                'scales': {
                    'x': {
                        'beginAtZero': True,
                        'min': 0,
                        'max': 3,
                        'ticks': {'stepSize': 1},
                        'title': {'display': True, 'text': 'Low to high stockout risk'},
                        'grid': {'color': '#f0ebe1'},
                    },
                    'y': {'grid': {'display': False}, 'ticks': {'autoSkip': False}},
                },
            },
            'chart_message': None,
            'chart_explanation': 'Top products by stockout risk.',
            'chart_summary': 'Higher values may need earlier review.',
        }

    if report_type == 'monthly_sales_and_demand_summary':
        rows = _top_rows(report_rows, lambda row: _to_number(row.get('total_sales_value')) or 0)
        labels = [_short_name(row.get('product_name')) for row in rows]
        sales_values = [_to_number(row.get('total_sales_value')) or 0 for row in rows]
        demand_values = [_to_number(row.get('average_daily_demand')) or 0 for row in rows]
        if not any(sales_values) and not any(demand_values):
            base['chart_message'] = 'No chart data available yet.'
            base['chart_explanation'] = 'This chart compares product sales value and average daily demand when report data is available.'
            return base
        return {
            'chart_type': 'bar',
            'chart_labels': labels,
            'chart_datasets': [
                {
                    'label': 'Sales Value',
                    'data': sales_values,
                    'backgroundColor': '#FFE58A',
                    'borderColor': '#EAD98A',
                    'borderWidth': 0,
                    'borderRadius': 10,
                    'yAxisID': 'y',
                },
                {
                    'label': 'Average Daily Demand',
                    'data': demand_values,
                    'backgroundColor': '#FFF1B8',
                    'borderColor': '#EAD98A',
                    'borderWidth': 0,
                    'borderRadius': 10,
                    'yAxisID': 'y1',
                },
            ],
            'chart_options': {
                'plugins': {'legend': {'position': 'bottom', 'labels': {'boxWidth': 10, 'usePointStyle': True}}},
                'scales': {
                    'y': {
                        'beginAtZero': True,
                        'position': 'left',
                        'title': {'display': True, 'text': 'Sales value (₱)'},
                        'grid': {'color': '#f0ebe1'},
                    },
                    'y1': {
                        'beginAtZero': True,
                        'position': 'right',
                        'title': {'display': True, 'text': 'Average daily demand'},
                        'grid': {'drawOnChartArea': False},
                    },
                    'x': {'grid': {'display': False}, 'ticks': {'maxRotation': 0, 'minRotation': 0, 'autoSkip': False}},
                },
            },
            'chart_message': None,
            'chart_explanation': 'Top products by sales value and demand.',
            'chart_summary': '',
        }

    rows = _top_rows(report_rows, lambda row: _to_number(row.get('forecast_demand')) or 0)
    labels = []
    values = []
    for row in rows:
        value = _to_number(row.get('forecast_demand'))
        if value is None:
            continue
        labels.append(_short_name(row.get('product_name')))
        values.append(value)
    if not values:
        base['chart_message'] = 'No chart data available yet.'
        base['chart_explanation'] = 'This chart compares predicted demand for products in the selected report.'
        return base
    return {
        'chart_type': 'bar',
        'chart_labels': labels,
        'chart_datasets': [{
            'label': 'Predicted Demand',
            'data': values,
            'backgroundColor': _build_ranked_bar_colors(values, 'forecast'),
            'borderColor': '#EAD98A',
            'borderWidth': 0,
            'borderRadius': 10,
            'barPercentage': 0.68,
            'categoryPercentage': 0.72,
        }],
        'chart_options': common_options | {
            'scales': {
                'y': {
                    'beginAtZero': True,
                    'title': {'display': True, 'text': 'Predicted demand units'},
                    'grid': {'color': '#f0ebe1'},
                },
                'x': {'grid': {'display': False}, 'ticks': {'maxRotation': 0, 'minRotation': 0, 'autoSkip': False}},
            },
        },
        'chart_message': None,
        'chart_explanation': 'This bar chart compares up to 10 products by predicted demand in the current report.',
        'chart_summary': 'Taller bars indicate products with higher predicted demand in this view.',
    }


def get_report_context(
    df: pd.DataFrame | None,
    report_type: str,
    period_key: str,
    product_filter: str = "",
    risk_filter: str = "",
    category_filter: str = "",
    product_filters: list[str] | None = None,
) -> dict[str, Any]:
    allowed_report_types = {option["key"] for option in get_report_type_options()}
    allowed_periods = {option["key"] for option in get_report_period_options()}
    if report_type not in allowed_report_types:
        report_type = "demand_forecast_summary"
    if period_key not in allowed_periods:
        period_key = "last_30_days"

    selected_products = [str(name).strip() for name in (product_filters or []) if str(name).strip()]
    if product_filter and not selected_products:
        selected_products = [product_filter]
    selected_products = sorted(set(selected_products), key=str.casefold)

    cache_key = _analytics_cache_key(df) + (
        'report_context',
        report_type,
        period_key,
        category_filter or '',
        risk_filter or '',
        tuple(selected_products),
    )
    report_context_cache = _state_get('report_context_cache') or {}
    if cache_key in report_context_cache:
        cached = report_context_cache[cache_key]
        cached_copy = cached.copy()
        cached_copy['rows'] = [row.copy() for row in cached.get('rows', [])]
        cached_copy['summary'] = cached.get('summary', {}).copy()
        cached_copy['filters'] = cached.get('filters', {}).copy()
        return cached_copy

    has_processed_data = df is not None and not df.empty
    filtered_df = filter_dataset_for_report(
        df,
        period_key=period_key,
        product_names=selected_products,
        category_filter=category_filter,
    )
    filtered_has_records = filtered_df is not None and not filtered_df.empty
    rows = get_report_rows(filtered_df)
    if risk_filter:
        rows = [row for row in rows if row.get("risk_level") == risk_filter]

    total_sales_volume = round(sum(row.get("total_sales_value", 0) for row in rows), 2)
    most_sold_product = max(rows, key=lambda x: x.get("total_monthly_sales_qty", 0))["product_name"] if rows else "No report data available yet"
    highest_risk_product = next((row["product_name"] for row in rows if row.get("risk_level") == "High"), "No report data available yet")
    average_daily_demand = round(sum(float(row.get("average_daily_demand") or 0) for row in rows), 2)
    period_label = next((opt["label"] for opt in get_report_period_options() if opt["key"] == period_key), "Last 30 Days")
    type_label = next((opt["label"] for opt in get_report_type_options() if opt["key"] == report_type), "Predicted Demand Summary")

    category_label = standardize_product_category(category_filter) if category_filter else "All categories"
    risk_label = f"{risk_filter} risk" if risk_filter else "All stockout risks"
    product_label = f"{len(selected_products)} selected product" + ("s" if len(selected_products) != 1 else "") if selected_products else "All products"
    has_specific_filters = (
        report_type != "demand_forecast_summary"
        or period_key != "last_30_days"
        or bool(category_filter)
        or bool(risk_filter)
        or bool(selected_products)
    )
    if has_specific_filters:
        filter_summary = "Report view: " + " · ".join([type_label, period_label, category_label, risk_label, product_label])
    else:
        filter_summary = "Report view: Latest processed results · Last 30 Days · All categories · All stockout risks · All products"

    empty_message = "No report data available yet."
    if has_processed_data and not rows:
        empty_message = "No report records match the selected filters."

    context = {
        "rows": rows,
        "summary": {
            "total_sales_volume": total_sales_volume,
            "most_sold_product": most_sold_product,
            "highest_risk_product": highest_risk_product,
            "average_daily_demand": average_daily_demand,
        },
        "period_label": period_label,
        "type_label": type_label,
        "coverage_period": infer_coverage_period(filtered_df),
        "chart": get_report_chart_data(rows, report_type),
        "has_processed_data": has_processed_data,
        "filtered_has_records": filtered_has_records,
        "empty_message": empty_message,
        "filter_summary": filter_summary,
        "filters": {
            "report_type": report_type,
            "period_key": period_key,
            "product_filter": product_filter,
            "product_filters": selected_products,
            "category_filter": standardize_product_category(category_filter) if category_filter else "",
            "risk_filter": risk_filter,
        },
    }
    report_context_cache[cache_key] = context
    _state_set('report_context_cache', report_context_cache)
    return context
# =========================================
# SHARED DATA DERIVATION HELPERS
# =========================================
def get_product_names(df: pd.DataFrame | None) -> list[str]:
    if df is None or "product_name" not in df.columns:
        return []
    return sorted(df["product_name"].dropna().astype(str).unique().tolist())


def get_available_risk_levels(rows: list[dict[str, Any]] | None) -> list[str]:
    order = {'High': 0, 'Moderate': 1, 'Low': 2, 'Unavailable': 3}
    values = sorted({str(row.get('risk_level') or 'Unavailable') for row in (rows or []) if str(row.get('risk_level') or '').strip()}, key=lambda item: order.get(item, 9))
    return values













def _calculate_stock_cover_days(current_stock: Any, average_daily_demand: Any) -> float | None:
    stock = to_float(current_stock)
    demand = to_float(average_daily_demand)
    if stock is None or demand is None or demand <= 0:
        return None
    return round(stock / demand, 2)


def _calibrate_risk_outcome(
    *,
    probability: Any,
    current_stock: Any,
    reorder_point: Any,
    forecast_demand: Any,
    average_daily_demand: Any,
    trend: str,
    forecast_status: str | None,
    factors: list[str] | None = None,
    risk_note: str | None = None,
) -> dict[str, Any]:
    probability_value = to_float(probability)
    stock = to_float(current_stock)
    reorder = to_float(reorder_point)
    forecast_qty = to_float(forecast_demand)
    avg_daily = to_float(average_daily_demand, 0.0) or 0.0
    stock_cover_days = _calculate_stock_cover_days(stock, avg_daily)
    forecast_ready = (forecast_status or '').strip().lower() == 'completed' and forecast_qty is not None

    if stock is None or reorder is None:
        return {
            'risk_level': 'Unavailable',
            'reason': 'Stock details need review.',
            'suggested_action': 'Gather more sales history',
            'priority': 'Low',
            'stock_cover_days': stock_cover_days,
            'forecast_pressure': False,
            'reorder_pressure': False,
            'near_reorder': False,
        }

    reorder_pressure = stock <= reorder
    near_reorder = stock <= (reorder * 1.15)
    forecast_pressure = forecast_ready and forecast_qty > stock
    approaching_pressure = forecast_ready and forecast_qty > (stock * 0.8)
    low_cover = stock_cover_days is not None and stock_cover_days < 3
    tight_cover = stock_cover_days is not None and stock_cover_days < 5
    comfortable_cover = stock_cover_days is not None and stock_cover_days >= 8
    strong_probability = probability_value is not None and probability_value >= 0.78
    moderate_probability = probability_value is not None and probability_value >= 0.56
    low_probability = probability_value is not None and probability_value < 0.33
    rising_demand = (trend or '').strip().lower() == 'rising'

    reasons = []
    if reorder_pressure:
        reasons.append('Stock is near the reorder point.')
    elif near_reorder:
        reasons.append('Stock is close to the reorder point.')
    if forecast_pressure:
        reasons.append('Predicted demand may be higher than available stock.')
    elif approaching_pressure:
        reasons.append('Predicted demand may use much of the available stock.')
    if low_cover:
        reasons.append('Available stock may only cover a few days of demand.')
    elif tight_cover:
        reasons.append('Available stock may need attention soon.')
    if rising_demand:
        reasons.append('Demand has been rising recently.')
    if not reasons and (factors or []):
        reasons.extend(_friendly_factor_phrase(item) for item in factors if str(item).strip())

    forecast_limited = (forecast_status or '').strip().lower() != 'completed'
    limited_probability = probability_value is None

    if forecast_limited and limited_probability and not reorder_pressure and not near_reorder and not tight_cover:
        return {
            'risk_level': 'Unavailable',
            'reason': 'More sales history is needed for this item.',
            'suggested_action': 'Gather more sales history',
            'priority': 'Low',
            'stock_cover_days': stock_cover_days,
            'forecast_pressure': forecast_pressure,
            'reorder_pressure': reorder_pressure,
            'near_reorder': near_reorder,
        }

    if forecast_pressure and (reorder_pressure or low_cover or strong_probability):
        risk_level = 'High'
    elif reorder_pressure and (tight_cover or moderate_probability or rising_demand):
        risk_level = 'High'
    elif forecast_pressure or (near_reorder and (moderate_probability or rising_demand)) or low_cover:
        risk_level = 'Moderate'
    elif approaching_pressure or tight_cover or moderate_probability:
        risk_level = 'Moderate'
    elif low_probability and comfortable_cover and not rising_demand:
        risk_level = 'Low'
    elif near_reorder and rising_demand:
        risk_level = 'Moderate'
    else:
        risk_level = 'Low'

    if risk_level == 'High':
        suggested_action = 'Consider restocking soon'
        priority = 'High'
    elif risk_level == 'Moderate' and (forecast_pressure or near_reorder or rising_demand):
        suggested_action = 'Review before next restocking'
        priority = 'Moderate'
    elif risk_level == 'Moderate':
        suggested_action = 'Monitor this week'
        priority = 'Moderate'
    elif forecast_limited and limited_probability:
        suggested_action = 'Gather more sales history'
        priority = 'Low'
    elif rising_demand:
        suggested_action = 'Monitor this week'
        priority = 'Low'
    else:
        suggested_action = 'Stock appears sufficient'
        priority = 'Low'

    if not reasons:
        if risk_level == 'High':
            reasons = ['This item may need attention soon.']
        elif risk_level == 'Moderate':
            reasons = ['This item should be monitored closely.']
        else:
            reasons = ['Latest recorded stock appears sufficient.']

    return {
        'risk_level': risk_level,
        'reason': ' '.join(dict.fromkeys(reasons[:2])),
        'suggested_action': suggested_action,
        'priority': priority,
        'stock_cover_days': stock_cover_days,
        'forecast_pressure': forecast_pressure,
        'reorder_pressure': reorder_pressure,
        'near_reorder': near_reorder,
    }



























def get_forecast_ranges() -> list[dict[str, Any]]:
    return [
        {"key": "7", "label": "Next 7 Days", "days": 7},
        {"key": "14", "label": "Next 14 Days", "days": 14},
        {"key": "30", "label": "Next 30 Days", "days": 30},
        {"key": "monthly", "label": "Monthly", "days": 30},
    ]


def get_forecast_view_options() -> list[dict[str, Any]]:
    """Display-only forecast grouping options for the Insights forecast graph.

    These options keep the chart balanced by showing a matching amount of
    observed sales and predicted demand without rerunning the models.
    """
    return [
        {"key": "daily", "label": "Daily", "days": 7, "group_by": "daily", "history_days": 7, "history_points": 7, "forecast_points": 7},
        {"key": "weekly", "label": "Weekly", "days": 28, "group_by": "weekly", "history_days": 28, "history_points": 4, "forecast_points": 4},
        {"key": "monthly", "label": "Monthly", "days": 90, "group_by": "monthly", "history_days": 90, "history_points": 3, "forecast_points": 3},
    ]


def _forecast_view_config(view_key: str | None) -> dict[str, Any]:
    options = {option["key"]: option for option in get_forecast_view_options()}
    return options.get(view_key or "daily", options["daily"])


def _format_forecast_axis_label(value: Any, group_by: str = "daily") -> str:
    date_value = pd.to_datetime(value, errors="coerce")
    if pd.isna(date_value):
        return str(value or "")
    if group_by == "weekly":
        week_end = date_value + pd.Timedelta(days=6)
        if date_value.month == week_end.month:
            return f"{date_value.strftime('%b %d')}–{week_end.strftime('%d')}"
        return f"{date_value.strftime('%b %d')}–{week_end.strftime('%b %d')}"
    if group_by == "monthly":
        return date_value.strftime("%b %Y")
    if group_by == "yearly":
        return date_value.strftime("%Y")
    return date_value.strftime("%b %d")


def _aggregate_series_for_forecast_view(series: pd.Series, view_key: str | None = "daily") -> pd.Series:
    if series is None or series.empty:
        return pd.Series(dtype=float)
    clean = series.copy()
    clean.index = pd.to_datetime(clean.index, errors="coerce")
    clean = clean[pd.notna(clean.index)]
    clean = pd.to_numeric(clean, errors="coerce").fillna(0)
    if clean.empty:
        return pd.Series(dtype=float)

    group_by = _forecast_view_config(view_key).get("group_by", "daily")
    if group_by == "weekly":
        grouped_index = clean.index.to_period("W-SUN").start_time
        return clean.groupby(grouped_index).sum().sort_index()
    if group_by == "monthly":
        grouped_index = clean.index.to_period("M").start_time
        return clean.groupby(grouped_index).sum().sort_index()
    if group_by == "yearly":
        grouped_index = clean.index.to_period("Y").start_time
        return clean.groupby(grouped_index).sum().sort_index()
    return clean.groupby(clean.index.normalize()).sum().sort_index()


def _limit_forecast_view_points(series: pd.Series, view_key: str | None, point_type: str = "history") -> pd.Series:
    """Limit grouped chart data to the balanced view required by the selected forecast view."""
    if series is None or series.empty:
        return pd.Series(dtype=float)

    view_config = _forecast_view_config(view_key)
    key = "forecast_points" if point_type == "forecast" else "history_points"
    point_limit = int(view_config.get(key) or 0)
    if point_limit <= 0:
        return series
    if point_type == "forecast":
        return series.head(point_limit)
    return series.tail(point_limit)


def _build_forecast_line_datasets(historical_values: list[float], forecast_values: list[float], future_count: int) -> tuple[list[float | None], list[float | None]]:
    """Return aligned Observed Sales and Predicted Demand datasets for Chart.js.

    The predicted series shares the last observed point before the future labels.
    This keeps the forecast line visually connected while still keeping the
    observed and forecast sections clearly separated in the legend.
    """
    observed_values = list(historical_values or [])
    future_values = list(forecast_values or [])
    observed_data = observed_values + ([None] * future_count)

    if observed_values:
        predicted_data = ([None] * max(len(observed_values) - 1, 0)) + [observed_values[-1]] + future_values
    else:
        predicted_data = future_values

    expected_length = len(observed_values) + future_count
    if len(predicted_data) < expected_length:
        predicted_data.extend([None] * (expected_length - len(predicted_data)))
    elif len(predicted_data) > expected_length:
        predicted_data = predicted_data[:expected_length]

    return observed_data, predicted_data


def _forecast_chart_options(view_key: str | None = "daily") -> dict[str, Any]:
    group_by = _forecast_view_config(view_key).get("group_by", "daily")
    daily_ticks = group_by == "daily"
    weekly_ticks = group_by == "weekly"
    return {
        "layout": {"padding": {"top": 2, "right": 8, "bottom": 2, "left": 6}},
        "plugins": {
            "legend": {"position": "top", "labels": {"usePointStyle": True, "boxWidth": 8, "padding": 10}},
        },
        "scales": {
            "y": {
                "beginAtZero": True,
                "grid": {"color": "#f0f0f0"},
                "title": {"display": True, "text": "Demand"},
            },
            "x": {
                "grid": {"display": False},
                "ticks": {
                    "autoSkip": True,
                    "maxTicksLimit": 9 if daily_ticks else (8 if weekly_ticks else 7),
                    "maxRotation": 0,
                    "minRotation": 0,
                    "padding": 3,
                },
            },
        },
    }


def _forecast_total_from_entry(forecast_entry: dict[str, Any] | None, forecast_days: int) -> float | None:
    if not isinstance(forecast_entry, dict):
        return None

    daily_points = forecast_entry.get("daily") or []
    total = 0.0
    found = False
    for index, item in enumerate(daily_points, start=1):
        if not isinstance(item, dict):
            continue
        horizon = item.get("horizon") or index
        try:
            horizon = int(horizon)
        except Exception:
            horizon = index
        if horizon > forecast_days:
            continue
        quantity = pd.to_numeric(item.get("quantity"), errors="coerce")
        if pd.notna(quantity):
            total += float(quantity)
            found = True

    if found:
        return round(total, 2)

    horizons = forecast_entry.get("horizons") or {}
    direct_value = horizons.get(forecast_days) or horizons.get(str(forecast_days))
    numeric = _extract_number_from_display(direct_value)
    if numeric is not None:
        return round(float(numeric), 2)

    available = []
    for key, value in horizons.items():
        try:
            key_days = int(key)
        except Exception:
            continue
        numeric_value = _extract_number_from_display(value)
        if numeric_value is not None and key_days <= forecast_days:
            available.append((key_days, numeric_value))
    if available:
        available.sort(key=lambda pair: pair[0], reverse=True)
        return round(float(available[0][1]), 2)

    return None


def get_available_categories(df: pd.DataFrame | None) -> list[str]:
    """Return the fixed StockWise category list used by all category filters.

    Uploaded records may contain older or custom category names. Those values are
    standardized when rows are displayed or filtered, so the dropdown remains
    predictable and does not crash when an unknown category appears.
    """
    return get_category_filter_options()




def get_recent_sales_points(df: pd.DataFrame | None, product_name: str, days: int = 7) -> list[dict[str, Any]]:
    if df is None or "product_name" not in df.columns or "date" not in df.columns or "quantity_sold" not in df.columns:
        return []
    working = df.copy()
    working = working[(working["product_name"] == product_name) & working["date"].notna()].copy()
    if working.empty:
        return []
    working["quantity_sold"] = pd.to_numeric(working["quantity_sold"], errors="coerce").fillna(0)
    daily_totals = (
        working.groupby(working["date"].dt.normalize())["quantity_sold"]
        .sum()
        .sort_index()
        .tail(days)
    )
    return [
        {"date": index.strftime("%b %d"), "quantity": round(float(value), 2)}
        for index, value in daily_totals.items()
    ]















# =========================================
# MODEL + DATABASE-BACKED ANALYTICS HELPERS
# =========================================

def _state_get(name: str, default=None):
    return getattr(get_app_state(), name, default)


def _state_set(name: str, value):
    setattr(get_app_state(), name, value)


def clear_model_cache() -> None:
    for key in [
        "latest_upload_id",
        "latest_model_run_id",
        "latest_model_artifacts",
        "model_ui_summary",
    ]:
        _state_set(key, None)
    try:
        clear_insights_cache()
    except NameError:
        pass


def store_processed_dataset(df: pd.DataFrame, filename: str, upload_id: int | None = None) -> None:
    state = get_app_state()
    state.processed_data = df.copy()
    state.processed_filename = filename
    state.processed_at = datetime.now()
    if upload_id is not None:
        state.latest_upload_id = upload_id
    clear_insights_cache()


def clear_processed_dataset() -> None:
    state = get_app_state()
    state.processed_data = None
    state.processed_filename = None
    state.processed_at = None
    clear_model_cache()
    clear_insights_cache()


def _get_current_database(cursor) -> str:
    cursor.execute("SELECT DATABASE() AS db_name")
    row = cursor.fetchone()
    if isinstance(row, dict):
        return row.get('db_name')
    return row[0] if row else ''


def _fetch_table_columns(cursor, table_name: str) -> dict[str, dict[str, Any]]:
    db_name = _get_current_database(cursor)
    cursor.execute(
        """
        SELECT COLUMN_NAME, COLUMN_TYPE, IS_NULLABLE, COLUMN_DEFAULT, EXTRA
        FROM information_schema.COLUMNS
        WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s
        """,
        (db_name, table_name),
    )
    rows = cursor.fetchall() or []
    columns: dict[str, dict[str, Any]] = {}
    for row in rows:
        if isinstance(row, dict):
            columns[row['COLUMN_NAME']] = row
        else:
            columns[row[0]] = {
                'COLUMN_NAME': row[0],
                'COLUMN_TYPE': row[1],
                'IS_NULLABLE': row[2],
                'COLUMN_DEFAULT': row[3],
                'EXTRA': row[4],
            }
    return columns


def _fetch_table_indexes(cursor, table_name: str) -> set[str]:
    db_name = _get_current_database(cursor)
    cursor.execute(
        """
        SELECT DISTINCT INDEX_NAME
        FROM information_schema.STATISTICS
        WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s
        """,
        (db_name, table_name),
    )
    rows = cursor.fetchall() or []
    names = set()
    for row in rows:
        if isinstance(row, dict):
            names.add(row['INDEX_NAME'])
        else:
            names.add(row[0])
    return names


def _ensure_index(cursor, table_name: str, index_name: str, definition_sql: str) -> None:
    existing_indexes = _fetch_table_indexes(cursor, table_name)
    if index_name not in existing_indexes:
        cursor.execute(f"ALTER TABLE {table_name} ADD INDEX {index_name} {definition_sql}")


def _drop_single_column_unique_index(cursor, table_name: str, column_name: str) -> None:
    """Remove an old single-column unique index when the system now scopes records per user."""
    db_name = _get_current_database(cursor)
    cursor.execute(
        """
        SELECT s.INDEX_NAME, COUNT(*) AS column_count
        FROM information_schema.STATISTICS s
        WHERE s.TABLE_SCHEMA = %s
          AND s.TABLE_NAME = %s
          AND s.NON_UNIQUE = 0
          AND s.INDEX_NAME <> 'PRIMARY'
        GROUP BY s.INDEX_NAME
        HAVING column_count = 1
           AND SUM(CASE WHEN s.COLUMN_NAME = %s THEN 1 ELSE 0 END) = 1
        """,
        (db_name, table_name, column_name),
    )
    rows = cursor.fetchall() or []
    for row in rows:
        index_name = row['INDEX_NAME'] if isinstance(row, dict) else row[0]
        safe_index = str(index_name).replace('`', '``')
        cursor.execute(f"ALTER TABLE {table_name} DROP INDEX `{safe_index}`")


def ensure_model_tables() -> None:
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        add_column_if_missing(cursor, "products", "unit_type", "VARCHAR(100) NULL")
        add_column_if_missing(cursor, "products", "user_id", "INT(11) NULL", after_column="product_id")
        _drop_single_column_unique_index(cursor, 'products', 'product_name')
        _ensure_index(cursor, 'products', 'idx_products_user_name', '(user_id, product_name)')

        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS model_runs (
                model_run_id INT AUTO_INCREMENT PRIMARY KEY,
                upload_id INT NOT NULL,
                user_id INT NOT NULL,
                run_status VARCHAR(32) DEFAULT 'started',
                sarima_status VARCHAR(32) DEFAULT 'pending',
                xgboost_status VARCHAR(32) DEFAULT 'pending',
                sarima_version VARCHAR(100) NULL,
                xgboost_version VARCHAR(100) NULL,
                notes TEXT NULL,
                started_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                completed_at DATETIME NULL
            )
            """
        )
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS forecasts (
                forecast_id BIGINT AUTO_INCREMENT PRIMARY KEY,
                model_run_id INT NOT NULL,
                upload_id INT NULL,
                user_id INT NULL,
                product_id INT NULL,
                product_name VARCHAR(255) NOT NULL,
                forecast_date DATE NULL,
                horizon_days INT NULL,
                forecast_quantity DECIMAL(12,4) NULL,
                model_source VARCHAR(32) DEFAULT 'SARIMA',
                status_label VARCHAR(32) DEFAULT 'completed',
                note TEXT NULL,
                generated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS stockout_predictions (
                prediction_id BIGINT AUTO_INCREMENT PRIMARY KEY,
                model_run_id INT NOT NULL,
                upload_id INT NULL,
                user_id INT NULL,
                product_id INT NULL,
                product_name VARCHAR(255) NOT NULL,
                prediction_date DATE NULL,
                forecast_horizon_days INT NULL,
                stockout_probability DECIMAL(10,6) NULL,
                risk_level VARCHAR(32) DEFAULT 'Unavailable',
                model_source VARCHAR(32) DEFAULT 'XGBoost',
                top_factors TEXT NULL,
                note TEXT NULL,
                recommendation VARCHAR(255) NULL,
                generated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
            """
        )

        # model_runs schema harmonization
        model_run_cols = _fetch_table_columns(cursor, 'model_runs')
        if 'run_id' in model_run_cols and 'model_run_id' not in model_run_cols:
            cursor.execute("ALTER TABLE model_runs CHANGE COLUMN run_id model_run_id INT(11) NOT NULL AUTO_INCREMENT")
            model_run_cols = _fetch_table_columns(cursor, 'model_runs')
        if 'initiated_by_user_id' in model_run_cols and 'user_id' not in model_run_cols:
            cursor.execute("ALTER TABLE model_runs CHANGE COLUMN initiated_by_user_id user_id INT(11) NOT NULL")
            model_run_cols = _fetch_table_columns(cursor, 'model_runs')
        if 'sarima_model_version' in model_run_cols and 'sarima_version' not in model_run_cols:
            cursor.execute("ALTER TABLE model_runs CHANGE COLUMN sarima_model_version sarima_version VARCHAR(100) NULL")
            model_run_cols = _fetch_table_columns(cursor, 'model_runs')
        if 'xgboost_model_version' in model_run_cols and 'xgboost_version' not in model_run_cols:
            cursor.execute("ALTER TABLE model_runs CHANGE COLUMN xgboost_model_version xgboost_version VARCHAR(100) NULL")
            model_run_cols = _fetch_table_columns(cursor, 'model_runs')
        if 'created_at' in model_run_cols and 'started_at' not in model_run_cols:
            cursor.execute("ALTER TABLE model_runs CHANGE COLUMN created_at started_at DATETIME DEFAULT CURRENT_TIMESTAMP")
            model_run_cols = _fetch_table_columns(cursor, 'model_runs')
        for column_name, column_definition, after_column in [
            ("user_id", "INT(11) NOT NULL", "upload_id"),
            ("run_status", "VARCHAR(32) DEFAULT 'started'", "user_id"),
            ("sarima_status", "VARCHAR(32) DEFAULT 'pending'", "run_status"),
            ("xgboost_status", "VARCHAR(32) DEFAULT 'pending'", "sarima_status"),
            ("sarima_version", "VARCHAR(100) NULL", "xgboost_status"),
            ("xgboost_version", "VARCHAR(100) NULL", "sarima_version"),
            ("notes", "TEXT NULL", "xgboost_version"),
            ("started_at", "DATETIME DEFAULT CURRENT_TIMESTAMP", "notes"),
            ("completed_at", "DATETIME NULL", "started_at"),
        ]:
            add_column_if_missing(cursor, "model_runs", column_name, column_definition, after_column=after_column)
        _ensure_index(cursor, 'model_runs', 'idx_model_runs_user_upload', '(user_id, upload_id)')

        # forecasts schema harmonization
        forecast_cols = _fetch_table_columns(cursor, 'forecasts')
        if 'run_id' in forecast_cols and 'model_run_id' not in forecast_cols:
            cursor.execute("ALTER TABLE forecasts CHANGE COLUMN run_id model_run_id INT(11) NOT NULL")
            forecast_cols = _fetch_table_columns(cursor, 'forecasts')
        if 'forecast_for_date' in forecast_cols and 'forecast_date' not in forecast_cols:
            cursor.execute("ALTER TABLE forecasts CHANGE COLUMN forecast_for_date forecast_date DATE NULL")
            forecast_cols = _fetch_table_columns(cursor, 'forecasts')
        if 'predicted_demand' in forecast_cols and 'forecast_quantity' not in forecast_cols:
            cursor.execute("ALTER TABLE forecasts CHANGE COLUMN predicted_demand forecast_quantity DECIMAL(12,4) NULL")
            forecast_cols = _fetch_table_columns(cursor, 'forecasts')
        if 'created_at' in forecast_cols and 'generated_at' not in forecast_cols:
            cursor.execute("ALTER TABLE forecasts CHANGE COLUMN created_at generated_at DATETIME DEFAULT CURRENT_TIMESTAMP")
            forecast_cols = _fetch_table_columns(cursor, 'forecasts')
        for column_name, column_definition, after_column in [
            ("upload_id", "INT(11) NULL", "model_run_id"),
            ("user_id", "INT(11) NULL", "upload_id"),
            ("product_name", "VARCHAR(255) NULL", "product_id"),
            ("horizon_days", "INT(11) NULL", "forecast_date"),
            ("model_source", "VARCHAR(32) DEFAULT 'SARIMA'", "forecast_quantity"),
            ("status_label", "VARCHAR(32) DEFAULT 'completed'", "model_source"),
            ("note", "TEXT NULL", "status_label"),
            ("generated_at", "DATETIME DEFAULT CURRENT_TIMESTAMP", "note"),
        ]:
            add_column_if_missing(cursor, "forecasts", column_name, column_definition, after_column=after_column)
        cursor.execute(
            """
            ALTER TABLE forecasts
            MODIFY COLUMN product_id INT(11) NULL,
            MODIFY COLUMN forecast_date DATE NULL,
            MODIFY COLUMN forecast_quantity DECIMAL(12,4) NULL,
            MODIFY COLUMN horizon_days INT(11) NULL,
            MODIFY COLUMN model_source VARCHAR(32) NULL DEFAULT 'SARIMA',
            MODIFY COLUMN status_label VARCHAR(32) NULL DEFAULT 'completed'
            """
        )
        _ensure_index(cursor, 'forecasts', 'idx_forecasts_lookup', '(user_id, upload_id, product_name, horizon_days)')

        # stockout_predictions schema harmonization
        prediction_cols = _fetch_table_columns(cursor, 'stockout_predictions')
        if 'run_id' in prediction_cols and 'model_run_id' not in prediction_cols:
            cursor.execute("ALTER TABLE stockout_predictions CHANGE COLUMN run_id model_run_id INT(11) NOT NULL")
            prediction_cols = _fetch_table_columns(cursor, 'stockout_predictions')
        if 'predicted_for_date' in prediction_cols and 'prediction_date' not in prediction_cols:
            cursor.execute("ALTER TABLE stockout_predictions CHANGE COLUMN predicted_for_date prediction_date DATE NULL")
            prediction_cols = _fetch_table_columns(cursor, 'stockout_predictions')
        if 'predicted_stockout_probability' in prediction_cols and 'stockout_probability' not in prediction_cols:
            cursor.execute("ALTER TABLE stockout_predictions CHANGE COLUMN predicted_stockout_probability stockout_probability DECIMAL(10,6) NULL")
            prediction_cols = _fetch_table_columns(cursor, 'stockout_predictions')
        if 'created_at' in prediction_cols and 'generated_at' not in prediction_cols:
            cursor.execute("ALTER TABLE stockout_predictions CHANGE COLUMN created_at generated_at DATETIME DEFAULT CURRENT_TIMESTAMP")
            prediction_cols = _fetch_table_columns(cursor, 'stockout_predictions')
        for column_name, column_definition, after_column in [
            ("upload_id", "INT(11) NULL", "model_run_id"),
            ("user_id", "INT(11) NULL", "upload_id"),
            ("product_name", "VARCHAR(255) NULL", "product_id"),
            ("forecast_horizon_days", "INT(11) NULL", "prediction_date"),
            ("model_source", "VARCHAR(32) DEFAULT 'XGBoost'", "risk_level"),
            ("top_factors", "TEXT NULL", "model_source"),
            ("note", "TEXT NULL", "top_factors"),
            ("recommendation", "VARCHAR(255) NULL", "note"),
            ("generated_at", "DATETIME DEFAULT CURRENT_TIMESTAMP", "recommendation"),
        ]:
            add_column_if_missing(cursor, "stockout_predictions", column_name, column_definition, after_column=after_column)
        cursor.execute(
            """
            ALTER TABLE stockout_predictions
            MODIFY COLUMN product_id INT(11) NULL,
            MODIFY COLUMN prediction_date DATE NULL,
            MODIFY COLUMN forecast_horizon_days INT(11) NULL,
            MODIFY COLUMN stockout_probability DECIMAL(10,6) NULL,
            MODIFY COLUMN risk_level VARCHAR(32) NULL DEFAULT 'Unavailable',
            MODIFY COLUMN model_source VARCHAR(32) NULL DEFAULT 'XGBoost'
            """
        )
        _ensure_index(cursor, 'stockout_predictions', 'idx_predictions_lookup', '(user_id, upload_id, product_name)')

        conn.commit()
    finally:
        cursor.close()
        conn.close()


def get_or_create_product_id(cursor, user_id: int, product_name: str, category_id: int, unit_price: float, reorder_point: int, unit_type: str = "Unit") -> int:
    safe_name = (product_name or "").strip()
    if not safe_name:
        raise ValueError("Product name is required.")

    cursor.execute(
        "SELECT product_id FROM products WHERE user_id = %s AND product_name = %s LIMIT 1",
        (user_id, safe_name)
    )
    row = cursor.fetchone()

    if row:
        product_id = row[0]
        cursor.execute(
            """
            UPDATE products
            SET category_id = %s,
                standard_price = %s,
                reorder_point = %s,
                unit_type = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE product_id = %s
            """,
            (category_id, unit_price, reorder_point, unit_type, product_id)
        )
        return product_id

    cursor.execute(
        """
        INSERT INTO products (user_id, category_id, product_name, standard_price, reorder_point, unit_type, is_active)
        VALUES (%s, %s, %s, %s, %s, %s, 1)
        """,
        (user_id, category_id, safe_name, unit_price, reorder_point, unit_type)
    )
    return cursor.lastrowid


def _chunked(iterable: list[Any], chunk_size: int = 1000) -> list[list[Any]]:
    return [iterable[index:index + chunk_size] for index in range(0, len(iterable), chunk_size)]


def save_processed_dataset_to_database(user_id: int, filename: str, processed_df: pd.DataFrame, upload_mode: str = "new") -> int:
    if user_id is None:
        raise ValueError("A logged-in user is required before saving uploads.")
    if processed_df is None or processed_df.empty:
        raise ValueError("No processed sales data is available to save.")

    ensure_model_tables()

    file_type = filename.rsplit('.', 1)[1].lower() if '.' in filename else 'csv'
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            """
            INSERT INTO uploads (user_id, file_name, file_type, upload_status, row_count, remarks)
            VALUES (%s, %s, %s, %s, %s, %s)
            """,
            (user_id, filename, file_type, 'validated', len(processed_df), f'Processed and saved from Flask upload workflow ({set_selected_upload_mode(upload_mode)})')
        )
        upload_id = cursor.lastrowid

        category_cache: dict[str, int] = {}
        product_map: dict[str, int] = {}
        latest_inventory_snapshots: dict[int, dict[str, Any]] = {}
        sales_rows: list[tuple[Any, ...]] = []
        inventory_history_rows: list[tuple[Any, ...]] = []
        data_format_preferences = get_data_format_preferences()
        saved_time_format = data_format_preferences.get("data_time_format", "auto")

        for _, row in processed_df.iterrows():
            product_name = str(row.get('product_name', '')).strip()
            if not product_name:
                continue

            parsed_date = pd.to_datetime(row.get('date'), errors='coerce')
            if pd.isna(parsed_date):
                continue

            quantity_sold = to_int(row.get('quantity_sold'), 0)
            if quantity_sold <= 0:
                continue

            category_name = str(row.get('category', 'Uncategorized')).strip() or 'Uncategorized'
            unit_price = to_float(row.get('unit_price'), 0.0)
            reorder_point = max(to_int(row.get('reorder_point'), 0), 0)
            unit_type = str(row.get('unit_type', 'Unit')).strip() or 'Unit'

            category_id = category_cache.get(category_name)
            if category_id is None:
                category_id = get_or_create_category_id(cursor, category_name)
                category_cache[category_name] = category_id

            product_id = product_map.get(product_name)
            if product_id is None:
                product_id = get_or_create_product_id(cursor, user_id, product_name, category_id, unit_price, reorder_point, unit_type)
                product_map[product_name] = product_id

            transaction_time = parse_time_for_sql(row.get('time'), saved_time_format)
            time_of_day = infer_time_of_day_label(transaction_time)
            day_of_week = parsed_date.day_name()
            is_payday = to_int(row.get('is_payday'), 1 if parsed_date.day in {15, 30} else 0)

            sales_rows.append((
                upload_id,
                product_id,
                quantity_sold,
                unit_price,
                parsed_date.date(),
                transaction_time,
                time_of_day,
                day_of_week,
                is_payday,
            ))

            current_stock_raw = row.get('current_stock')
            if current_stock_raw is not None and not pd.isna(current_stock_raw):
                current_stock = max(to_int(current_stock_raw, 0), 0)
                existing_snapshot = latest_inventory_snapshots.get(product_id)
                if existing_snapshot is None or parsed_date.to_pydatetime() >= existing_snapshot['recorded_at']:
                    latest_inventory_snapshots[product_id] = {
                        'recorded_at': parsed_date.to_pydatetime(),
                        'stock_on_hand': current_stock,
                        'is_stockout': 1 if current_stock <= 0 else 0,
                    }

        if not sales_rows:
            raise ValueError("No valid sales rows were available to save after processing.")

        insert_sales_sql = """
            INSERT INTO sales_transactions (
                upload_id, product_id, quantity_sold, unit_price, transaction_date, transaction_time,
                time_of_day, day_of_week, is_payday
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        for chunk in _chunked(sales_rows, 1000):
            cursor.executemany(insert_sales_sql, chunk)

        for product_id, snapshot in latest_inventory_snapshots.items():
            upsert_inventory(cursor, product_id, snapshot['stock_on_hand'])
            inventory_history_rows.append((
                product_id,
                upload_id,
                snapshot['recorded_at'],
                snapshot['stock_on_hand'],
                snapshot['is_stockout'],
                'Captured from uploaded sales data',
            ))

        if inventory_history_rows:
            insert_history_sql = """
                INSERT INTO inventory_history (product_id, upload_id, recorded_at, stock_on_hand, is_stockout, notes)
                VALUES (%s, %s, %s, %s, %s, %s)
            """
            for chunk in _chunked(inventory_history_rows, 1000):
                cursor.executemany(insert_history_sql, chunk)

        cursor.execute(
            """
            UPDATE uploads
            SET upload_status = %s, row_count = %s, processed_at = CURRENT_TIMESTAMP
            WHERE upload_id = %s
            """,
            ('processed', len(sales_rows), upload_id)
        )
        _apply_processed_upload_mode_effects(cursor, user_id, upload_id, upload_mode)
        conn.commit()
        _state_set('latest_upload_id', upload_id)
        return upload_id
    except Exception:
        conn.rollback()
        raise
    finally:
        cursor.close()
        conn.close()

def _apply_processed_upload_mode_effects(cursor, user_id: int, new_upload_id: int, upload_mode: str) -> None:
    normalized_mode = set_selected_upload_mode(upload_mode)
    if normalized_mode in {'append', 'replace'}:
        action_text = 'Superseded by appended upload' if normalized_mode == 'append' else 'Replaced by upload'
        store_user_ids = get_store_user_ids(user_id)
        placeholders, params = make_in_clause(store_user_ids)
        cursor.execute(
            f"""
            UPDATE uploads
            SET remarks = CONCAT(COALESCE(remarks, ''), CASE WHEN COALESCE(remarks, '') = '' THEN '' ELSE ' | ' END, %s, ' ', %s)
            WHERE user_id IN ({placeholders}) AND upload_id <> %s AND upload_status = 'processed'
            """,
            (action_text, str(new_upload_id), *params, new_upload_id),
        )




def _load_latest_upload_metadata(user_id: int) -> dict[str, Any] | None:
    if user_id is None:
        return None
    try:
        store_user_ids = get_store_user_ids(user_id)
        placeholders, params = make_in_clause(store_user_ids)
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        try:
            cursor.execute(
                f"""
                SELECT upload_id, user_id, file_name, processed_at
                FROM uploads
                WHERE user_id IN ({placeholders}) AND upload_status = 'processed'
                ORDER BY COALESCE(processed_at, uploaded_at) DESC, upload_id DESC
                LIMIT 1
                """,
                params,
            )
            return cursor.fetchone()
        finally:
            cursor.close()
            conn.close()
    except Exception:
        return None




def _load_processed_dataset_from_db(user_id: int) -> tuple[pd.DataFrame | None, dict[str, Any] | None]:
    meta = _load_latest_upload_metadata(user_id)
    if not meta:
        return None, None
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        try:
            cursor.execute(
                """
                SELECT st.transaction_date AS date,
                       p.product_name,
                       st.quantity_sold,
                       st.transaction_time AS time,
                       c.category_name AS category,
                       COALESCE(ih.stock_on_hand, i.current_stock) AS current_stock,
                       p.reorder_point,
                       st.unit_price,
                       COALESCE(p.unit_type, 'Unit') AS unit_type
                FROM sales_transactions st
                JOIN products p ON p.product_id = st.product_id
                LEFT JOIN categories c ON c.category_id = p.category_id
                LEFT JOIN inventory i ON i.product_id = p.product_id
                LEFT JOIN (
                    SELECT h.product_id, h.upload_id, h.stock_on_hand
                    FROM inventory_history h
                    JOIN (
                        SELECT product_id, upload_id, MAX(recorded_at) AS latest_recorded_at
                        FROM inventory_history
                        WHERE upload_id = %s
                        GROUP BY product_id, upload_id
                    ) latest_h
                    ON latest_h.product_id = h.product_id
                    AND latest_h.upload_id = h.upload_id
                    AND latest_h.latest_recorded_at = h.recorded_at
                ) ih ON ih.product_id = st.product_id AND ih.upload_id = st.upload_id
                WHERE st.upload_id = %s
                ORDER BY st.transaction_date, p.product_name
                """,
                (meta['upload_id'], meta['upload_id'])
            )
        except Exception:
            cursor.execute(
                """
                SELECT st.transaction_date AS date,
                       p.product_name,
                       st.quantity_sold,
                       st.transaction_time AS time,
                       c.category_name AS category,
                       COALESCE(ih.stock_on_hand, i.current_stock) AS current_stock,
                       p.reorder_point,
                       st.unit_price,
                       'Unit' AS unit_type
                FROM sales_transactions st
                JOIN products p ON p.product_id = st.product_id
                LEFT JOIN categories c ON c.category_id = p.category_id
                LEFT JOIN inventory i ON i.product_id = p.product_id
                LEFT JOIN (
                    SELECT h.product_id, h.upload_id, h.stock_on_hand
                    FROM inventory_history h
                    JOIN (
                        SELECT product_id, upload_id, MAX(recorded_at) AS latest_recorded_at
                        FROM inventory_history
                        WHERE upload_id = %s
                        GROUP BY product_id, upload_id
                    ) latest_h
                    ON latest_h.product_id = h.product_id
                    AND latest_h.upload_id = h.upload_id
                    AND latest_h.latest_recorded_at = h.recorded_at
                ) ih ON ih.product_id = st.product_id AND ih.upload_id = st.upload_id
                WHERE st.upload_id = %s
                ORDER BY st.transaction_date, p.product_name
                """,
                (meta['upload_id'], meta['upload_id'])
            )
        rows = cursor.fetchall()
    finally:
        cursor.close()
        conn.close()
    if not rows:
        return None, meta
    df = normalize_dataframe(pd.DataFrame(rows))
    if 'date' in df.columns:
        df['date'] = pd.to_datetime(df['date'], errors='coerce')
        if 'time' in df.columns:
            parsed_time = pd.to_datetime(df['time'].astype(str), errors='coerce')
            df['hour_of_day'] = parsed_time.dt.hour
        df['day_of_week'] = df['date'].dt.dayofweek
        df['day_name'] = df['date'].dt.day_name()
        df['is_weekend'] = df['day_of_week'].isin([5,6]).astype(int)
        df['is_payday'] = df['date'].dt.day.isin([15,30]).astype(int)
        df['month_period'] = df['date'].dt.to_period('M').astype(str)
    return df, meta


def _ensure_state_loaded_from_db() -> None:
    user_id = get_current_user_id()
    if not user_id:
        return
    state = get_app_state()
    try:
        latest_meta = _load_latest_upload_metadata(user_id)
    except Exception:
        latest_meta = None

    latest_upload_id = latest_meta.get('upload_id') if latest_meta else None
    if state.processed_data is not None and getattr(state, 'latest_upload_id', None) and latest_upload_id == getattr(state, 'latest_upload_id', None):
        return

    try:
        df, meta = _load_processed_dataset_from_db(user_id)
    except Exception:
        state.processed_data = None
        state.processed_filename = None
        state.processed_at = None
        state.latest_upload_id = None
        return
    if df is not None and meta is not None:
        state.processed_data = df
        state.processed_filename = meta.get('file_name')
        state.processed_at = meta.get('processed_at')
        state.latest_upload_id = meta.get('upload_id')

        warmup_upload_id = meta.get('upload_id')
        warmup_key = f"db:{user_id}:{warmup_upload_id}" if warmup_upload_id else None
        if warmup_key and _state_get("db_loaded_context_warmup_key") != warmup_key:
            try:
                warm_up_generated_page_contexts(df)
                _state_set("db_loaded_context_warmup_key", warmup_key)
            except Exception:
                # Display cache warm-up should never block page loading.
                pass
    elif latest_meta is None:
        state.processed_data = None
        state.processed_filename = None
        state.processed_at = None
        state.latest_upload_id = None


def get_processed_dataset() -> pd.DataFrame | None:
    _ensure_state_loaded_from_db()
    return get_app_state().processed_data


def get_processed_filename() -> str | None:
    _ensure_state_loaded_from_db()
    return get_app_state().processed_filename


def get_last_processed_label() -> str:
    _ensure_state_loaded_from_db()
    return format_datetime(get_app_state().processed_at)


def get_upload_status() -> dict[str, Any]:
    _ensure_state_loaded_from_db()
    state = get_app_state()
    selected_df = state.selected_data
    processed_df = state.processed_data
    active_df = processed_df if processed_df is not None else selected_df
    active_name = state.processed_filename if processed_df is not None else state.selected_filename
    return {
        'selected_filename': state.selected_filename,
        'processed_filename': state.processed_filename,
        'selected_at': state.selected_at,
        'processed_at': state.processed_at,
        'selected_coverage_period': infer_coverage_period(selected_df),
        'processed_coverage_period': infer_coverage_period(processed_df),
        'coverage_period': infer_coverage_period(active_df),
        'freshness_label': infer_upload_freshness(state.processed_at),
        'last_upload_mode': state.last_upload_mode,
        'active_filename': active_name,
        'missing_date_gaps': find_missing_date_gaps(processed_df if processed_df is not None else selected_df),
        'coverage_overlap': coverage_overlap(selected_df, processed_df),
        'recommended_upload_frequency': 'Daily uploads are recommended for more reliable forecast and stockout insights.',
    }


def _prepare_daily_model_frame(df: pd.DataFrame | None) -> pd.DataFrame:
    if df is None or df.empty or 'product_name' not in df.columns or 'date' not in df.columns:
        return pd.DataFrame()
    working = df.copy()
    working['date'] = pd.to_datetime(working['date'], errors='coerce').dt.normalize()
    working = working.dropna(subset=['product_name', 'date'])
    if working.empty:
        return pd.DataFrame()
    for col in ['quantity_sold','current_stock','reorder_point','unit_price','hour_of_day']:
        if col not in working.columns:
            working[col] = np.nan
        working[col] = pd.to_numeric(working[col], errors='coerce')
    if 'category' not in working.columns:
        working['category'] = 'Seasonal / Miscellaneous Items'
    if 'unit_type' not in working.columns:
        working['unit_type'] = 'Unit'
    grouped = (
        working.groupby(['product_name','date'], as_index=False)
        .agg({
            'quantity_sold': 'sum',
            'category': 'last',
            'current_stock': 'last',
            'reorder_point': 'last',
            'unit_price': 'last',
            'unit_type': 'last',
            'hour_of_day': 'mean',
        })
    )
    daily_frames=[]
    for product_name, grp in grouped.groupby('product_name'):
        grp = grp.sort_values('date').copy()
        full_range = pd.date_range(grp['date'].min(), grp['date'].max(), freq='D')
        grp = grp.set_index('date').reindex(full_range)
        grp.index.name='date'
        grp['product_name']=product_name
        grp['quantity_sold']=pd.to_numeric(grp['quantity_sold'], errors='coerce').fillna(0)
        for col, default in [('category','Uncategorized'),('unit_type','Unit')]:
            grp[col]=grp[col].ffill().bfill().fillna(default)
        for col in ['current_stock','reorder_point','unit_price','hour_of_day']:
            grp[col]=pd.to_numeric(grp[col], errors='coerce')
            grp[col]=grp[col].ffill().bfill()
        grp = grp.reset_index().rename(columns={'index':'date'})
        grp['day_of_week'] = grp['date'].dt.dayofweek
        grp['day_name'] = grp['date'].dt.day_name()
        grp['is_weekend'] = grp['day_of_week'].isin([5,6]).astype(int)
        grp['is_payday'] = grp['date'].dt.day.isin([15,30]).astype(int)
        grp['month'] = grp['date'].dt.month
        daily_frames.append(grp)
    if not daily_frames:
        return pd.DataFrame()
    return pd.concat(daily_frames, ignore_index=True)


def _fit_sarima_forecast(series: pd.Series, horizon: int = 30) -> dict[str, Any]:
    y = pd.Series(series).astype(float).fillna(0.0)
    if len(y) < 14 or y.nunique() < 2:
        return {'status':'limited','note':'Insufficient history for SARIMA forecasting.','daily_forecast':[],'model_source':'SARIMA'}
    seasonal = len(y) >= 21
    candidate_specs = [
        ((1,1,1),(1,1,0,7) if seasonal else (0,0,0,0)),
        ((0,1,1),(0,0,0,0)),
        ((1,0,1),(0,0,0,0)),
    ]
    last_error = None
    for order, seasonal_order in candidate_specs:
        try:
            with warnings.catch_warnings():
                warnings.simplefilter('ignore')
                model = SARIMAX(y, order=order, seasonal_order=seasonal_order, enforce_stationarity=False, enforce_invertibility=False)
                fitted = model.fit(disp=False)
                forecast_res = fitted.get_forecast(steps=horizon)
                predicted = np.maximum(np.asarray(forecast_res.predicted_mean, dtype=float), 0)
                return {
                    'status':'completed',
                    'note':f'SARIMA forecast generated using order={order} and seasonal_order={seasonal_order}.',
                    'daily_forecast':[round(float(x),4) for x in predicted.tolist()],
                    'model_source':'SARIMA',
                }
        except Exception as exc:
            last_error = str(exc)
            continue
    return {'status':'limited','note':f'SARIMA could not be fitted reliably. {last_error or "Try more history."}','daily_forecast':[],'model_source':'SARIMA'}


def _build_xgb_training_frame(daily_df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    if daily_df.empty:
        return pd.DataFrame(), []
    frames = []
    feature_cols = [
        'rolling_3','rolling_7','rolling_14','recent_growth','is_payday','is_weekend','day_of_week',
        'month','current_stock','reorder_point','stock_to_reorder_ratio','demand_stock_ratio','expected_next_7d','category_code'
    ]
    for product_name, grp in daily_df.groupby('product_name'):
        grp = grp.sort_values('date').copy()
        grp['rolling_3'] = grp['quantity_sold'].rolling(3, min_periods=1).mean()
        grp['rolling_7'] = grp['quantity_sold'].rolling(7, min_periods=1).mean()
        grp['rolling_14'] = grp['quantity_sold'].rolling(14, min_periods=1).mean()
        grp['recent_growth'] = ((grp['rolling_3'] + 1) / (grp['rolling_7'] + 1)) - 1
        grp['future_1'] = grp['quantity_sold'].shift(-1).fillna(0)
        grp['future_2'] = grp['quantity_sold'].shift(-2).fillna(0)
        grp['future_3'] = grp['quantity_sold'].shift(-3).fillna(0)
        grp['future_3d'] = grp['future_1'] + grp['future_2'] + grp['future_3']
        grp['expected_next_7d'] = grp['rolling_7'] * 7
        grp['current_stock'] = pd.to_numeric(grp['current_stock'], errors='coerce')
        grp['reorder_point'] = pd.to_numeric(grp['reorder_point'], errors='coerce')
        grp['stock_to_reorder_ratio'] = grp['current_stock'] / (grp['reorder_point'].replace(0, np.nan))
        grp['stock_to_reorder_ratio'] = grp['stock_to_reorder_ratio'].replace([np.inf, -np.inf], np.nan).fillna(0)
        grp['demand_stock_ratio'] = grp['future_3d'] / (grp['current_stock'].replace(0, np.nan))
        grp['demand_stock_ratio'] = grp['demand_stock_ratio'].replace([np.inf, -np.inf], np.nan).fillna(0)
        grp['category_code'] = grp['category'].astype('category').cat.codes
        grp['target_stockout'] = ((grp['current_stock'] <= grp['reorder_point']) | (grp['future_3d'] > grp['current_stock'])).astype(int)
        grp = grp.dropna(subset=['current_stock','reorder_point'])
        if len(grp) > 3:
            grp = grp.iloc[:-3].copy()
        frames.append(grp)
    if not frames:
        return pd.DataFrame(), feature_cols
    training_df = pd.concat(frames, ignore_index=True)
    training_df = training_df.dropna(subset=feature_cols + ['target_stockout'])
    return training_df, feature_cols


def _derive_top_factor_notes(feature_importances: dict[str, float], row: dict[str, Any]) -> list[str]:
    notes = []
    expected_next_7d = to_float(row.get('expected_next_7d'))
    current_stock = to_float(row.get('current_stock'))
    reorder_point = to_float(row.get('reorder_point'))
    recent_growth = to_float(row.get('recent_growth'), 0.0)
    stock_cover_days = _calculate_stock_cover_days(current_stock, to_float(row.get('rolling_7'), 0.0))

    for feature, _ in sorted(feature_importances.items(), key=lambda item: item[1], reverse=True)[:5]:
        if feature == 'expected_next_7d' and expected_next_7d is not None and current_stock is not None and expected_next_7d > current_stock:
            notes.append('predicted demand may exceed available stock')
        elif feature == 'stock_to_reorder_ratio' and current_stock is not None and reorder_point is not None and current_stock <= reorder_point * 1.1:
            notes.append('stock is at or below the reorder point')
        elif feature == 'recent_growth' and recent_growth > 0.08:
            notes.append('recent demand is rising')
        elif feature == 'is_payday' and row.get('is_payday', 0) == 1:
            notes.append('payday-related demand pattern detected')
        elif feature == 'is_weekend' and row.get('is_weekend', 0) == 1:
            notes.append('weekend demand pattern detected')
        elif feature == 'rolling_7' and stock_cover_days is not None and stock_cover_days < 5:
            notes.append('available stock may only cover a few days of demand')
        elif feature == 'rolling_7' and to_float(row.get('rolling_7'), 0.0) > 0:
            notes.append('recent average demand remains active')
    if not notes:
        notes.append('risk is based on latest recorded stock, reorder point, and recent demand behavior')
    return list(dict.fromkeys(notes))[:3]


def _train_xgboost_and_predict(daily_df: pd.DataFrame, sarima_summary: dict[str, Any]) -> dict[str, Any]:
    training_df, feature_cols = _build_xgb_training_frame(daily_df)
    if training_df.empty or len(training_df) < 20 or training_df['target_stockout'].nunique() < 2:
        return {
            'status': 'limited',
            'note': 'XGBoost could not be trained reliably because the uploaded data has limited stockout class variety or too few usable rows.',
            'predictions': {},
            'feature_importances': {},
            'model_source': 'XGBoost',
        }

    X = training_df[feature_cols].astype(float)
    y = training_df['target_stockout'].astype(int)
    model = XGBClassifier(
        n_estimators=120,
        max_depth=3,
        learning_rate=0.08,
        subsample=0.9,
        colsample_bytree=0.9,
        objective='binary:logistic',
        eval_metric='logloss',
        random_state=42,
        n_jobs=1,
        verbosity=0,
    )
    model.fit(X, y)
    feature_importances = {feature: float(importance) for feature, importance in zip(feature_cols, model.feature_importances_)}

    predictions = {}
    for product_name, grp in daily_df.groupby('product_name'):
        grp = grp.sort_values('date').copy()
        grp['rolling_3'] = grp['quantity_sold'].rolling(3, min_periods=1).mean()
        grp['rolling_7'] = grp['quantity_sold'].rolling(7, min_periods=1).mean()
        grp['rolling_14'] = grp['quantity_sold'].rolling(14, min_periods=1).mean()
        grp['recent_growth'] = ((grp['rolling_3'] + 1) / (grp['rolling_7'] + 1)) - 1
        latest = grp.iloc[-1].copy()
        current_stock = to_float(latest.get('current_stock'), np.nan)
        reorder_point = to_float(latest.get('reorder_point'), np.nan)
        if pd.isna(current_stock) or pd.isna(reorder_point):
            predictions[product_name] = {
                'risk_level': 'Unavailable',
                'probability': None,
                'top_factors': ['latest recorded stock or reorder point is missing'],
                'note': 'XGBoost could not score this product because inventory fields are incomplete.',
                'model_source': 'XGBoost',
                'suggested_action': 'Gather more sales history',
                'priority': 'Low',
                'stock_cover_days': None,
            }
            continue

        forecast_entry = sarima_summary.get(product_name, {})
        forecast_7 = forecast_entry.get('horizons', {}).get(7)
        if forecast_7 is None:
            forecast_7 = float(latest.get('rolling_7', 0) * 7)
        feature_row = {
            'rolling_3': float(latest.get('rolling_3', 0)),
            'rolling_7': float(latest.get('rolling_7', 0)),
            'rolling_14': float(latest.get('rolling_14', 0)),
            'recent_growth': float(latest.get('recent_growth', 0)),
            'is_payday': int(latest.get('is_payday', 0)),
            'is_weekend': int(latest.get('is_weekend', 0)),
            'day_of_week': int(latest.get('day_of_week', 0)),
            'month': int(pd.to_datetime(latest.get('date')).month),
            'current_stock': float(current_stock),
            'reorder_point': float(reorder_point),
            'stock_to_reorder_ratio': float(current_stock / reorder_point) if reorder_point else 0.0,
            'demand_stock_ratio': float(forecast_7 / current_stock) if current_stock else 0.0,
            'expected_next_7d': float(forecast_7),
            'category_code': int(pd.Series([latest.get('category', 'Uncategorized')]).astype('category').cat.codes.iloc[0]),
        }
        X_pred = pd.DataFrame([feature_row], columns=feature_cols).astype(float)
        probability = float(model.predict_proba(X_pred)[0][1])
        trend = 'Rising' if feature_row['recent_growth'] > 0.08 else ('Falling' if feature_row['recent_growth'] < -0.08 else 'Stable')
        factors = _derive_top_factor_notes(feature_importances, feature_row)
        calibrated = _calibrate_risk_outcome(
            probability=probability,
            current_stock=current_stock,
            reorder_point=reorder_point,
            forecast_demand=forecast_7,
            average_daily_demand=feature_row['rolling_7'],
            trend=trend,
            forecast_status=forecast_entry.get('status', 'limited'),
            factors=factors,
            risk_note='XGBoost classification based on stock position and recent demand.',
        )
        predictions[product_name] = {
            'risk_level': calibrated['risk_level'],
            'probability': round(probability, 4),
            'top_factors': factors,
            'note': calibrated['reason'],
            'model_source': 'XGBoost',
            'suggested_action': calibrated['suggested_action'],
            'priority': calibrated['priority'],
            'stock_cover_days': calibrated['stock_cover_days'],
        }
    return {
        'status': 'completed',
        'note': 'XGBoost stockout classification completed successfully.',
        'predictions': predictions,
        'feature_importances': feature_importances,
        'model_source': 'XGBoost',
    }

def _save_model_outputs_to_db(user_id: int, upload_id: int, daily_df: pd.DataFrame, sarima_runs: dict[str, Any], xgb_output: dict[str, Any]) -> dict[str, Any]:
    ensure_model_tables()
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        sarima_status = 'completed' if any(v.get('status') == 'completed' for v in sarima_runs.values()) else 'limited'
        xgboost_status = xgb_output.get('status', 'limited')
        overall_status = 'completed' if sarima_status == 'completed' or xgboost_status == 'completed' else 'limited'
        run_notes = []
        if sarima_status != 'completed':
            run_notes.append('Some forecast results are limited.')
        if xgboost_status != 'completed':
            run_notes.append(xgb_output.get('note') or 'Some stock risk results are limited.')
        if not run_notes:
            run_notes.append('Actual SARIMA and XGBoost outputs saved from the latest processing run.')

        cursor.execute(
            """
            INSERT INTO model_runs (user_id, upload_id, run_status, sarima_status, xgboost_status, sarima_version, xgboost_version, notes, completed_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
            """,
            (
                user_id,
                upload_id,
                overall_status,
                sarima_status,
                xgboost_status,
                'statsmodels-SARIMAX',
                'xgboost',
                ' '.join(note.strip() for note in run_notes if note).strip(),
            ),
        )
        model_run_id = cursor.lastrowid

        product_ids: dict[str, Any] = {}
        cursor.execute("SELECT product_id, product_name FROM products WHERE user_id = %s", (user_id,))
        for row in cursor.fetchall() or []:
            product_ids[row['product_name']] = row['product_id']

        cursor.execute("DELETE FROM forecasts WHERE user_id = %s AND upload_id = %s", (user_id, upload_id))
        cursor.execute("DELETE FROM stockout_predictions WHERE user_id = %s AND upload_id = %s", (user_id, upload_id))

        base_last_date = pd.to_datetime(daily_df['date']).max().date() if not daily_df.empty else None
        for product_name, result in sarima_runs.items():
            product_id = product_ids.get(product_name)
            daily_forecast = result.get('daily') or result.get('daily_forecast') or []
            if result.get('status') == 'completed' and daily_forecast and base_last_date is not None:
                for idx, forecast_qty in enumerate(daily_forecast, start=1):
                    forecast_date = pd.Timestamp(base_last_date) + pd.Timedelta(days=idx)
                    cursor.execute(
                        """
                        INSERT INTO forecasts (model_run_id, upload_id, user_id, product_id, product_name, forecast_date, horizon_days, forecast_quantity, model_source, status_label, note)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """,
                        (
                            model_run_id,
                            upload_id,
                            user_id,
                            product_id,
                            product_name,
                            forecast_date.date(),
                            idx,
                            float(forecast_qty),
                            result.get('model_source', 'SARIMA'),
                            'completed',
                            result.get('note'),
                        ),
                    )
            else:
                cursor.execute(
                    """
                    INSERT INTO forecasts (model_run_id, upload_id, user_id, product_id, product_name, forecast_date, horizon_days, forecast_quantity, model_source, status_label, note)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        model_run_id,
                        upload_id,
                        user_id,
                        product_id,
                        product_name,
                        None,
                        None,
                        None,
                        result.get('model_source', 'SARIMA'),
                        result.get('status', 'limited'),
                        result.get('note'),
                    ),
                )

        latest_rows = daily_df.sort_values('date').groupby('product_name', as_index=False).tail(1) if not daily_df.empty else pd.DataFrame()
        latest_lookup = {row['product_name']: row for _, row in latest_rows.iterrows()} if not latest_rows.empty else {}
        product_names = sorted(set(daily_df['product_name'].unique()) if not daily_df.empty else set(xgb_output.get('predictions', {}).keys()))
        for product_name in product_names:
            pred = xgb_output.get('predictions', {}).get(
                product_name,
                {
                    'risk_level': 'Unavailable',
                    'probability': None,
                    'top_factors': ['insufficient data'],
                    'note': xgb_output.get('note', 'No model output.'),
                    'model_source': 'XGBoost',
                },
            )
            latest_row = latest_lookup.get(product_name)
            pred_date = pd.to_datetime(latest_row['date']).date() if latest_row is not None and pd.notna(latest_row.get('date')) else None
            cursor.execute(
                """
                INSERT INTO stockout_predictions (model_run_id, upload_id, user_id, product_id, product_name, prediction_date, forecast_horizon_days, stockout_probability, risk_level, model_source, top_factors, note)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    model_run_id,
                    upload_id,
                    user_id,
                    product_ids.get(product_name),
                    product_name,
                    pred_date,
                    7,
                    float(pred.get('probability')) if pred.get('probability') is not None else None,
                    pred.get('risk_level', 'Unavailable'),
                    pred.get('model_source', 'XGBoost'),
                    '; '.join(pred.get('top_factors', [])),
                    pred.get('note'),
                ),
            )
        conn.commit()
        return {'model_run_id': model_run_id, 'run_status': overall_status, 'sarima_status': sarima_status, 'xgboost_status': xgboost_status}
    except Exception:
        conn.rollback()
        raise
    finally:
        cursor.close()
        conn.close()


def _load_latest_model_artifacts_from_db(user_id: int, upload_id: int | None = None) -> dict[str, Any]:
    if user_id is None:
        return {'available': False, 'forecast_by_product': {}, 'risk_by_product': {}, 'model_ui_summary': {}}
    if upload_id is None:
        meta = _load_latest_upload_metadata(user_id)
        upload_id = meta['upload_id'] if meta else None
    if upload_id is None:
        return {'available': False, 'forecast_by_product': {}, 'risk_by_product': {}, 'model_ui_summary': {}}
    ensure_model_tables()
    store_user_ids = get_store_user_ids(user_id)
    placeholders, params = make_in_clause(store_user_ids)
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute(
            f"""
            SELECT * FROM model_runs
            WHERE user_id IN ({placeholders}) AND upload_id = %s
            ORDER BY model_run_id DESC LIMIT 1
            """,
            (*params, upload_id)
        )
        model_run = cursor.fetchone()
        if not model_run:
            return {'available': False, 'forecast_by_product': {}, 'risk_by_product': {}, 'model_ui_summary': {}}
        cursor.execute(
            f"SELECT * FROM forecasts WHERE user_id IN ({placeholders}) AND upload_id = %s AND model_run_id = %s ORDER BY product_name, horizon_days",
            (*params, upload_id, model_run['model_run_id'])
        )
        forecast_rows = cursor.fetchall() or []
        cursor.execute(
            f"SELECT * FROM stockout_predictions WHERE user_id IN ({placeholders}) AND upload_id = %s AND model_run_id = %s ORDER BY product_name",
            (*params, upload_id, model_run['model_run_id'])
        )
        prediction_rows = cursor.fetchall() or []
    finally:
        cursor.close()
        conn.close()

    forecast_by_product: dict[str, Any] = {}
    for row in forecast_rows:
        product = row.get('product_name') or 'Unknown Product'
        entry = forecast_by_product.setdefault(product, {
            'status': row.get('status_label', 'limited'),
            'note': row.get('note'),
            'model_source': row.get('model_source', 'SARIMA'),
            'daily': [],
            'horizons': {7: None, 14: None, 28: None, 30: None, 180: None},
        })
        if row.get('forecast_date') is not None and row.get('forecast_quantity') is not None:
            entry['daily'].append({
                'date': row['forecast_date'],
                'quantity': float(row['forecast_quantity']),
                'horizon': int(row.get('horizon_days') or 0),
            })
        else:
            entry['status'] = row.get('status_label', entry['status'])
            entry['note'] = row.get('note') or entry.get('note')
    for product, entry in forecast_by_product.items():
        if entry['daily']:
            for horizon in [7, 14, 28, 30, 180]:
                matching = [item['quantity'] for item in entry['daily'] if (item.get('horizon') or 0) <= horizon]
                entry['horizons'][horizon] = round(sum(matching), 2) if matching else None

    risk_by_product: dict[str, Any] = {}
    for row in prediction_rows:
        risk_by_product[row.get('product_name') or 'Unknown Product'] = {
            'risk_level': row.get('risk_level', 'Unavailable') or 'Unavailable',
            'probability': float(row['stockout_probability']) if row.get('stockout_probability') is not None else None,
            'top_factors': [part.strip() for part in (row.get('top_factors') or '').split(';') if part.strip()],
            'note': row.get('note') or 'No model output.',
            'model_source': row.get('model_source', 'XGBoost'),
        }

    run_timestamp = model_run.get('completed_at') or model_run.get('started_at') or model_run.get('created_at')
    sarima_status = model_run.get('sarima_status', 'limited')
    xgboost_status = model_run.get('xgboost_status', 'limited')
    limited_note = None
    if str(sarima_status).strip().lower() == 'limited' and str(xgboost_status).strip().lower() == 'limited':
        limited_note = 'Some items still need more sales history before both model results are fully ready.'
    elif str(sarima_status).strip().lower() == 'limited':
        limited_note = 'Some forecast results still need more sales history.'
    elif str(xgboost_status).strip().lower() == 'limited':
        limited_note = 'Some stock risk results still need review because the data is limited.'

    model_ui_summary = {
        'available': True,
        'model_run_id': model_run['model_run_id'],
        'model_run_label': f"Run #{model_run['model_run_id']}",
        'model_run_timestamp': format_datetime(run_timestamp),
        'forecast_source_label': _friendly_model_source_labels()[0],
        'risk_source_label': _friendly_model_source_labels()[1],
        'sarima_status': sarima_status,
        'xgboost_status': xgboost_status,
        'sarima_status_label': _friendly_model_status_label(sarima_status, 'sarima'),
        'xgboost_status_label': _friendly_model_status_label(xgboost_status, 'xgboost'),
        'run_status': model_run.get('run_status', 'limited'),
        'limited_note': limited_note,
        'compact_badges': [
            _friendly_model_source_labels()[0],
            _friendly_model_source_labels()[1],
            f"Latest generated results: {format_datetime(run_timestamp)}" if run_timestamp else 'Latest generated results: latest processed data',
        ],
    }
    return {
        'available': True,
        'upload_id': upload_id,
        'model_run_id': model_run['model_run_id'],
        'forecast_by_product': forecast_by_product,
        'risk_by_product': risk_by_product,
        'model_ui_summary': model_ui_summary,
    }


def _load_latest_model_run_id_from_db(user_id: int | None, upload_id: int | None) -> int | None:
    """Return the newest model run id for the current store/upload.

    Model outputs can be regenerated for the same upload id. The display cache
    must therefore be validated against model_run_id, not upload_id only.
    """
    if user_id is None or upload_id is None:
        return None
    try:
        ensure_model_tables()
        store_user_ids = get_store_user_ids(user_id)
        placeholders, params = make_in_clause(store_user_ids)
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        try:
            cursor.execute(
                f"""
                SELECT model_run_id
                FROM model_runs
                WHERE user_id IN ({placeholders}) AND upload_id = %s
                ORDER BY model_run_id DESC
                LIMIT 1
                """,
                (*params, upload_id),
            )
            row = cursor.fetchone()
            return int(row["model_run_id"]) if row and row.get("model_run_id") is not None else None
        finally:
            cursor.close()
            conn.close()
    except Exception:
        return None


def get_model_artifacts() -> dict[str, Any]:
    _ensure_state_loaded_from_db()
    cached = _state_get('latest_model_artifacts')
    upload_id = _state_get('latest_upload_id')
    user_id = get_current_user_id()
    latest_model_run_id = _load_latest_model_run_id_from_db(user_id, upload_id)

    if (
        cached
        and cached.get('upload_id') == upload_id
        and cached.get('model_run_id') == latest_model_run_id
    ):
        return cached

    previous_model_run_id = _state_get('latest_model_run_id')
    try:
        artifacts = _load_latest_model_artifacts_from_db(user_id, upload_id)
    except Exception:
        artifacts = {'available': False, 'upload_id': upload_id, 'model_run_id': latest_model_run_id, 'forecast_by_product': {}, 'risk_by_product': {}, 'model_ui_summary': {}}

    new_model_run_id = artifacts.get('model_run_id')
    if previous_model_run_id and new_model_run_id and previous_model_run_id != new_model_run_id:
        clear_insights_cache()

    _state_set('latest_model_artifacts', artifacts)
    _state_set('model_ui_summary', artifacts.get('model_ui_summary'))
    _state_set('latest_model_run_id', new_model_run_id)
    return artifacts


def _empty_model_ui_summary() -> dict[str, Any]:
    return {
        'available': False,
        'model_run_label': 'No generated results yet',
        'model_run_timestamp': 'No uploaded sales data yet',
        'forecast_source_label': _friendly_model_source_labels()[0],
        'risk_source_label': _friendly_model_source_labels()[1],
        'sarima_status': 'pending',
        'xgboost_status': 'pending',
        'sarima_status_label': _friendly_model_status_label('pending', 'sarima'),
        'xgboost_status_label': _friendly_model_status_label('pending', 'xgboost'),
        'limited_note': None,
        'compact_badges': [
            _friendly_model_source_labels()[0],
            _friendly_model_source_labels()[1],
            'Latest generated results: not available yet',
        ],
    }


def get_model_ui_summary() -> dict[str, Any]:
    artifacts = get_model_artifacts()
    return artifacts.get('model_ui_summary') or _empty_model_ui_summary()


def clear_insights_cache() -> None:
    """Clear display-only Insights caches after a new upload, clear action, or model refresh."""
    for key in [
        "insights_daily_context_key",
        "insights_daily_context",
        "product_insights_cache",
        "dashboard_summary_cache_key",
        "dashboard_summary_cache",
        "forecast_summary_cache_key",
        "forecast_summary_cache",
        "stock_risk_summary_cache_key",
        "stock_risk_summary_cache",
        "product_summary_cache",
        "report_context_cache",
        "report_summary_cache",
    ]:
        _state_set(key, None)


def _dataset_cache_key(df: pd.DataFrame | None) -> tuple[Any, ...]:
    """Stable fingerprint for the current user's prepared dataset.

    The previous key included ``id(df)``, which changes when the same processed
    upload is reloaded into memory and can make the first page visit rebuild
    display summaries again.  Using the upload id, processed timestamp, and row
    count keeps caches tied to the latest generated results without depending on
    the Python object's memory address.
    """
    state = get_app_state()
    upload_id = getattr(state, "latest_upload_id", None)
    processed_at = str(getattr(state, "processed_at", None))
    store_id = get_current_store_id() or get_current_user_id()
    if df is None:
        return (store_id, "empty", upload_id, processed_at, 0)
    try:
        row_count = int(len(df))
    except Exception:
        row_count = 0
    return (
        store_id,
        upload_id,
        processed_at,
        row_count,
    )


def _analytics_cache_key(df: pd.DataFrame | None, forecast_days: int | None = None, artifacts: dict[str, Any] | None = None) -> tuple[Any, ...]:
    model_run_id = None
    if isinstance(artifacts, dict):
        model_run_id = artifacts.get("model_run_id")
    if model_run_id is None:
        model_run_id = getattr(get_app_state(), "latest_model_run_id", None)
    return _dataset_cache_key(df) + (forecast_days, model_run_id)


def _empty_forecast_summary() -> dict[str, Any]:
    return {
        "products": [],
        "categories": [],
        "default_product": "__total__",
        "range_options": get_forecast_view_options(),
        "chart_map": {"__total__": {}},
        "category_chart_map": {},
        "summary_map": {"__total__": {}},
        "category_summary_map": {},
        "details_by_range": {},
        "priority_rows": [],
        "filter_metrics": {},
        "initial_view": None,
        "model_ui_summary": get_model_ui_summary(),
    }


def _empty_stock_risk_summary() -> dict[str, Any]:
    return {
        "inventory_rows": [],
        "has_data": False,
        "high_count": 0,
        "needs_restock_count": 0,
        "moderate_count": 0,
        "safe_count": 0,
        "top_high_risk": [],
        "fastest_rising": [],
        "priority_reorder": [],
        "main_risk_drivers": ["Upload sales data to view the main stockout risk drivers."],
        "top_high_risk_empty_message": "No uploaded sales data yet.",
        "fastest_rising_empty_message": "No uploaded sales data yet.",
        "priority_reorder_empty_message": "No uploaded sales data yet.",
        "model_ui_summary": get_model_ui_summary(),
    }


def _get_daily_sales_context(df: pd.DataFrame | None) -> dict[str, Any]:
    """Prepare reusable daily sales series for Insights charts and summaries.

    The Insights page calls several helpers in one request. Reusing this context avoids
    repeatedly copying the full dataset and grouping the same dates/products.
    """
    empty_context = {
        "product_daily": {},
        "category_daily": {},
        "category_products": {},
        "product_meta": {},
        "total_daily": pd.Series(dtype=float),
        "products": [],
        "categories_present": [],
    }
    if df is None or df.empty or "product_name" not in df.columns or "date" not in df.columns:
        return empty_context

    cache_key = _dataset_cache_key(df)
    cached_key = _state_get("insights_daily_context_key")
    cached_context = _state_get("insights_daily_context")
    if cached_key == cache_key and isinstance(cached_context, dict):
        return cached_context

    working = df.copy()
    working["date"] = pd.to_datetime(working["date"], errors="coerce")
    working = working.dropna(subset=["date", "product_name"])
    if working.empty:
        _state_set("insights_daily_context_key", cache_key)
        _state_set("insights_daily_context", empty_context)
        return empty_context

    if "quantity_sold" not in working.columns:
        working["quantity_sold"] = 0
    working["quantity_sold"] = pd.to_numeric(working["quantity_sold"], errors="coerce").fillna(0)

    if "category" not in working.columns:
        working["category"] = "Seasonal / Miscellaneous Items"
    working["category"] = working["category"].fillna("Seasonal / Miscellaneous Items").apply(standardize_product_category)

    for col in ["current_stock", "reorder_point", "unit_price"]:
        if col not in working.columns:
            working[col] = None
        working[col] = pd.to_numeric(working[col], errors="coerce")
    if "unit_type" not in working.columns:
        working["unit_type"] = "Unit"

    working["normalized_date"] = working["date"].dt.normalize()
    daily = (
        working.groupby(["product_name", "normalized_date"], as_index=False)["quantity_sold"]
        .sum()
    )

    product_daily: dict[str, pd.Series] = {}
    for product_name, product_rows in daily.groupby("product_name"):
        series = (
            product_rows.set_index("normalized_date")["quantity_sold"]
            .sort_index()
            .astype(float)
        )
        product_daily[str(product_name)] = series

    category_daily: dict[str, pd.Series] = {}
    category_products: dict[str, list[str]] = {}
    for category, category_rows in working.groupby("category"):
        category_name = str(category)
        category_products[category_name] = sorted(
            category_rows["product_name"].dropna().astype(str).unique().tolist(),
            key=str.casefold,
        )
        category_daily[category_name] = (
            category_rows.groupby(category_rows["date"].dt.normalize())["quantity_sold"]
            .sum()
            .sort_index()
            .astype(float)
        )

    total_daily = (
        working.groupby(working["date"].dt.normalize())["quantity_sold"]
        .sum()
        .sort_index()
        .astype(float)
    )

    product_meta: dict[str, dict[str, Any]] = {}
    sorted_working = working.sort_values(["product_name", "date"])
    for product_name, product_rows in sorted_working.groupby("product_name"):
        product_key = str(product_name)
        meta = {
            "category": "Seasonal / Miscellaneous Items",
            "current_stock": None,
            "reorder_point": None,
            "unit_price": None,
            "unit_type": "Unit",
        }

        category_values = product_rows["category"].dropna().astype(str)
        if not category_values.empty:
            meta["category"] = standardize_product_category(category_values.iloc[-1])

        stock_values = pd.to_numeric(product_rows["current_stock"], errors="coerce").dropna()
        if not stock_values.empty:
            meta["current_stock"] = float(stock_values.iloc[-1])

        reorder_values = pd.to_numeric(product_rows["reorder_point"], errors="coerce").dropna()
        if not reorder_values.empty:
            meta["reorder_point"] = float(reorder_values.iloc[-1])

        price_values = pd.to_numeric(product_rows["unit_price"], errors="coerce").dropna()
        if not price_values.empty:
            meta["unit_price"] = float(price_values.iloc[-1])

        unit_type_values = product_rows["unit_type"].dropna().astype(str).str.strip()
        unit_type_values = unit_type_values[unit_type_values != ""]
        if not unit_type_values.empty:
            meta["unit_type"] = unit_type_values.iloc[-1]

        product_meta[product_key] = meta

    context = {
        "product_daily": product_daily,
        "category_daily": category_daily,
        "category_products": category_products,
        "product_meta": product_meta,
        "total_daily": total_daily,
        "products": sorted(product_daily.keys(), key=str.casefold),
        "categories_present": sorted(category_daily.keys(), key=str.casefold),
    }
    _state_set("insights_daily_context_key", cache_key)
    _state_set("insights_daily_context", context)
    return context


def _build_lightweight_model_outputs(daily_df: pd.DataFrame, horizon: int = 180) -> tuple[dict[str, Any], dict[str, Any]]:
    """Generate fast, database-backed results when hosted CPU/request limits are tight.

    This prevents a medium upload from blocking Render while SARIMA is fitted for
    many products.  The UI still receives forecast/risk rows immediately, and the
    notes clearly label them as lightweight results.
    """
    sarima_summary: dict[str, Any] = {}
    predictions: dict[str, Any] = {}
    if daily_df is None or daily_df.empty or 'product_name' not in daily_df.columns:
        return sarima_summary, {
            'status': 'limited',
            'note': 'Lightweight model outputs could not be generated because no daily sales frame was available.',
            'predictions': {},
            'feature_importances': {},
            'model_source': 'Fast Risk Rules',
        }

    for product_name, grp in daily_df.groupby('product_name'):
        product_key = str(product_name)
        grp = grp.sort_values('date').copy()
        quantity = pd.to_numeric(grp.get('quantity_sold'), errors='coerce').fillna(0.0)
        if quantity.empty:
            daily_average = 0.0
        else:
            recent_window = quantity.tail(min(14, len(quantity)))
            daily_average = max(float(recent_window.mean()), 0.0)
        forecast_daily = [round(daily_average, 4)] * horizon
        sarima_summary[product_key] = {
            'status': 'completed' if daily_average > 0 else 'limited',
            'note': 'Fast hosted forecast generated from recent average daily sales. Full SARIMA was skipped to keep Render responsive.',
            'model_source': 'Fast Forecast',
            'daily': forecast_daily,
            'horizons': {
                7: round(sum(forecast_daily[:7]), 2),
                14: round(sum(forecast_daily[:14]), 2),
                28: round(sum(forecast_daily[:28]), 2),
                30: round(sum(forecast_daily[:30]), 2),
                180: round(sum(forecast_daily[:180]), 2),
            },
        }

        latest = grp.iloc[-1].copy()
        current_stock = to_float(latest.get('current_stock'), np.nan)
        reorder_point = to_float(latest.get('reorder_point'), np.nan)
        rolling_7 = max(float(quantity.tail(min(7, len(quantity))).mean()) if not quantity.empty else 0.0, 0.0)
        rolling_3 = max(float(quantity.tail(min(3, len(quantity))).mean()) if not quantity.empty else 0.0, 0.0)
        forecast_7 = sarima_summary[product_key]['horizons'][7] or 0.0
        trend = 'Rising' if rolling_3 > rolling_7 * 1.08 else ('Falling' if rolling_3 < rolling_7 * 0.92 else 'Stable')

        if pd.isna(current_stock) or pd.isna(reorder_point):
            predictions[product_key] = {
                'risk_level': 'Unavailable',
                'probability': None,
                'top_factors': ['latest recorded stock or reorder point is missing'],
                'note': 'Fast risk rules could not score this product because inventory fields are incomplete.',
                'model_source': 'Fast Risk Rules',
                'suggested_action': 'Gather more sales history',
                'priority': 'Low',
                'stock_cover_days': None,
            }
            continue

        stock_pressure = 0.0
        if current_stock <= reorder_point:
            stock_pressure += 0.45
        if current_stock and forecast_7 > current_stock:
            stock_pressure += 0.35
        if trend == 'Rising':
            stock_pressure += 0.15
        probability = max(0.05, min(0.95, stock_pressure or 0.2))
        factors = []
        if current_stock <= reorder_point:
            factors.append('stock is at or below the reorder point')
        if current_stock and forecast_7 > current_stock:
            factors.append('predicted demand may exceed available stock')
        if trend == 'Rising':
            factors.append('recent demand is rising')
        if not factors:
            factors.append('risk is based on latest stock, reorder point, and recent demand')

        calibrated = _calibrate_risk_outcome(
            probability=probability,
            current_stock=float(current_stock),
            reorder_point=float(reorder_point),
            forecast_demand=float(forecast_7),
            average_daily_demand=float(rolling_7),
            trend=trend,
            forecast_status=sarima_summary[product_key]['status'],
            factors=factors,
            risk_note='Fast hosted risk rules based on stock position and recent demand.',
        )
        predictions[product_key] = {
            'risk_level': calibrated['risk_level'],
            'probability': round(float(probability), 4),
            'top_factors': factors[:3],
            'note': calibrated['reason'],
            'model_source': 'Fast Risk Rules',
            'suggested_action': calibrated['suggested_action'],
            'priority': calibrated['priority'],
            'stock_cover_days': calibrated['stock_cover_days'],
        }

    return sarima_summary, {
        'status': 'completed',
        'note': 'Fast hosted stockout risk rules completed. Full XGBoost/SARIMA was skipped for this upload to avoid Render request timeouts.',
        'predictions': predictions,
        'feature_importances': {
            'current_stock': 0.35,
            'reorder_point': 0.25,
            'expected_next_7d': 0.25,
            'recent_growth': 0.15,
        },
        'model_source': 'Fast Risk Rules',
    }


def run_model_pipeline(user_id: int, upload_id: int, processed_df: pd.DataFrame) -> dict[str, Any]:
    daily_df = _prepare_daily_model_frame(processed_df)
    if daily_df.empty:
        artifacts = {
            'available': False,
            'forecast_by_product': {},
            'risk_by_product': {},
            'model_ui_summary': {
                'available': False,
                'model_run_label': 'No generated results yet',
                'model_run_timestamp': 'No uploaded sales data yet',
                'forecast_source_label': _friendly_model_source_labels()[0],
                'risk_source_label': _friendly_model_source_labels()[1],
                'sarima_status': 'limited',
                'xgboost_status': 'limited',
                'sarima_status_label': _friendly_model_status_label('limited', 'sarima'),
                'xgboost_status_label': _friendly_model_status_label('limited', 'xgboost'),
                'limited_note': 'Some items still need more sales history before model results are fully ready.',
                'compact_badges': [
                    _friendly_model_source_labels()[0],
                    _friendly_model_source_labels()[1],
                    'Latest generated results: not available yet',
                ],
            }
        }
        _state_set('latest_model_artifacts', artifacts)
        clear_insights_cache()
        return artifacts

    product_count = int(daily_df['product_name'].nunique()) if 'product_name' in daily_df.columns else 0
    uploaded_rows = len(processed_df) if processed_df is not None else len(daily_df)
    use_lightweight_models = uploaded_rows > MAX_SYNC_MODEL_ROWS or product_count > MAX_SYNC_SARIMA_PRODUCTS

    if use_lightweight_models:
        sarima_summary, xgb_output = _build_lightweight_model_outputs(daily_df, horizon=180)
    else:
        started_at = datetime.now()
        sarima_results = {}
        for product_name, grp in daily_df.groupby('product_name'):
            if (datetime.now() - started_at).total_seconds() > MAX_SYNC_MODEL_SECONDS:
                use_lightweight_models = True
                break
            sarima_results[product_name] = _fit_sarima_forecast(grp.sort_values('date')['quantity_sold'], horizon=180)

        if use_lightweight_models:
            sarima_summary, xgb_output = _build_lightweight_model_outputs(daily_df, horizon=180)
        else:
            sarima_summary = {}
            for product_name, result in sarima_results.items():
                if result.get('status') == 'completed' and result.get('daily_forecast'):
                    daily = result['daily_forecast']
                    sarima_summary[product_name] = {
                        'status': result['status'],
                        'note': result['note'],
                        'model_source': result['model_source'],
                        'daily': daily,
                        'horizons': {7: round(sum(daily[:7]),2), 14: round(sum(daily[:14]),2), 28: round(sum(daily[:28]),2), 30: round(sum(daily[:30]),2), 180: round(sum(daily[:180]),2)},
                    }
                else:
                    sarima_summary[product_name] = {
                        'status': result.get('status','limited'),
                        'note': result.get('note'),
                        'model_source': 'SARIMA',
                        'daily': [],
                        'horizons': {7: None, 14: None, 28: None, 30: None, 180: None},
                    }
            xgb_output = _train_xgboost_and_predict(daily_df, sarima_summary)

    save_info = _save_model_outputs_to_db(user_id, upload_id, daily_df, sarima_summary, xgb_output)
    artifacts = _load_latest_model_artifacts_from_db(user_id, upload_id)
    artifacts['model_run_id'] = save_info.get('model_run_id')
    _state_set('latest_model_artifacts', artifacts)
    _state_set('latest_model_run_id', save_info.get('model_run_id'))
    clear_insights_cache()
    return artifacts


def get_product_chart_data(df: pd.DataFrame | None, product_name: str, forecast_days: int = 7, history_days: int = 14, view_key: str = "daily") -> dict[str, Any]:
    empty = {
        "chart_type": "line",
        "chart_labels": [],
        "chart_datasets": [],
        "chart_message": "No sales history is available for this item yet.",
        "has_forecast": False,
        "forecast_start_label": None,
        "chart_options": _forecast_chart_options(view_key),
    }
    if df is None or not product_name:
        return empty

    context = _get_daily_sales_context(df)
    daily_totals = context.get("product_daily", {}).get(product_name)
    if daily_totals is None or daily_totals.empty:
        return empty

    daily_totals = daily_totals[daily_totals >= 0].sort_index()
    if daily_totals.empty:
        return empty

    view_config = _forecast_view_config(view_key)
    group_by = view_config.get("group_by", "daily")
    historical_daily = daily_totals.tail(max(history_days, 7))
    historical = _aggregate_series_for_forecast_view(historical_daily, view_key)
    historical = _limit_forecast_view_points(historical, view_key, "history")
    labels = [_format_forecast_axis_label(d, group_by) for d in historical.index]
    historical_data = [round(float(v), 2) for v in historical.tolist()]
    chart = {
        "chart_type": "line",
        "chart_labels": labels,
        "chart_datasets": [{
            "label": "Observed Sales",
            "data": historical_data,
            "borderColor": "#f4d35e",
            "backgroundColor": "rgba(244, 211, 94, 0.18)",
            "tension": 0.32,
            "fill": True,
        }],
        "chart_message": None,
        "has_forecast": False,
        "forecast_start_label": None,
        "chart_options": _forecast_chart_options(view_key),
    }

    artifacts = get_model_artifacts()
    forecast_entry = artifacts.get("forecast_by_product", {}).get(product_name, {}) if isinstance(artifacts, dict) else {}
    forecast_status = str(forecast_entry.get("status", "") or "").strip().lower() if isinstance(forecast_entry, dict) else ""
    forecast_daily = forecast_entry.get("daily", []) if isinstance(forecast_entry, dict) else []

    valid_forecast_points = []
    if forecast_status == "completed":
        for item in (forecast_daily or [])[:forecast_days]:
            item_date = pd.to_datetime(item.get("date"), errors="coerce") if isinstance(item, dict) else pd.NaT
            item_qty = pd.to_numeric(item.get("quantity"), errors="coerce") if isinstance(item, dict) else None
            if pd.notna(item_date) and item_qty is not None and pd.notna(item_qty):
                valid_forecast_points.append({"date": item_date, "quantity": round(float(item_qty), 2)})

    if valid_forecast_points:
        future_series = pd.Series(
            [item["quantity"] for item in valid_forecast_points],
            index=[item["date"] for item in valid_forecast_points],
            dtype=float,
        )
        future_grouped = _aggregate_series_for_forecast_view(future_series, view_key)
        future_grouped = _limit_forecast_view_points(future_grouped, view_key, "forecast")
        future_dates = list(future_grouped.index)
        future_labels = [_format_forecast_axis_label(d, group_by) for d in future_dates]
        forecast_values = [round(float(value), 2) for value in future_grouped.tolist()]
        observed_data, forecast_data = _build_forecast_line_datasets(historical_data, forecast_values, len(future_labels))
        chart["chart_labels"] = labels + future_labels
        chart["chart_datasets"][0]["data"] = observed_data
        chart["chart_datasets"].append({
            "label": "Predicted Demand",
            "data": forecast_data,
            "borderColor": "#8d6e63",
            "backgroundColor": "rgba(141, 110, 99, 0.08)",
            "borderDash": [6, 4],
            "tension": 0.32,
            "fill": False,
        })
        chart["has_forecast"] = True
        if future_dates:
            chart["forecast_start_label"] = _format_forecast_axis_label(future_dates[0], group_by)
    else:
        note = str(forecast_entry.get("note") or "").strip().lower() if isinstance(forecast_entry, dict) else ""
        if forecast_status in {"limited", "failed", "unavailable"} or "insufficient" in note:
            chart["chart_message"] = "Forecast line not available for this product yet."

    return chart

def get_dashboard_chart_data(df: pd.DataFrame | None) -> dict[str, Any]:
    empty = {
        'chart_type': 'line',
        'chart_labels': [],
        'chart_datasets': [],
        'chart_message': 'No chart data available yet.',
        'chart_explanation': None,
        'chart_focus_label': 'Overall Demand',
        'chart_options': {
            'layout': {'padding': {'top': 2, 'right': 8, 'bottom': 2, 'left': 6}},
            'plugins': {
                'legend': {'position': 'top', 'labels': {'usePointStyle': True, 'boxWidth': 8, 'padding': 10}},
            },
            'scales': {
                'x': {'grid': {'display': False}, 'ticks': {'maxRotation': 0, 'minRotation': 0, 'autoSkip': True, 'padding': 3}},
                'y': {'beginAtZero': True, 'grid': {'color': '#f0f0f0'}, 'ticks': {'padding': 8}},
            },
        },
        'has_forecast': False,
        'forecast_start_label': None,
    }
    if df is None or df.empty or 'date' not in df.columns or 'quantity_sold' not in df.columns:
        return empty

    working = df.copy()
    working['date'] = pd.to_datetime(working['date'], errors='coerce')
    working['quantity_sold'] = pd.to_numeric(working['quantity_sold'], errors='coerce').fillna(0)
    working = working.dropna(subset=['date'])
    if working.empty:
        return empty

    daily_totals = working.groupby(working['date'].dt.normalize())['quantity_sold'].sum().sort_index()
    daily_totals = daily_totals[daily_totals > 0]
    if daily_totals.empty:
        return empty

    historical = daily_totals.tail(7)
    labels = [d.strftime('%b %d') for d in historical.index]
    historical_data = [round(float(v), 2) for v in historical.tolist()]
    chart = {
        'chart_type': 'line',
        'chart_labels': labels,
        'chart_datasets': [{
            'label': 'Observed Sales',
            'data': historical_data,
            'borderColor': '#f4d35e',
            'backgroundColor': 'rgba(244, 211, 94, 0.18)',
            'tension': 0.32,
            'fill': True,
        }],
        'chart_message': None,
        'chart_explanation': None,
        'chart_focus_label': 'Overall Demand',
        'chart_options': {
            'layout': {'padding': {'top': 2, 'right': 8, 'bottom': 2, 'left': 6}},
            'plugins': {
                'legend': {'position': 'top', 'labels': {'usePointStyle': True, 'boxWidth': 8, 'padding': 10}},
            },
            'scales': {
                'x': {'grid': {'display': False}, 'ticks': {'maxRotation': 0, 'minRotation': 0, 'autoSkip': True, 'padding': 3}},
                'y': {'beginAtZero': True, 'grid': {'color': '#f0f0f0'}, 'ticks': {'padding': 8}},
            },
        },
        'has_forecast': False,
        'forecast_start_label': None,
    }

    artifacts = get_model_artifacts()
    forecast_map = artifacts.get('forecast_by_product', {}) if isinstance(artifacts, dict) else {}
    future_totals: dict[pd.Timestamp, float] = {}
    for forecast_entry in forecast_map.values():
        if not isinstance(forecast_entry, dict):
            continue
        if str(forecast_entry.get('status', '') or '').strip().lower() != 'completed':
            continue
        for item in (forecast_entry.get('daily') or [])[:7]:
            item_date = pd.to_datetime(item.get('date'), errors='coerce') if isinstance(item, dict) else pd.NaT
            item_qty = pd.to_numeric(item.get('quantity'), errors='coerce') if isinstance(item, dict) else None
            if pd.notna(item_date) and item_qty is not None and pd.notna(item_qty):
                normalized_date = item_date.normalize()
                future_totals[normalized_date] = future_totals.get(normalized_date, 0.0) + float(item_qty)

    if future_totals:
        future_dates = sorted(future_totals.keys())[:7]
        future_labels = [d.strftime('%b %d') for d in future_dates]
        forecast_values = [round(float(future_totals[d]), 2) for d in future_dates]
        observed_data, forecast_data = _build_forecast_line_datasets(historical_data, forecast_values, len(future_labels))
        chart['chart_labels'] = labels + future_labels
        chart['chart_datasets'][0]['data'] = observed_data
        chart['chart_datasets'].append({
            'label': 'Predicted Demand',
            'data': forecast_data,
            'borderColor': '#8d6e63',
            'backgroundColor': 'rgba(141, 110, 99, 0.08)',
            'borderDash': [6, 4],
            'tension': 0.32,
            'fill': False,
        })
        chart['has_forecast'] = True
        chart['forecast_start_label'] = future_dates[0].strftime('%b %d')
    else:
        chart['chart_message'] = None

    return chart


def _friendly_model_source_labels() -> tuple[str, str]:
    return "Predicted Demand", "Stockout Risk"


def _friendly_model_status_label(status: str | None, model: str) -> str:
    status_key = (status or '').strip().lower()
    if model == 'sarima':
        mapping = {
            'completed': 'Forecast ready',
            'limited': 'Need more sales history',
            'failed': 'Forecast not ready',
            'started': 'Processing',
            'pending': 'Pending',
            'unavailable': 'Forecast not ready',
        }
        return mapping.get(status_key, 'Forecast not ready')
    mapping = {
        'completed': 'Risk ready',
        'limited': 'Risk needs review',
        'failed': 'Risk needs review',
        'started': 'Processing',
        'pending': 'Pending',
        'unavailable': 'Risk needs review',
    }
    return mapping.get(status_key, 'Risk needs review')


def _friendly_factor_phrase(text: str) -> str:
    value = (text or '').strip().lower()
    mapping = {
        'predicted demand may exceed available stock': 'Predicted demand may be higher than available stock.',
        'stock is at or below the reorder point': 'Stock is near the reorder point.',
        'recent demand is rising': 'Demand has been rising recently.',
        'payday-related demand pattern detected': 'Payday sales patterns were noticed.',
        'weekend demand pattern detected': 'Weekend sales patterns were noticed.',
        'recent average demand remains active': 'Sales have stayed active recently.',
        'risk is based on latest recorded stock, reorder point, and recent demand behavior': 'This item was reviewed using stock and demand patterns.',
        'latest recorded stock or reorder point is missing': 'Stock details need review.',
        'insufficient data': 'More sales history is needed for this item.',
        'insufficient history for sarima forecasting.': 'More sales history is needed for this item.',
    }
    if value in mapping:
        return mapping[value]
    cleaned = (text or '').strip()
    if not cleaned:
        return 'This item needs review.'
    cleaned = cleaned[0].upper() + cleaned[1:]
    if not cleaned.endswith('.'):
        cleaned += '.'
    return cleaned


def _friendly_forecast_value(forecast_demand: Any, forecast_status: str | None) -> str | int:
    if forecast_demand is None or (isinstance(forecast_demand, float) and pd.isna(forecast_demand)):
        return 'Forecast not ready'
    try:
        return int(math.ceil(float(forecast_demand)))
    except Exception:
        return 'Forecast not ready'


def _friendly_forecast_note(forecast_status: str | None, forecast_note: str | None) -> str:
    status_key = (forecast_status or '').strip().lower()
    if status_key == 'completed':
        return 'Forecast uses your recent sales history for this period.'
    if forecast_note and 'insufficient' in forecast_note.lower():
        return 'More sales history is needed for this item.'
    return 'Forecast is not ready for this item yet.'


def _friendly_risk_reason(risk_level: str | None, factors: list[str] | None, risk_note: str | None, forecast_status: str | None) -> str:
    factor_list = [
        _friendly_factor_phrase(item)
        for item in (factors or [])
        if str(item).strip()
    ]
    if factor_list:
        return ' '.join(dict.fromkeys(factor_list))
    if risk_note and 'limited' in risk_note.lower():
        return 'The current data is too limited for a strong stockout estimate.'
    if (forecast_status or '').strip().lower() != 'completed':
        return 'More sales history is needed for this item.'
    if (risk_level or '').strip().lower() == 'high':
        return 'This item may need attention soon.'
    if (risk_level or '').strip().lower() == 'moderate':
        return 'This item should be monitored closely.'
    if (risk_level or '').strip().lower() == 'low':
        return 'Stock currently looks sufficient.'
    return 'This item needs review.'


def _friendly_priority_label(priority: str | None, risk_level: str | None) -> str:
    key = (priority or '').strip().lower()
    if key in {'high', 'moderate', 'low'}:
        return key.title()
    level = (risk_level or '').strip().lower()
    if level in {'high', 'moderate', 'low'}:
        return level.title()
    return 'Needs review'


def _friendly_suggested_action_label(action: str | None) -> str:
    key = (action or '').strip().lower()
    mapping = {
        'restock now': 'Consider restocking soon',
        'consider restocking soon': 'Consider restocking soon',
        'prepare reorder soon': 'Review before next restocking',
        'review before next restocking': 'Review before next restocking',
        'monitor this week': 'Monitor only',
        'monitor this item': 'Monitor only',
        'monitor only': 'Monitor only',
        'sufficient stock': 'Stock appears sufficient',
        'sufficient stock for now': 'Stock appears sufficient',
        'stock appears sufficient': 'Stock appears sufficient',
        'needs more history': 'Gather more sales history',
        'gather more sales history': 'Gather more sales history',
    }
    return mapping.get(key, action or 'Needs review')


def _friendly_probability_note(probability: Any) -> str | None:
    if probability is None:
        return None
    try:
        return f"Stockout estimate: {round(float(probability) * 100, 1)}%"
    except Exception:
        return None


def _friendly_table_note(risk_reason: str, forecast_status: str | None, probability: Any) -> str:
    forecast_hint = _friendly_forecast_note(forecast_status, None)
    if (forecast_status or '').strip().lower() == 'completed':
        return risk_reason
    if forecast_hint == risk_reason:
        return forecast_hint
    return f"{risk_reason} {forecast_hint}".strip()


def _format_unit_gap(value: Any) -> str:
    try:
        number = float(value)
    except Exception:
        return ""
    rounded = int(math.ceil(abs(number)))
    return f"{rounded} unit" if rounded == 1 else f"{rounded} units"


def _build_demand_detail_note(
    *,
    forecast_status: str | None,
    forecast_demand: Any,
    current_stock: Any,
    reorder_point: Any,
    risk_level: str | None,
    trend: str | None,
    fallback_reason: str,
    probability: Any = None,
) -> str:
    """Create product-specific Demand Details notes using available computed values only."""
    if (forecast_status or "").strip().lower() != "completed" or forecast_demand is None:
        return _friendly_table_note(fallback_reason, forecast_status, probability)

    try:
        forecast_value = float(forecast_demand)
    except Exception:
        return _friendly_table_note(fallback_reason, forecast_status, probability)

    stock_value = None
    reorder_value = None
    try:
        if current_stock is not None and not pd.isna(current_stock):
            stock_value = float(current_stock)
    except Exception:
        stock_value = None
    try:
        if reorder_point is not None and not pd.isna(reorder_point):
            reorder_value = float(reorder_point)
    except Exception:
        reorder_value = None

    trend_label = (trend or "").strip().lower()
    risk_label = (risk_level or "").strip().lower()

    if stock_value is None:
        if trend_label == "rising":
            return "Predicted demand is available and recent demand is rising, so review this product before the next selling period."
        return "Predicted demand is available, but the latest stock value is missing. Review this product once stock is updated."

    stock_gap = forecast_value - stock_value
    if stock_gap > 0:
        gap_label = _format_unit_gap(stock_gap)
        if risk_label == "high":
            return f"Predicted demand is about {gap_label} higher than latest recorded stock, so this product may need review before the next selling period."
        if risk_label == "moderate":
            return f"Predicted demand is about {gap_label} higher than latest recorded stock. Monitor this product and prepare stock if demand continues."
        return f"Predicted demand is slightly above latest recorded stock. Review this product if sales remain active."

    if reorder_value is not None and stock_value <= reorder_value:
        return "Latest stock is at or below the reorder point, so this product should still be reviewed even if the current forecast is manageable."

    if trend_label == "rising":
        return "Latest stock can cover the current forecast, but demand is rising, so continue monitoring this product closely."

    if forecast_value > 0 and stock_value >= forecast_value * 1.5:
        return "Latest stock is above predicted demand, so immediate restocking may not be needed. Continue monitoring regular sales movement."

    return "Latest stock can cover the current forecast. Continue monitoring this product during the next selling period."


def _short_display_note(note: Any, max_chars: int = 132) -> str:
    text = str(note or '').strip()
    if not text:
        return 'Needs review.'
    first = text.split('.') [0].strip() if '.' in text else text
    short = first if len(first) >= 18 else text
    if len(short) > max_chars:
        short = short[: max_chars - 1].rstrip() + '…'
    if short and short[-1] not in '.!?':
        short += '.'
    return short


def _dashboard_main_reason(note: Any, fallback: Any = None, max_chars: int = 120) -> str:
    """Return a brief dashboard-only reason without changing detailed notes elsewhere."""
    text = str(note or fallback or '').strip()
    if not text:
        return 'Needs review.'

    lowered = text.lower()
    cut_index = None
    for marker in (', so ', ' so ', '. '):
        idx = lowered.find(marker)
        if idx > 0:
            cut_index = idx if cut_index is None else min(cut_index, idx)

    if cut_index is not None:
        text = text[:cut_index].strip()

    if len(text) > max_chars:
        text = text[: max_chars - 1].rstrip() + '…'

    if text and text[-1] not in '.!?':
        text += '.'
    return text or 'Needs review.'


def _detail_factor_list(raw_factors: list[str] | None, probability_label: str | None, stock_cover_label: str | None) -> list[str]:
    items: list[str] = []
    for factor in raw_factors or []:
        cleaned = _friendly_factor_phrase(factor)
        if cleaned and cleaned not in items:
            items.append(cleaned)
    if stock_cover_label and stock_cover_label not in items:
        items.append(stock_cover_label)
    if probability_label and probability_label not in items:
        items.append(probability_label)
    return items[:5]


def compute_product_insights(df: pd.DataFrame | None, forecast_days: int = 7) -> list[dict[str, Any]]:
    if df is None or "product_name" not in df.columns or "quantity_sold" not in df.columns or "date" not in df.columns:
        return []

    artifacts = get_model_artifacts()
    cache_key = _analytics_cache_key(df, forecast_days=forecast_days, artifacts=artifacts)
    insights_cache = _state_get("product_insights_cache") or {}
    if cache_key in insights_cache:
        cached_rows = [row.copy() for row in insights_cache[cache_key]]
        for row in cached_rows:
            if "dashboard_main_reason" not in row:
                row["dashboard_main_reason"] = _dashboard_main_reason(
                    row.get("display_note"),
                    row.get("why_flagged"),
                )
            if "stock_status" not in row:
                row["stock_status"] = _product_stock_status_from_decision(row.get("suggested_action", ""))
            if "assistant_stock_status" not in row:
                row["assistant_stock_status"] = _assistant_stock_status_from_store_status(row.get("stock_status"))
        return cached_rows

    context = _get_daily_sales_context(df)
    product_daily = context.get("product_daily", {})
    product_meta = context.get("product_meta", {})
    if not product_daily:
        return []

    forecast_map = artifacts.get("forecast_by_product", {}) if isinstance(artifacts, dict) else {}
    risk_map = artifacts.get("risk_by_product", {}) if isinstance(artifacts, dict) else {}

    rows: list[dict[str, Any]] = []
    for product_name, series in product_daily.items():
        if series is None or series.empty:
            continue

        product_rows = series.sort_index()
        last_date = product_rows.index.max()
        recent_start = last_date - pd.Timedelta(days=6)
        previous_start = last_date - pd.Timedelta(days=13)
        previous_end = last_date - pd.Timedelta(days=7)

        recent_rows = product_rows[product_rows.index >= recent_start]
        previous_rows = product_rows[(product_rows.index >= previous_start) & (product_rows.index <= previous_end)]

        recent_total = float(recent_rows.sum())
        previous_total = float(previous_rows.sum())
        recent_day_count = int(recent_rows.index.nunique())
        previous_day_count = int(previous_rows.index.nunique())
        recent_days = max(recent_day_count, 1)
        average_daily_demand = recent_total / recent_days if recent_days else 0
        previous_average_daily = previous_total / previous_day_count if previous_day_count else None

        growth_pct = None
        comparable_history = recent_day_count >= 3 and previous_day_count >= 3
        if comparable_history and previous_average_daily is not None:
            if previous_average_daily <= 0 and average_daily_demand > 0:
                growth_pct = 1.0
            elif previous_average_daily > 0:
                growth_pct = (average_daily_demand - previous_average_daily) / previous_average_daily

        if growth_pct is not None and growth_pct >= 0.08 and (average_daily_demand - (previous_average_daily or 0)) >= 1:
            trend = "Rising"
        elif growth_pct is not None and growth_pct <= -0.08 and ((previous_average_daily or 0) - average_daily_demand) >= 1:
            trend = "Falling"
        else:
            trend = "Stable"

        meta = product_meta.get(product_name, {})
        current_stock = meta.get("current_stock")
        reorder_point = meta.get("reorder_point")
        category = standardize_product_category(meta.get("category", "Seasonal / Miscellaneous Items"))

        forecast_entry = forecast_map.get(product_name, {}) if isinstance(forecast_map, dict) else {}
        forecast_demand_raw = _forecast_total_from_entry(forecast_entry, forecast_days)
        forecast_status = forecast_entry.get("status", "limited") if isinstance(forecast_entry, dict) else "limited"

        risk_entry = risk_map.get(product_name, {}) if isinstance(risk_map, dict) else {}
        probability = risk_entry.get("probability")
        factors = risk_entry.get("top_factors", [])

        calibrated = _calibrate_risk_outcome(
            probability=probability,
            current_stock=current_stock,
            reorder_point=reorder_point,
            forecast_demand=forecast_demand_raw,
            average_daily_demand=average_daily_demand,
            trend=trend,
            forecast_status=forecast_status,
            factors=factors,
            risk_note=risk_entry.get("note"),
        )
        risk_level = calibrated["risk_level"]
        suggested_action = calibrated["suggested_action"]
        priority = calibrated["priority"]
        risk_reason = calibrated["reason"]
        stock_cover_days = calibrated.get("stock_cover_days")

        stock_number = int(current_stock) if current_stock is not None and not pd.isna(current_stock) else None
        reorder_number = int(reorder_point) if reorder_point is not None and not pd.isna(reorder_point) else None
        recommended_additional_stock = (
            max(int(math.ceil(forecast_demand_raw or 0)) - (stock_number or 0), 0)
            if stock_number is not None and forecast_demand_raw is not None
            else None
        )
        forecast_display = _friendly_forecast_value(forecast_demand_raw, forecast_status)
        suggested_action_label = _friendly_suggested_action_label(suggested_action)
        stock_status_value = _product_stock_status_from_decision(suggested_action_label)
        assistant_stock_status_value = _assistant_stock_status_from_store_status(stock_status_value)
        display_note = _build_demand_detail_note(
            forecast_status=forecast_status,
            forecast_demand=forecast_demand_raw,
            current_stock=stock_number,
            reorder_point=reorder_number,
            risk_level=risk_level,
            trend=trend,
            fallback_reason=risk_reason,
            probability=probability,
        )
        stock_cover_label = None if stock_cover_days is None else f"About {stock_cover_days:.1f} days of stock at the recent sales pace."

        rows.append({
            "product_name": product_name,
            "category": category,
            "current_stock": stock_number if stock_number is not None else "No uploaded sales data yet",
            "reorder_point": reorder_number if reorder_number is not None else "No uploaded sales data yet",
            "forecast_demand": forecast_display,
            "trend": trend,
            "risk_level": risk_level,
            "risk_display": risk_level if str(risk_level).strip().lower() in {"high", "moderate", "low"} else "Needs review",
            "why_flagged": risk_reason,
            "suggested_action": suggested_action_label,
            "stock_status": stock_status_value,
            "assistant_stock_status": assistant_stock_status_value,
            "priority": priority,
            "priority_display": _friendly_priority_label(priority, risk_level),
            "recommended_additional_stock": recommended_additional_stock,
            "average_daily_demand": round(average_daily_demand, 2),
            "recent_total": round(recent_total, 2),
            "previous_total": round(previous_total, 2),
            "stockout_probability": probability,
            "stockout_probability_label": _friendly_probability_note(probability),
            "forecast_source": "SARIMA",
            "risk_source": "XGBoost",
            "forecast_status": forecast_status,
            "forecast_status_label": _friendly_model_status_label(forecast_status, "sarima"),
            "forecast_note": _friendly_forecast_note(forecast_status, forecast_entry.get("note") if isinstance(forecast_entry, dict) else None),
            "risk_note": risk_entry.get("note"),
            "display_note": display_note,
            "display_note_short": _short_display_note(display_note),
            "dashboard_main_reason": _dashboard_main_reason(display_note, risk_reason),
            "detail_factors": _detail_factor_list(factors, _friendly_probability_note(probability), stock_cover_label),
            "stock_cover_days": stock_cover_days,
            "stock_cover_label": stock_cover_label,
            "probability_band": "High confidence" if probability is not None and probability >= 0.78 else ("Watchlist" if probability is not None and probability >= 0.56 else ("Lower concern" if probability is not None else "Needs review")),
            "comparable_history": comparable_history,
            "growth_pct": round(float(growth_pct), 4) if growth_pct is not None else None,
            "is_low_stock": bool(stock_number is not None and reorder_number is not None and stock_number <= reorder_number),
        })

    priority_order = {"High": 0, "Moderate": 1, "Low": 2, "Unavailable": 3}
    rows.sort(key=lambda item: (priority_order.get(item["priority"], 9), priority_order.get(item["risk_level"], 9), -(item["average_daily_demand"])))

    insights_cache[cache_key] = [row.copy() for row in rows]
    _state_set("product_insights_cache", insights_cache)
    return [row.copy() for row in rows]


def get_report_rows(df: pd.DataFrame | None) -> list[dict[str, Any]]:
    if df is None or 'product_name' not in df.columns:
        return []
    working = df.copy()
    for col in ['quantity_sold', 'unit_price', 'current_stock', 'reorder_point']:
        if col not in working.columns:
            working[col] = 0
        working[col] = pd.to_numeric(working[col], errors='coerce').fillna(0)
    if 'category' not in working.columns:
        working['category'] = 'Seasonal / Miscellaneous Items'
    working['category'] = working['category'].fillna('Seasonal / Miscellaneous Items').apply(standardize_product_category)
    working['sales_value'] = working['quantity_sold'] * working['unit_price']
    grouped = working.groupby('product_name', dropna=True, as_index=False).agg({
        'quantity_sold': 'sum',
        'sales_value': 'sum',
        'current_stock': 'last',
        'reorder_point': 'last',
        'category': 'last',
    })
    insight_map = {row['product_name']: row for row in compute_product_insights(df)}
    rows = []
    for _, row in grouped.iterrows():
        insight = insight_map.get(row['product_name'], {})
        rows.append({
            'product_name': row['product_name'],
            'category': standardize_product_category(row.get('category')),
            'total_monthly_sales_qty': int(row['quantity_sold']),
            'total_sales_value': round(float(row['sales_value']), 2),
            'average_daily_demand': insight.get('average_daily_demand', round(float(row['quantity_sold']) / 30, 2)),
            'forecast_demand': insight.get('forecast_demand', 'Forecast not ready'),
            'current_stock': int(float(row['current_stock'])) if not pd.isna(row['current_stock']) else 'No uploaded sales data yet',
            'risk_level': insight.get('risk_level', 'Unavailable'),
            'suggested_action': insight.get('suggested_action', 'Gather more sales history'),
            'display_note': insight.get('display_note', insight.get('why_flagged', 'Needs review.')),
            'display_note_short': insight.get('display_note_short', _short_display_note(insight.get('display_note', insight.get('why_flagged', 'Needs review.')))),
        })
    risk_order = {'High': 0, 'Moderate': 1, 'Low': 2, 'Unavailable': 3}
    rows.sort(key=lambda item: (risk_order.get(item.get('risk_level', 'Unavailable'), 3), str(item.get('product_name', '')).casefold()))
    return rows
def get_dashboard_metrics(df: pd.DataFrame | None) -> dict[str, Any] | None:
    if df is None:
        return None
    insight_rows = compute_product_insights(df)
    report_rows = get_report_rows(df)
    total_products = len(insight_rows)
    high_risk_items = sum(1 for item in insight_rows if item['risk_level'] == 'High')
    moderate_risk_items = sum(1 for item in insight_rows if item['risk_level'] == 'Moderate')
    low_risk_items = sum(1 for item in insight_rows if item['risk_level'] == 'Low')
    total_sales_volume = round(sum(row['total_sales_value'] for row in report_rows),2)
    rising_demand_products = sum(1 for item in insight_rows if item['trend'] == 'Rising')
    most_sold_product = max(report_rows, key=lambda x: x['total_monthly_sales_qty'])['product_name'] if report_rows else 'No uploaded sales data yet'
    high_priority_rows=[row for row in insight_rows if row['risk_level']=='High']
    highest_risk_product = high_priority_rows[0]['product_name'] if high_priority_rows else 'No uploaded sales data yet'
    return {
        'total_products': total_products,
        'high_risk_items': high_risk_items,
        'moderate_risk_items': moderate_risk_items,
        'low_risk_items': low_risk_items,
        'rising_demand_products': rising_demand_products,
        'total_sales_volume': total_sales_volume,
        'most_sold_product': most_sold_product,
        'highest_risk_product': highest_risk_product,
    }


def get_buyer_behavior_insights(df: pd.DataFrame | None, product_insights: list[dict[str, Any]] | None = None) -> list[dict[str, str]]:
    if df is None or df.empty:
        return []

    working = df.copy()
    if 'date' in working.columns:
        working['date'] = pd.to_datetime(working['date'], errors='coerce')
    if 'quantity_sold' in working.columns:
        working['quantity_sold'] = pd.to_numeric(working['quantity_sold'], errors='coerce').fillna(0)

    product_insights = product_insights or []
    high_risk_count = sum(1 for item in product_insights if item.get('risk_level') == 'High')
    moderate_risk_count = sum(1 for item in product_insights if item.get('risk_level') == 'Moderate')
    rising_count = sum(1 for item in product_insights if item.get('trend') == 'Rising')

    insight_rows: list[dict[str, str]] = []

    payday_message = 'Demand may rise near payday dates.'
    if 'date' in working.columns and 'quantity_sold' in working.columns:
        dated_rows = working.dropna(subset=['date']).copy()
        if not dated_rows.empty:
            if 'is_payday_period' in dated_rows.columns:
                payday_mask = dated_rows['is_payday_period'].astype(str).str.lower().isin(['1', 'true', 'yes', 'y', 'payday'])
            else:
                payday_mask = dated_rows['date'].dt.day.isin([15, 30])
            payday_average = float(dated_rows.loc[payday_mask, 'quantity_sold'].mean()) if payday_mask.any() else None
            regular_average = float(dated_rows.loc[~payday_mask, 'quantity_sold'].mean()) if (~payday_mask).any() else None
            if payday_average is None or regular_average is None or payday_average <= regular_average * 1.05:
                payday_message = 'Check demand near the 15th and 30th before restocking.'
    insight_rows.append({
        'title': 'Payday Pattern',
        'message': payday_message,
    })

    if rising_count > 0:
        fast_message = f'{rising_count} product(s) show repeated or rising purchases.'
    else:
        fast_message = 'Repeated purchases may need earlier review.'
    insight_rows.append({
        'title': 'Fast-Moving Items',
        'message': fast_message,
    })

    if high_risk_count > 0:
        stockout_message = 'Predicted demand is higher than latest recorded stock.'
    elif moderate_risk_count > 0:
        stockout_message = 'Some products may need review before restocking.'
    else:
        stockout_message = 'Most products can be monitored based on current records.'
    insight_rows.append({
        'title': 'Stockout Driver',
        'message': stockout_message,
    })

    category_message = 'Some categories may need review before restocking.'
    if 'category' in working.columns and 'quantity_sold' in working.columns:
        category_rows = working.dropna(subset=['category']).copy()
        if not category_rows.empty:
            category_rows['category'] = category_rows['category'].fillna('Seasonal / Miscellaneous Items').apply(standardize_product_category)
            category_totals = category_rows.groupby('category')['quantity_sold'].sum().sort_values(ascending=False)
            if not category_totals.empty:
                top_category = category_totals.index[0]
                category_message = f'{top_category} may need review before restocking.'
    insight_rows.append({
        'title': 'Category Watch',
        'message': category_message,
    })

    return insight_rows[:4]

def get_dashboard_summary(df: pd.DataFrame | None) -> dict[str, Any]:
    cache_key = _analytics_cache_key(df, forecast_days=7) + ("dashboard_summary",)
    cached_key = _state_get("dashboard_summary_cache_key")
    cached_summary = _state_get("dashboard_summary_cache")
    if cached_key == cache_key and isinstance(cached_summary, dict):
        summary = copy.deepcopy(cached_summary)
        summary.setdefault('dashboard_priority_rows', (summary.get('priority_rows') or [])[:5])
        return summary

    metrics = get_dashboard_metrics(df)
    insights = compute_product_insights(df)
    chart_data = get_dashboard_chart_data(df)
    buyer_behavior_insights = get_buyer_behavior_insights(df, insights)
    summary = {
        'metrics': metrics,
        'priority_rows': insights,
        'dashboard_priority_rows': insights[:5],
        'buyer_behavior_insights': buyer_behavior_insights,
        'chart_type': chart_data.get('chart_type', 'line'),
        'chart_labels': chart_data.get('chart_labels', []),
        'chart_datasets': chart_data.get('chart_datasets', []),
        'chart_message': chart_data.get('chart_message'),
        'chart_options': chart_data.get('chart_options', {}),
        'chart_explanation': chart_data.get('chart_explanation') or ('No chart data available yet.' if not chart_data.get('chart_labels') else 'Overall demand based on uploaded sales records.'),
        'chart_focus_label': chart_data.get('chart_focus_label', 'Overall Demand'),
        'has_forecast': chart_data.get('has_forecast', False),
        'forecast_start_label': chart_data.get('forecast_start_label'),
        'model_ui_summary': get_model_ui_summary(),
    }
    _state_set("dashboard_summary_cache_key", cache_key)
    _state_set("dashboard_summary_cache", copy.deepcopy(summary))
    return copy.deepcopy(summary)




def get_assistant_stock_status(row: dict[str, Any]) -> str:
    """Return a simple shelf-check label for the Operational Assistant view.

    The assistant label must come from the same stock-status result used by the
    Owner and Store Manager views. Avoid recalculating a separate status from
    risk/trend here because that makes role dashboards show conflicting counts.
    """
    explicit_status = str(row.get("assistant_stock_status") or "").strip()
    if explicit_status in {"Low", "Monitor", "Stable"}:
        return explicit_status

    raw_status = str(row.get("stock_status") or "").strip()
    if raw_status:
        return _assistant_stock_status_from_store_status(raw_status)

    if bool(row.get("is_low_stock")):
        return "Low"
    risk_level = str(row.get("risk_level") or "").strip().lower()
    priority = str(row.get("priority") or "").strip().lower()
    if risk_level in {"high", "moderate"} or priority in {"high", "moderate"}:
        return "Monitor"
    return "Stable"


def get_assistant_stock_concern(row: dict[str, Any]) -> str:
    """Return a non-technical stock concern message for frontline checking."""
    assistant_status = get_assistant_stock_status(row)
    if assistant_status == "Low":
        return "Stock is low. Check the shelf and storage area, then report if more supply is needed."
    if assistant_status == "Monitor":
        return "This item may need attention soon. Keep checking the shelf and update the Store Manager if it moves fast."
    return "Stock looks stable. Continue normal shelf checking during the day."


def get_assistant_suggested_action(row: dict[str, Any]) -> str:
    """Return a simple action label for the Operational Assistant checklist."""
    assistant_status = get_assistant_stock_status(row)
    if assistant_status == "Low":
        return "Check and report"
    if assistant_status == "Monitor":
        return "Monitor closely"
    return "Continue checking"

def get_role_dashboard_context(
    role: str,
    dashboard_summary: dict[str, Any],
    upload_status: dict[str, Any] | None,
) -> dict[str, Any]:
    """Build lightweight role-specific dashboard content without changing access rules."""
    normalized_role = normalize_role(role)
    metrics = dashboard_summary.get("metrics") or {}
    priority_rows = dashboard_summary.get("priority_rows") or []
    upload_status = upload_status or {}
    model_ui_summary = dashboard_summary.get("model_ui_summary") or get_model_ui_summary()

    all_high_risk_rows = [
        row for row in priority_rows
        if str(row.get("risk_level") or "").strip().lower() == "high"
    ]
    high_risk_rows = all_high_risk_rows[:6]

    all_low_stock_rows = [
        row for row in priority_rows
        if str(row.get("stock_status") or "").strip().casefold() == "low stock"
    ]
    low_stock_rows = all_low_stock_rows[:6]

    all_monitor_rows = [
        row for row in priority_rows
        if str(row.get("stock_status") or "").strip().casefold() == "needs review"
    ]
    all_stable_rows = [
        row for row in priority_rows
        if str(row.get("stock_status") or "").strip().casefold() == "sufficient"
    ]
    all_review_rows = all_low_stock_rows + all_monitor_rows
    monitored_count = len(all_monitor_rows)
    review_rows = all_review_rows[:8]
    if not review_rows:
        review_rows = priority_rows[:8]

    assistant_source_rows = [
        row for row in priority_rows
        if get_assistant_stock_status(row) in {"Low", "Monitor"}
    ][:8] or review_rows or priority_rows[:8]
    assistant_task_rows = []
    for row in assistant_source_rows[:8]:
        assistant_task_rows.append({
            "product_name": row.get("product_name") or "Unnamed product",
            "category": row.get("category") or "Uncategorized",
            "current_stock": row.get("current_stock", "No uploaded sales data yet"),
            "stock_status": get_assistant_stock_status(row),
            "stock_concern": get_assistant_stock_concern(row),
            "action_label": get_assistant_suggested_action(row),
            "action_url": url_for("product_list"),
        })

    has_data = bool(priority_rows or dashboard_summary.get("chart_labels") or metrics.get("total_products"))
    last_processed_label = upload_status.get("last_processed_label") or get_last_processed_label()
    latest_upload_status_label = upload_status.get("freshness_label") or "No uploaded sales data yet"
    latest_results_label = model_ui_summary.get("model_run_timestamp") or "No generated results yet"

    base = {
        "role": normalized_role,
        "title": "Dashboard",
        "subtitle": "",
        "has_data": has_data,
        "last_processed_label": last_processed_label,
        "latest_upload_status_label": latest_upload_status_label,
        "latest_results_label": latest_results_label,
        "review_rows": review_rows,
        "high_risk_rows": high_risk_rows,
        "low_stock_rows": low_stock_rows,
        "review_count": len(all_review_rows) if all_review_rows else len(priority_rows),
        "high_risk_count": len(all_high_risk_rows),
        "low_stock_count": len(all_low_stock_rows),
        "stable_count": len(all_stable_rows),
        "monitored_count": monitored_count,
        "assistant_task_rows": assistant_task_rows,
        "show_owner_dashboard": normalized_role == "Owner",
        "show_manager_dashboard": normalized_role == "Store Manager",
        "show_assistant_dashboard": normalized_role == "Operational Assistant",
        "show_full_owner_dashboard": normalized_role == "Owner",
        "show_operational_chart": normalized_role in {"Owner", "Store Manager"},
        "show_buyer_behavior": normalized_role == "Owner",
        "show_priority_table": normalized_role in {"Owner", "Store Manager"},
        "show_simple_tasks": normalized_role == "Operational Assistant",
        "empty_prompt": "Upload sales data to start generating store insights.",
        "primary_action_label": "",
        "primary_action_url": "",
        "secondary_actions": [],
        "secondary_action_label": "",
        "secondary_action_url": "",
        "manager_focus_items": [],
        "assistant_reminders": [],
    }

    if normalized_role == "Owner":
        base.update({
            "primary_action_label": "Upload New Data",
            "primary_action_url": url_for("upload_data"),
        })
    elif normalized_role == "Store Manager":
        base.update({
            "subtitle": "",
            "primary_action_label": "Upload Sales Data",
            "primary_action_url": url_for("upload_data"),
            "secondary_actions": [],
            "secondary_action_label": "",
            "secondary_action_url": "",
            "empty_prompt": "Upload store sales records to view demand outlook, stockout risk, and products needing review.",
            "manager_focus_items": [
                {"title": "Review high-risk products first.", "message": f"{len(all_high_risk_rows)} high-risk product(s) need attention from the latest results."},
                {"title": "Check products with low recorded stock.", "message": f"{len(all_low_stock_rows)} product(s) are at or below their reorder point."},
                {"title": "Upload latest sales file if new records are available.", "message": latest_upload_status_label},
                {"title": "Generate results after uploading updated data.", "message": latest_results_label},
            ],
        })
    else:
        base.update({
            "subtitle": "",
            "primary_action_label": "",
            "primary_action_url": "",
            "show_operational_chart": False,
            "show_buyer_behavior": False,
            "show_priority_table": False,
            "show_simple_tasks": True,
            "empty_prompt": "No product data available yet.",
            "assistant_reminders": [
                "Check low-stock products first.",
                "Report unavailable items to the Store Manager.",
                "Review products marked Monitor.",
            ] if has_data else [],
        })

    return base


def _extract_number_from_display(value: Any) -> float | None:
    """Return a float from a display value such as 12, '12', or '12 units'."""
    if value is None:
        return None
    if isinstance(value, (int, float, np.integer, np.floating)):
        try:
            if pd.isna(value):
                return None
            return float(value)
        except Exception:
            return None
    match = re.search(r"-?\d+(?:\.\d+)?", str(value))
    if not match:
        return None
    try:
        return float(match.group(0))
    except Exception:
        return None


def _sum_forecast_horizon(forecast_map: dict[str, Any], product_names: list[str], forecast_days: int) -> float | None:
    total = 0.0
    found = False
    for product_name in product_names:
        entry = forecast_map.get(product_name, {}) if isinstance(forecast_map, dict) else {}
        numeric = _forecast_total_from_entry(entry, forecast_days)
        if numeric is not None:
            total += numeric
            found = True
    return round(total, 2) if found else None


def _build_forecast_scope_summary(rows: list[dict[str, Any]], scope_label: str, forecast_days: int, forecast_total: float | None = None) -> dict[str, Any]:
    if not rows:
        return {
            'forecast_demand': 'Forecast not ready',
            'trend': 'No data yet',
            'suggested_action': 'Upload sales data',
            'recommended_additional_stock': 0,
            'why_flagged': 'No uploaded sales data yet.',
            'display_note': 'No uploaded sales data yet.',
            'forecast_note': 'No chart data available yet.',
            'category': scope_label,
            'forecast_status_label': 'Need more sales history',
        }

    high_count = sum(1 for row in rows if row.get('risk_level') == 'High')
    moderate_count = sum(1 for row in rows if row.get('risk_level') == 'Moderate')
    rising_count = sum(1 for row in rows if row.get('trend') == 'Rising')
    decision_rows = [row for row in rows if row.get('suggested_action') in {'Consider restocking soon', 'Review before next restocking'}]

    if high_count:
        suggested = 'Review high-risk products first'
        note = f'{high_count} product(s) may need attention because demand and stock patterns show higher stockout risk.'
    elif moderate_count:
        suggested = 'Review before next restocking'
        note = f'{moderate_count} product(s) should be monitored before the next restocking schedule.'
    elif decision_rows:
        suggested = 'Review before next restocking'
        note = 'Some products may need review based on predicted demand and latest recorded stock.'
    else:
        suggested = 'Monitor only'
        note = 'Stock appears sufficient for the current processed records.'

    if forecast_total is None:
        forecast_values = [_extract_number_from_display(row.get('forecast_demand')) for row in rows]
        valid_values = [value for value in forecast_values if value is not None]
        forecast_total = round(sum(valid_values), 2) if valid_values else None

    if rising_count >= max(1, math.ceil(len(rows) * 0.25)):
        trend = 'Rising'
    elif rising_count:
        trend = 'Mixed'
    else:
        trend = 'Stable'

    return {
        'forecast_demand': int(math.ceil(float(forecast_total))) if forecast_total is not None else 'Forecast not ready',
        'trend': trend,
        'suggested_action': suggested,
        'recommended_additional_stock': 0,
        'why_flagged': note,
        'display_note': note,
        'forecast_note': f'Based on uploaded sales records for the selected {forecast_days}-day view.',
        'category': scope_label,
        'forecast_status_label': 'Based on uploaded sales records',
    }


def get_insights_aggregate_chart_data(df: pd.DataFrame | None, forecast_days: int = 7, history_days: int = 14, category: str | None = None, view_key: str = "daily") -> dict[str, Any]:
    category = standardize_product_category(category) if category else None
    focus_label = f"{category} Demand" if category else "All Products"
    empty = {
        "chart_type": "line",
        "chart_labels": [],
        "chart_datasets": [],
        "chart_message": "No chart data available yet.",
        "chart_explanation": None,
        "chart_focus_label": focus_label,
        "has_forecast": False,
        "forecast_start_label": None,
        "chart_options": _forecast_chart_options(view_key),
    }
    if df is None or df.empty:
        return empty

    context = _get_daily_sales_context(df)
    if category:
        daily_totals = context.get("category_daily", {}).get(category, pd.Series(dtype=float))
        selected_products = context.get("category_products", {}).get(category, [])
    else:
        daily_totals = context.get("total_daily", pd.Series(dtype=float))
        selected_products = context.get("products", [])

    if daily_totals is None or daily_totals.empty:
        return empty

    daily_totals = daily_totals[daily_totals > 0].sort_index()
    if daily_totals.empty:
        return empty

    view_config = _forecast_view_config(view_key)
    group_by = view_config.get("group_by", "daily")
    historical_daily = daily_totals.tail(max(history_days, 7))
    historical = _aggregate_series_for_forecast_view(historical_daily, view_key)
    historical = _limit_forecast_view_points(historical, view_key, "history")
    labels = [_format_forecast_axis_label(d, group_by) for d in historical.index]
    historical_data = [round(float(v), 2) for v in historical.tolist()]
    chart = {
        "chart_type": "line",
        "chart_labels": labels,
        "chart_datasets": [{
            "label": "Observed Sales",
            "data": historical_data,
            "borderColor": "#f4d35e",
            "backgroundColor": "rgba(244, 211, 94, 0.18)",
            "tension": 0.32,
            "fill": True,
        }],
        "chart_message": None,
        "chart_explanation": None,
        "chart_focus_label": focus_label,
        "has_forecast": False,
        "forecast_start_label": None,
        "chart_options": _forecast_chart_options(view_key),
    }

    artifacts = get_model_artifacts()
    forecast_map = artifacts.get("forecast_by_product", {}) if isinstance(artifacts, dict) else {}
    future_totals: dict[pd.Timestamp, float] = {}
    for product_name in selected_products:
        forecast_entry = forecast_map.get(product_name, {}) if isinstance(forecast_map, dict) else {}
        if not isinstance(forecast_entry, dict):
            continue
        if str(forecast_entry.get("status", "") or "").strip().lower() != "completed":
            continue
        for item in (forecast_entry.get("daily") or [])[:forecast_days]:
            item_date = pd.to_datetime(item.get("date"), errors="coerce") if isinstance(item, dict) else pd.NaT
            item_qty = pd.to_numeric(item.get("quantity"), errors="coerce") if isinstance(item, dict) else None
            if pd.notna(item_date) and item_qty is not None and pd.notna(item_qty):
                normalized_date = item_date.normalize()
                future_totals[normalized_date] = future_totals.get(normalized_date, 0.0) + float(item_qty)

    if future_totals:
        ordered_dates = sorted(future_totals.keys())[:forecast_days]
        future_series = pd.Series(
            [future_totals[d] for d in ordered_dates],
            index=ordered_dates,
            dtype=float,
        )
        future_grouped = _aggregate_series_for_forecast_view(future_series, view_key)
        future_grouped = _limit_forecast_view_points(future_grouped, view_key, "forecast")
        future_dates = list(future_grouped.index)
        future_labels = [_format_forecast_axis_label(d, group_by) for d in future_dates]
        forecast_values = [round(float(value), 2) for value in future_grouped.tolist()]
        observed_data, forecast_data = _build_forecast_line_datasets(historical_data, forecast_values, len(future_labels))
        chart["chart_labels"] = labels + future_labels
        chart["chart_datasets"][0]["data"] = observed_data
        chart["chart_datasets"].append({
            "label": "Predicted Demand",
            "data": forecast_data,
            "borderColor": "#8d6e63",
            "backgroundColor": "rgba(141, 110, 99, 0.08)",
            "borderDash": [6, 4],
            "tension": 0.32,
            "fill": False,
        })
        chart["has_forecast"] = True
        if future_dates:
            chart["forecast_start_label"] = _format_forecast_axis_label(future_dates[0], group_by)
    else:
        chart["chart_message"] = None

    return chart

def _build_filter_metrics(rows: list[dict[str, Any]]) -> dict[str, Any]:
    total = len(rows)
    high_count = sum(1 for row in rows if row.get('risk_level') == 'High')
    rising_count = sum(1 for row in rows if row.get('trend') == 'Rising')
    low_stock_count = sum(1 for row in rows if row.get('is_low_stock'))
    urgent = rows[0].get('product_name') if rows else 'No records available yet'
    return {
        'total_products': total,
        'high_risk_products': high_count,
        'rising_demand_products': rising_count,
        'low_stock_products': low_stock_count,
        'most_urgent_product': urgent,
    }


def get_forecast_summary(df: pd.DataFrame | None) -> dict[str, Any]:
    base = _empty_forecast_summary()
    if df is None or "product_name" not in df.columns:
        return base

    artifacts = get_model_artifacts()
    cache_key = _analytics_cache_key(df, forecast_days=None, artifacts=artifacts) + ("forecast_summary",)
    cached_key = _state_get("forecast_summary_cache_key")
    cached_summary = _state_get("forecast_summary_cache")
    if cached_key == cache_key and isinstance(cached_summary, dict):
        return cached_summary

    try:
        context = _get_daily_sales_context(df)
        all_rows_daily = compute_product_insights(df, forecast_days=7)
        if not all_rows_daily:
            return base

        products = [row["product_name"] for row in all_rows_daily if row.get("product_name")]
        categories = get_available_categories(df)
        categories_with_rows = sorted({row.get("category") for row in all_rows_daily if row.get("category")}, key=str.casefold)

        base.update({"products": products, "categories": categories})
        for product_name in products:
            base["chart_map"][product_name] = {}
            base["summary_map"][product_name] = {}
        for category in categories_with_rows:
            base["category_chart_map"][category] = {}
            base["category_summary_map"][category] = {}

        forecast_map = artifacts.get("forecast_by_product", {}) if isinstance(artifacts, dict) else {}

        for option in get_forecast_view_options():
            range_key = option["key"]
            days = option["days"]
            history_days = option.get("history_days", 14 if days <= 14 else 21)
            rows = compute_product_insights(df, forecast_days=days)
            base["details_by_range"][range_key] = []
            insight_map = {row["product_name"]: row for row in rows}
            all_product_names = [row.get("product_name") for row in rows if row.get("product_name")]
            total_forecast = _sum_forecast_horizon(forecast_map, all_product_names, days)

            base["summary_map"]["__total__"][range_key] = _build_forecast_scope_summary(rows, "All Products", days, total_forecast)
            base["chart_map"]["__total__"][range_key] = get_insights_aggregate_chart_data(df, forecast_days=days, history_days=history_days, view_key=range_key)

            for category in categories_with_rows:
                category_rows = [row for row in rows if row.get("category") == category]
                category_product_names = [row.get("product_name") for row in category_rows if row.get("product_name")]
                category_forecast = _sum_forecast_horizon(forecast_map, category_product_names, days)
                base["category_summary_map"][category][range_key] = _build_forecast_scope_summary(category_rows, category, days, category_forecast)
                base["category_chart_map"][category][range_key] = get_insights_aggregate_chart_data(df, forecast_days=days, history_days=history_days, category=category, view_key=range_key)

            for row in rows:
                base["details_by_range"][range_key].append({
                    "product_name": row["product_name"],
                    "category": row["category"],
                    "forecast_period": option["label"],
                    "forecast_demand": row["forecast_demand"],
                    "forecast_numeric": _extract_number_from_display(row.get("forecast_demand")),
                    "trend": row["trend"],
                    "risk_level": row.get("risk_level", "Unavailable"),
                    "risk_display": row.get("risk_display", row.get("risk_level", "Needs review")),
                    "current_stock": row.get("current_stock", "No uploaded sales data yet"),
                    "reorder_point": row.get("reorder_point", "No uploaded sales data yet"),
                    "is_low_stock": row.get("is_low_stock", False),
                    "priority": row.get("priority", "Low"),
                    "priority_display": row.get("priority_display", row.get("priority", "Low")),
                    "suggested_action": row["suggested_action"],
                    "notes": row.get("display_note", row.get("why_flagged", "Needs review.")),
                    "note_detail": row.get("display_note", row["why_flagged"]),
                    "model_meta": row.get("stockout_probability_label"),
                    "average_daily_demand": row.get("average_daily_demand", 0),
                    "growth_pct": row.get("growth_pct"),
                })

            for product_name in products:
                insight = insight_map.get(product_name, {})
                base["summary_map"][product_name][range_key] = {
                    "forecast_demand": insight.get("forecast_demand", "Forecast not ready"),
                    "trend": insight.get("trend", "Stable"),
                    "suggested_action": insight.get("suggested_action", "Gather more sales history"),
                    "recommended_additional_stock": insight.get("recommended_additional_stock", 0),
                    "why_flagged": insight.get("why_flagged", "No uploaded sales data yet"),
                    "display_note": insight.get("display_note", insight.get("why_flagged", "No uploaded sales data yet")),
                    "forecast_note": insight.get("forecast_note", "Forecast is not ready for this item yet."),
                    "category": insight.get("category", "Uncategorized"),
                    "forecast_source": "SARIMA",
                    "risk_source": "XGBoost",
                    "stockout_probability": insight.get("stockout_probability"),
                    "stockout_probability_label": insight.get("stockout_probability_label"),
                    "forecast_status_label": insight.get("forecast_status_label", "Need more sales history"),
                }
                base["chart_map"][product_name][range_key] = get_product_chart_data(df, product_name, forecast_days=days, history_days=history_days, view_key=range_key)

        initial_rows = base["details_by_range"].get("daily", [])
        base["priority_rows"] = initial_rows[:5]
        base["filter_metrics"] = _build_filter_metrics(initial_rows)
        base["initial_view"] = {
            "product": "__total__",
            "range_key": "daily",
            "summary": base["summary_map"]["__total__"]["daily"],
            "table_rows": initial_rows,
            "priority_rows": initial_rows[:5],
            "filter_metrics": base["filter_metrics"],
            "chart": base["chart_map"]["__total__"]["daily"],
        }

        _state_set("forecast_summary_cache_key", cache_key)
        _state_set("forecast_summary_cache", base)
        return base
    except Exception:
        return _empty_forecast_summary()

def get_stock_risk_summary(df: pd.DataFrame | None) -> dict[str, Any]:
    if df is None:
        return _empty_stock_risk_summary()

    artifacts = get_model_artifacts()
    cache_key = _analytics_cache_key(df, forecast_days=7, artifacts=artifacts) + ("stock_risk_summary",)
    cached_key = _state_get("stock_risk_summary_cache_key")
    cached_summary = _state_get("stock_risk_summary_cache")
    if cached_key == cache_key and isinstance(cached_summary, dict):
        return cached_summary

    try:
        risk_rows = compute_product_insights(df, forecast_days=7)
        last_processed = get_last_processed_label()
        enriched = []
        for row in risk_rows:
            item = row.copy()
            item["last_processed_label"] = last_processed
            enriched.append(item)

        top_high_risk = [row for row in enriched if row.get("risk_level") == "High"][:5]
        priority_reorder = [row for row in enriched if row.get("suggested_action") in {"Consider restocking soon", "Review before next restocking"}]
        fastest_rising = [
            row for row in enriched
            if row.get("trend") == "Rising" or (row.get("growth_pct") is not None and row.get("growth_pct", 0) > 0.05)
        ]
        fastest_rising.sort(key=lambda item: (item.get("growth_pct") or 0, item.get("average_daily_demand", 0)), reverse=True)

        main_risk_drivers = []
        if enriched:
            if any(row.get("risk_level") == "High" for row in enriched):
                main_risk_drivers.append("Predicted demand is higher than latest recorded stock for some products.")
            if any(row.get("is_low_stock") for row in enriched):
                main_risk_drivers.append("Some products are near or below the reorder point.")
            if any(row.get("average_daily_demand", 0) > 0 for row in enriched):
                main_risk_drivers.append("Repeated purchases were detected in recent records.")
            if any(("payday" in str(row.get("display_note", "")).lower()) for row in enriched):
                main_risk_drivers.append("Demand may increase during payday or weekend periods.")
            main_risk_drivers.append("Stockout risk is based on demand and stock patterns from uploaded sales records.")
        else:
            main_risk_drivers.append("Upload sales data to view the main stockout risk drivers.")

        seen = set()
        unique_drivers = []
        for driver in main_risk_drivers:
            if driver not in seen:
                seen.add(driver)
                unique_drivers.append(driver)

        has_processed_data = bool(enriched)
        summary = {
            "inventory_rows": enriched,
            "has_data": has_processed_data,
            "high_count": sum(1 for item in enriched if item.get("risk_level") == "High"),
            "needs_restock_count": len(priority_reorder),
            "moderate_count": sum(1 for item in enriched if item.get("risk_level") == "Moderate"),
            "safe_count": sum(1 for item in enriched if item.get("risk_level") == "Low"),
            "top_high_risk": top_high_risk,
            "fastest_rising": fastest_rising[:5],
            "priority_reorder": priority_reorder[:5],
            "main_risk_drivers": unique_drivers[:5],
            "top_high_risk_empty_message": "No high-risk items were detected in the latest processed sales data." if has_processed_data else "No uploaded sales data yet.",
            "fastest_rising_empty_message": "No strong recent demand increase was detected." if has_processed_data else "No uploaded sales data yet.",
            "priority_reorder_empty_message": "No immediate reorder action is needed right now." if has_processed_data else "No uploaded sales data yet.",
            "model_ui_summary": get_model_ui_summary(),
        }
        _state_set("stock_risk_summary_cache_key", cache_key)
        _state_set("stock_risk_summary_cache", summary)
        return summary
    except Exception:
        return _empty_stock_risk_summary()

def _product_attention_sort_key(item: dict[str, Any]) -> tuple[int, int, int, float, str]:
    risk_order = {'High': 0, 'Moderate': 1, 'Low': 2, 'Unavailable': 3, 'Needs review': 3}
    stock_status_order = {'Low Stock': 0, 'Needs Review': 1, 'Sufficient': 2}
    priority_order = {'High': 0, 'Moderate': 1, 'Low': 2}
    forecast_value = _extract_number_from_display(item.get('forecast_demand')) or 0
    return (
        risk_order.get(str(item.get('risk_level') or 'Unavailable'), 4),
        stock_status_order.get(str(item.get('stock_status') or 'Needs Review'), 3),
        priority_order.get(str(item.get('priority') or 'Low'), 3),
        -forecast_value,
        str(item.get('product_name') or '').lower(),
    )


def _product_stock_status_from_decision(decision: str) -> str:
    if decision == 'Consider restocking soon':
        return 'Low Stock'
    if decision in {'Review before next restocking', 'Monitor this week', 'Gather more sales history'}:
        return 'Needs Review'
    return 'Sufficient'


def _assistant_stock_status_from_store_status(stock_status: Any) -> str:
    """Map the shared store stock status into the simpler assistant wording.

    This keeps Owner, Store Manager, and Operational Assistant views based on the
    same computed product status while allowing the assistant UI to use simpler
    labels.
    """
    normalized = str(stock_status or "").strip().casefold()
    if normalized in {"low stock", "low", "out of stock"}:
        return "Low"
    if normalized in {"needs review", "monitor", "review"}:
        return "Monitor"
    if normalized in {"sufficient", "stable"}:
        return "Stable"
    return "Monitor"


def _build_product_buyer_pattern(insight: dict[str, Any]) -> str:
    trend = str(insight.get('trend') or 'Stable')
    average_daily = _extract_number_from_display(insight.get('average_daily_demand')) or 0
    recent_total = _extract_number_from_display(insight.get('recent_total')) or 0
    display_note = str(insight.get('display_note') or insight.get('why_flagged') or '').lower()

    if 'payday' in display_note:
        return 'Demand may increase during payday periods based on uploaded sales records.'
    if trend == 'Rising':
        return 'Recent demand is rising based on uploaded sales records.'
    if average_daily > 0 and recent_total >= 3:
        return 'This product appears as a fast-moving item in uploaded records.'
    if average_daily > 0:
        return 'This product is frequently bought in small quantities based on recent records.'
    return 'Buyer behavior pattern is not ready for this item yet.'


def _build_stock_preparation_guide(insight: dict[str, Any], inventory: dict[str, Any]) -> str:
    risk_level = str(inventory.get('risk_level') or insight.get('risk_level') or 'Unavailable')
    decision = str(inventory.get('status') or insight.get('suggested_action') or 'Review before next restocking')
    forecast_value = _extract_number_from_display(inventory.get('forecast_demand') or insight.get('forecast_demand'))
    stock_value = _extract_number_from_display(inventory.get('current_stock'))
    reorder_value = _extract_number_from_display(inventory.get('reorder_point'))

    if risk_level == 'High' or decision == 'Consider restocking soon':
        if forecast_value is not None and stock_value is not None and forecast_value > stock_value:
            return 'Consider preparing additional stock because predicted demand is higher than the latest recorded stock.'
        return 'Review this item first before the next restocking schedule.'
    if reorder_value is not None and stock_value is not None and stock_value <= reorder_value:
        return 'Review before next restocking because this item is near or below the reorder point.'
    if risk_level == 'Moderate' or decision == 'Review before next restocking':
        return 'Review this item before the next restocking schedule and monitor recent sales movement.'
    return 'Stock appears sufficient. Continue monitoring this item during regular sales review.'


def _recent_points_from_series(series: pd.Series | None, days: int = 7) -> list[dict[str, Any]]:
    if series is None or series.empty:
        return []
    recent = series.sort_index().tail(days)
    return [
        {"date": index.strftime("%b %d"), "quantity": round(float(value), 2)}
        for index, value in recent.items()
    ]


def _recent_sales_chart_from_series(series: pd.Series | None, days: int = 14) -> dict[str, Any]:
    if series is None or series.empty:
        return {'labels': [], 'values': [], 'message': 'No chart data available yet.'}
    recent = series.sort_index().tail(days)
    recent = recent[recent >= 0]
    if recent.empty:
        return {'labels': [], 'values': [], 'message': 'No chart data available yet.'}
    return {
        'labels': [date.strftime('%b %d') for date in recent.index],
        'values': [round(float(value), 2) for value in recent.tolist()],
        'message': 'Recent sales trend based on uploaded sales records.',
    }


def _build_product_recent_sales_chart(df: pd.DataFrame | None, product_name: str) -> dict[str, Any]:
    if df is None or 'product_name' not in df.columns:
        return {'labels': [], 'values': [], 'message': 'No chart data available yet.'}
    context = _get_daily_sales_context(df)
    return _recent_sales_chart_from_series(context.get('product_daily', {}).get(product_name), days=14)


def get_inventory_products(df: pd.DataFrame | None, product_insights: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    if df is None or 'product_name' not in df.columns:
        return []
    context = _get_daily_sales_context(df)
    product_meta = context.get('product_meta', {})
    insight_rows = product_insights if product_insights is not None else compute_product_insights(df, forecast_days=7)
    insight_map = {row['product_name']: row for row in insight_rows}

    product_names = sorted(set(context.get('products') or get_product_names(df)), key=str.casefold)
    records = []
    for product_name in product_names:
        meta = product_meta.get(product_name, {})
        insight = insight_map.get(product_name, {})
        status = insight.get('suggested_action', 'Needs more history')
        stock_status = insight.get('stock_status') or _product_stock_status_from_decision(status)
        current_stock = meta.get('current_stock')
        reorder_point = meta.get('reorder_point')
        unit_price = meta.get('unit_price') if 'unit_price' in meta else None
        records.append({
            'product_name': product_name,
            'category': standardize_product_category(meta.get('category')) if meta.get('category') else 'Seasonal / Miscellaneous Items',
            'current_stock': int(current_stock) if current_stock is not None and not pd.isna(current_stock) else 'No uploaded sales data yet',
            'reorder_point': int(reorder_point) if reorder_point is not None and not pd.isna(reorder_point) else 'No uploaded sales data yet',
            'unit_price': float(unit_price) if unit_price is not None and not pd.isna(unit_price) else 'No uploaded sales data yet',
            'unit_type': meta.get('unit_type') if meta.get('unit_type') else 'Unit',
            'risk_level': insight.get('risk_level', 'Unavailable'),
            'status': status,
            'stock_status': stock_status,
            'forecast_demand': insight.get('forecast_demand', 'Forecast not ready'),
            'trend': insight.get('trend', 'Stable'),
            'why_flagged': insight.get('why_flagged', 'No uploaded sales data yet'),
            'display_note': insight.get('display_note', insight.get('why_flagged', 'No uploaded sales data yet')),
            'forecast_note': insight.get('forecast_note', 'Forecast is not ready for this item yet.'),
            'priority': insight.get('priority', 'Low'),
            'last_processed_label': get_last_processed_label(),
            'stockout_probability': insight.get('stockout_probability'),
            'stockout_probability_label': insight.get('stockout_probability_label'),
            'forecast_source': 'SARIMA',
            'risk_source': 'XGBoost',
        })
    records.sort(key=_product_attention_sort_key)
    return records


def get_product_detail_map(
    df: pd.DataFrame | None,
    inventory_rows: list[dict[str, Any]] | None = None,
    product_insights: list[dict[str, Any]] | None = None,
) -> dict[str, dict[str, Any]]:
    if df is None or 'product_name' not in df.columns:
        return {}
    insight_rows = product_insights if product_insights is not None else compute_product_insights(df, forecast_days=7)
    inventory_list = inventory_rows if inventory_rows is not None else get_inventory_products(df, insight_rows)
    inventory_map = {row['product_name']: row for row in inventory_list}
    insight_map = {row['product_name']: row for row in insight_rows}
    product_daily = _get_daily_sales_context(df).get('product_daily', {})

    detail_map = {}
    for product_name in [row['product_name'] for row in inventory_list]:
        inventory = inventory_map.get(product_name, {})
        insight = insight_map.get(product_name, {})
        series = product_daily.get(product_name)
        main_reason = insight.get('display_note_short', insight.get('display_note', insight.get('why_flagged', 'No uploaded sales data yet')))
        detail_note = insight.get('display_note', insight.get('why_flagged', 'No uploaded sales data yet'))
        detail_map[product_name] = {
            'product_name': product_name,
            'category': inventory.get('category', 'Seasonal / Miscellaneous Items'),
            'unit_type': inventory.get('unit_type', 'Unit'),
            'current_stock': inventory.get('current_stock', 'No uploaded sales data yet'),
            'reorder_point': inventory.get('reorder_point', 'No uploaded sales data yet'),
            'risk_level': inventory.get('risk_level', 'Unavailable'),
            'stock_status': inventory.get('stock_status', 'Needs Review'),
            'trend': insight.get('trend', 'Stable'),
            'forecast_demand': insight.get('forecast_demand', 'Forecast not ready'),
            'why_flagged': main_reason,
            'main_reason': main_reason,
            'detail_note': detail_note,
            'model_meta': 'Predicted Demand · Stockout Risk',
            'suggested_action': insight.get('suggested_action', 'Gather more sales history'),
            'average_daily_demand': insight.get('average_daily_demand', 0),
            'recent_sales_total': insight.get('recent_total', 0),
            'last_processed_label': get_last_processed_label(),
            'model_run_timestamp': get_model_ui_summary().get('model_run_timestamp'),
            'forecast_horizon_label': 'Next 7 days',
            'recent_sales_points': _recent_points_from_series(series, days=7),
            'recent_sales_chart': _recent_sales_chart_from_series(series, days=14),
            'stockout_probability': insight.get('stockout_probability'),
            'stockout_probability_label': insight.get('stockout_probability_label'),
            'top_factors': insight.get('detail_factors', []),
            'buyer_behavior_pattern': _build_product_buyer_pattern(insight),
            'stock_preparation_guide': _build_stock_preparation_guide(insight, inventory),
        }
    return detail_map


def get_product_summary(df: pd.DataFrame | None) -> dict[str, Any]:
    empty_summary = {
        'inventory_rows': [],
        'categories': get_category_filter_options(),
        'stock_statuses': [],
        'risk_levels': [],
        'detail_map': {},
        'summary_counts': {'total_products': 0, 'low_stock': 0, 'needs_review': 0},
    }
    if df is None or 'product_name' not in df.columns:
        return empty_summary

    cache_key = _analytics_cache_key(df, forecast_days=7) + ('product_summary',)
    product_summary_cache = _state_get('product_summary_cache') or {}
    if cache_key in product_summary_cache:
        cached = product_summary_cache[cache_key]
        return {
            'inventory_rows': [row.copy() for row in cached.get('inventory_rows', [])],
            'categories': list(cached.get('categories', [])),
            'stock_statuses': list(cached.get('stock_statuses', [])),
            'risk_levels': list(cached.get('risk_levels', [])),
            'detail_map': {name: detail.copy() for name, detail in cached.get('detail_map', {}).items()},
            'summary_counts': cached.get('summary_counts', {}).copy(),
        }

    product_insights = compute_product_insights(df, forecast_days=7)
    inventory_rows = get_inventory_products(df, product_insights)
    categories = get_category_filter_options()
    stock_statuses = sorted({row['stock_status'] for row in inventory_rows if row.get('stock_status')})
    summary = {
        'inventory_rows': inventory_rows,
        'categories': categories,
        'stock_statuses': stock_statuses,
        'risk_levels': get_available_risk_levels(inventory_rows),
        'detail_map': get_product_detail_map(df, inventory_rows, product_insights),
        'summary_counts': {
            'total_products': len(inventory_rows),
            'low_stock': sum(1 for row in inventory_rows if row.get('stock_status') == 'Low Stock'),
            'needs_review': sum(1 for row in inventory_rows if row.get('stock_status') == 'Needs Review'),
            'stable_products': sum(1 for row in inventory_rows if row.get('stock_status') == 'Sufficient'),
        },
    }
    product_summary_cache[cache_key] = summary
    _state_set('product_summary_cache', product_summary_cache)
    return {
        'inventory_rows': [row.copy() for row in summary['inventory_rows']],
        'categories': list(summary['categories']),
        'stock_statuses': list(summary['stock_statuses']),
        'risk_levels': list(summary['risk_levels']),
        'detail_map': {name: detail.copy() for name, detail in summary['detail_map'].items()},
        'summary_counts': summary['summary_counts'].copy(),
    }


def get_report_summary(df: pd.DataFrame | None) -> dict[str, Any]:
    empty_summary = {
        'report_rows': [],
        'product_names': [],
        'risk_levels': [],
        'categories': get_category_filter_options(),
        'product_category_groups': {category: [] for category in get_category_filter_options()},
    }
    if df is None or 'product_name' not in df.columns:
        return empty_summary

    cache_key = _analytics_cache_key(df) + ('report_summary',)
    report_summary_cache = _state_get('report_summary_cache') or {}
    if cache_key in report_summary_cache:
        cached = report_summary_cache[cache_key]
        return {
            'report_rows': [row.copy() for row in cached.get('report_rows', [])],
            'product_names': list(cached.get('product_names', [])),
            'risk_levels': list(cached.get('risk_levels', [])),
            'categories': list(cached.get('categories', [])),
            'product_category_groups': {
                category: list(products)
                for category, products in cached.get('product_category_groups', {}).items()
            },
        }

    report_rows = get_report_rows(df)
    product_category_groups: dict[str, list[str]] = {category: [] for category in get_category_filter_options()}
    for row in report_rows:
        category = standardize_product_category(row.get('category'))
        product_name = row.get('product_name')
        if product_name and product_name not in product_category_groups.setdefault(category, []):
            product_category_groups[category].append(product_name)
    for category in product_category_groups:
        product_category_groups[category] = sorted(product_category_groups[category], key=str.casefold)

    summary = {
        'report_rows': report_rows,
        'product_names': sorted({row['product_name'] for row in report_rows}, key=str.casefold),
        'risk_levels': get_available_risk_levels(report_rows),
        'categories': get_category_filter_options(),
        'product_category_groups': product_category_groups,
    }
    report_summary_cache[cache_key] = summary
    _state_set('report_summary_cache', report_summary_cache)
    return {
        'report_rows': [row.copy() for row in summary['report_rows']],
        'product_names': list(summary['product_names']),
        'risk_levels': list(summary['risk_levels']),
        'categories': list(summary['categories']),
        'product_category_groups': {category: list(products) for category, products in summary['product_category_groups'].items()},
    }

def render_page(template_name: str, **context):
    """Render pages without forcing a database reload when a route already handled a fallback.

    Render deployments can return a plain 500 if a page tries to reload the latest
    processed dataset while a previous upload is only partially available or a
    dashboard query fails. Routes can pass _skip_dataset_reload=True to render an
    empty-safe page instead of repeating the failing database load.
    """
    heavy_templates = {"dashboard.html", "insights.html", "product_list.html", "reports.html"}
    needs_dataset = template_name in heavy_templates or template_name == "upload_data.html"
    skip_dataset_reload = bool(context.pop("_skip_dataset_reload", False))

    dataset = context.get("data")
    if dataset is None and needs_dataset:
        dataset = context.get("processed_data")
        if dataset is None and template_name in heavy_templates and not skip_dataset_reload:
            try:
                dataset = get_processed_dataset()
            except Exception as exc:
                print(f"[StockWise render_page dataset fallback] {type(exc).__name__}: {exc}", flush=True)
                dataset = None

    context.setdefault("data", dataset if needs_dataset else None)

    if "processed_filename" not in context:
        processed_filename = None
        if needs_dataset and not skip_dataset_reload:
            try:
                processed_filename = get_processed_filename()
            except Exception as exc:
                print(f"[StockWise render_page filename fallback] {type(exc).__name__}: {exc}", flush=True)
        context["processed_filename"] = processed_filename

    try:
        processed_at = get_app_state().processed_at
    except Exception:
        processed_at = None

    if "page_state" not in context:
        try:
            last_processed_label = format_datetime(processed_at) if skip_dataset_reload else get_last_processed_label()
        except Exception as exc:
            print(f"[StockWise render_page page-state fallback] {type(exc).__name__}: {exc}", flush=True)
            last_processed_label = "No data yet"
        context["page_state"] = {
            "coverage_period": infer_coverage_period(dataset) if dataset is not None else "No processed sales data yet",
            "last_processed_label": last_processed_label,
            "upload_freshness": infer_upload_freshness(processed_at),
        }

    if template_name in {"dashboard.html", "insights.html"}:
        if "model_ui_summary" not in context:
            if skip_dataset_reload:
                context["model_ui_summary"] = _empty_model_ui_summary()
            else:
                try:
                    context["model_ui_summary"] = get_model_ui_summary()
                except Exception as exc:
                    print(f"[StockWise render_page model summary fallback] {type(exc).__name__}: {exc}", flush=True)
                    context["model_ui_summary"] = _empty_model_ui_summary()
    else:
        context.setdefault("model_ui_summary", {})

    return render_template(template_name, **context)


def validate_strong_password(password: str) -> tuple[bool, str]:
    """Return a friendly validation result for signup passwords."""
    message = "Password must be at least 8 characters and include uppercase, lowercase, number, and special character."
    if len(password or "") < 8:
        return False, message
    if not re.search(r"[A-Z]", password):
        return False, message
    if not re.search(r"[a-z]", password):
        return False, message
    if not re.search(r"\d", password):
        return False, message
    if not re.search(r"[^A-Za-z0-9]", password):
        return False, message
    return True, ""


def warm_up_generated_page_contexts(processed_df: pd.DataFrame | None) -> dict[str, Any]:
    """Prepare reusable page summaries after Generate Results completes.

    This keeps the slower analytics work inside the user's explicit Generate Results
    action, so Dashboard, Insights, Products, and Reports can read cached display
    summaries during normal page viewing.
    """
    status = {
        "dashboard": False,
        "insights": False,
        "products": False,
        "reports": False,
    }
    if processed_df is None or processed_df.empty:
        return status

    def safe_prepare(key: str, callback) -> None:
        try:
            callback()
            status[key] = True
        except Exception:
            status[key] = False

    # Prepare shared artifacts and grouped sales series first. These helpers are
    # display-only and reuse results that were already generated by the model
    # pipeline above; they do not rerun SARIMA or XGBoost.
    safe_prepare("dashboard", lambda: (get_model_artifacts(), _get_daily_sales_context(processed_df), get_dashboard_summary(processed_df)))

    def prepare_insights() -> None:
        get_model_artifacts()
        _get_daily_sales_context(processed_df)
        for option in get_forecast_view_options():
            compute_product_insights(processed_df, forecast_days=option.get("days", 7))
        get_forecast_summary(processed_df)
        get_stock_risk_summary(processed_df)

    safe_prepare("insights", prepare_insights)
    safe_prepare("products", lambda: get_product_summary(processed_df))

    def prepare_reports() -> None:
        preferences = get_settings_preferences()
        default_report_type = preferences.get("default_report_type", "demand_forecast_summary")
        default_period = preferences.get("default_report_period", "last_30_days")
        get_report_summary(processed_df)
        get_report_context(processed_df, default_report_type, default_period)
        get_report_context(processed_df, default_report_type, "all")

    safe_prepare("reports", prepare_reports)
    _state_set("page_context_warmup_status", status)
    return status


# =========================================
# AUTH ROUTES
# =========================================
@app.route("/", methods=["GET"])
def root():
    if session.get("user_logged_in"):
        return redirect(url_for("dashboard"))
    return redirect(url_for("auth"))


@app.route("/auth", methods=["GET", "POST"])
def auth():
    if session.get("user_logged_in"):
        return redirect(url_for("dashboard"))

    auth_message = None
    auth_message_type = None
    auth_field_errors: dict[str, str] = {}
    active_form = request.args.get("form", "login")
    login_email = ""
    signup_name = ""
    signup_email = ""

    if request.method == "POST":
        action = request.form.get("action", "").strip().lower()

        try:
            if action == "signup":
                active_form = "signup"

                signup_name = request.form.get("name", "").strip()
                signup_email = request.form.get("email", "").strip()
                password = request.form.get("password", "").strip()
                confirm_password = request.form.get("confirm_password", "").strip()

                if not signup_name:
                    auth_field_errors["name"] = "Full name is required."
                if not signup_email:
                    auth_field_errors["email"] = "Email is required."
                elif not is_valid_email_format(signup_email):
                    auth_field_errors["email"] = "Please enter a valid email address."
                if not password:
                    auth_field_errors["password"] = "Password is required."
                if not confirm_password:
                    auth_field_errors["confirm_password"] = "Please confirm your password."

                password_ok = True
                password_error = ""
                if password:
                    password_ok, password_error = validate_strong_password(password)
                    if not password_ok:
                        auth_field_errors["password"] = password_error

                if password and confirm_password and password_ok and password != confirm_password:
                    auth_field_errors["confirm_password"] = "Passwords do not match."

                if auth_field_errors:
                    required_fields = {"name", "email", "password", "confirm_password"}
                    if required_fields.issubset(auth_field_errors.keys()):
                        auth_message = "Please complete all required fields before signing up."
                    elif {"password", "confirm_password"}.issubset(auth_field_errors.keys()) and not password and not confirm_password:
                        auth_message = "Please enter and confirm your password."
                    else:
                        auth_message = next(iter(auth_field_errors.values()))
                    auth_message_type = "error"
                elif get_user_by_email(signup_email):
                    auth_field_errors["email"] = "An account with this email already exists."
                    auth_message = auth_field_errors["email"]
                    auth_message_type = "error"
                else:
                    user_id = create_user(
                        full_name=signup_name,
                        email=signup_email,
                        password=password,
                        position="Owner",
                    )

                    user = get_user_by_id(user_id)
                    if user is None:
                        auth_message = "The account was created, but login could not continue."
                        auth_message_type = "error"
                    else:
                        session["user_logged_in"] = True
                        session["user_id"] = user["user_id"]
                        session["user_name"] = user["full_name"]
                        session["user_email"] = user["email"]
                        ensure_user_workspace(user["user_id"])
                        refresh_session_user(user["user_id"])
                        session["force_onboarding"] = True
                        create_initial_user_settings(user["user_id"])
                        add_activity_log("Account created", "Authentication", "Success", user_id=user["user_id"], store_id=session.get("user_store_id"))

                        return redirect(url_for("first_time_setup"))

            elif action == "login":
                active_form = "login"

                login_email = request.form.get("email", "").strip()
                password = request.form.get("password", "").strip()

                if not login_email:
                    auth_field_errors["login_email"] = "Email is required."
                elif not is_valid_email_format(login_email):
                    auth_field_errors["login_email"] = "Please enter a valid email address."
                if not password:
                    auth_field_errors["login_password"] = "Password is required."

                if auth_field_errors:
                    if {"login_email", "login_password"}.issubset(auth_field_errors.keys()):
                        auth_message = "Please enter your email and password."
                    else:
                        auth_message = next(iter(auth_field_errors.values()))
                    auth_message_type = "error"
                else:
                    user = get_user_by_email(login_email)

                    if user is None or not check_password_hash(user["password_hash"], password):
                        auth_field_errors["login_email"] = "Email or password is incorrect."
                        auth_field_errors["login_password"] = "Email or password is incorrect."
                        auth_message = "Email or password is incorrect."
                        auth_message_type = "error"
                    else:
                        preferred_store_id = user.get("store_id")
                        membership = get_active_membership_for_user(user.get("user_id"), preferred_store_id)
                        if not membership and normalize_role(user.get("role") or user.get("position")) == "Owner" and str(user.get("account_status", "active")).lower() == "active" and user.get("is_active", 1):
                            ensured_store_id = ensure_user_workspace(user["user_id"])
                            membership = get_active_membership_for_user(user.get("user_id"), ensured_store_id)
                        if not membership:
                            auth_message = "This account is not active. Please contact the store Owner."
                            auth_message_type = "error"
                        else:
                            session["user_logged_in"] = True
                            session["user_id"] = user["user_id"]
                            session["user_name"] = user["full_name"]
                            session["user_email"] = user["email"]
                            session["user_store_id"] = membership.get("store_id")
                            session["user_membership_id"] = membership.get("membership_id")
                            refresh_session_user(user["user_id"])
                            update_last_login(user["user_id"])
                            add_activity_log("Login", "Authentication", "Success", user_id=user["user_id"], store_id=session.get("user_store_id"))

                            if user_needs_onboarding(user["user_id"]):
                                return redirect(url_for("first_time_setup"))
                            return redirect(url_for("preparing_dashboard"))
            else:
                active_form = "login"
                auth_message = "The authentication request could not be processed. Please try again."
                auth_message_type = "error"

        except Exception as exc:
            print("AUTH ERROR:", exc)
            auth_message = f"Error: {str(exc)}"
            auth_message_type = "error"

    return render_template(
        "auth.html",
        auth_message=auth_message,
        auth_message_type=auth_message_type,
        active_form=active_form,
        login_email=login_email,
        signup_name=signup_name,
        signup_email=signup_email,
        auth_field_errors=auth_field_errors,
    )


@app.route("/logout")
def logout():
    if session.get("user_logged_in"):
        add_activity_log("Logout", "Authentication", "Success")
    session.clear()
    return redirect(url_for("auth", form="login"))



@app.route("/preparing_dashboard")
@login_required
def preparing_dashboard():
    user_id = get_current_user_id()
    if user_needs_onboarding(user_id):
        return redirect(url_for("first_time_setup"))
    return render_template("preparing_dashboard.html")


# =========================================
# MAIN APP ROUTES
# =========================================
@app.route("/first_time_setup", methods=["GET", "POST"])
@login_required
def first_time_setup():
    user_id = get_current_user_id()
    if user_id is None:
        return redirect(url_for("auth", form="login"))

    setup_message = None
    setup_message_type = None
    current_user_row = get_user_by_id(user_id) or {}
    raw_preferences = get_user_settings_from_db(user_id)
    saved_preferences = get_settings_preferences()

    setup_incomplete = bool(session.get("force_onboarding")) or int(raw_preferences.get("onboarding_completed", 1) or 0) == 0
    store_name_value = raw_preferences.get("store_name") or ""
    store_type_value = raw_preferences.get("store_type") or ""
    location_value = raw_preferences.get("location_area") or ""
    upload_mode_value = raw_preferences.get("default_upload_mode") or ""
    report_period_value = raw_preferences.get("default_report_period") or ""
    currency_value = raw_preferences.get("currency") or DEFAULT_SETTINGS["currency"]
    if setup_incomplete:
        if store_name_value == DEFAULT_SETTINGS["store_name"]:
            store_name_value = ""
        if store_type_value == DEFAULT_SETTINGS["store_type"]:
            store_type_value = ""
        if location_value == DEFAULT_SETTINGS["location_area"]:
            location_value = ""
        if upload_mode_value == DEFAULT_SETTINGS["default_upload_mode"]:
            upload_mode_value = ""
        if report_period_value == DEFAULT_SETTINGS["default_report_period"]:
            report_period_value = ""

    form_values = {
        "display_name": current_user_row.get("full_name") or session.get("user_name", ""),
        "user_role": current_user_row.get("position") or session.get("user_position", "Owner") or "Owner",
        "profile_image": current_user_row.get("profile_image") or session.get("user_profile_image", ""),
        "store_name": store_name_value,
        "store_type": store_type_value,
        "location_area": location_value,
        "store_logo": saved_preferences.get("store_logo", ""),
        "has_store_logo": "yes" if saved_preferences.get("store_logo") else "",
        "currency": currency_value,
        "default_report_period": report_period_value,
        "default_upload_mode": upload_mode_value,
    }

    if form_values["user_role"] not in USER_ROLE_OPTIONS:
        form_values["user_role"] = "Owner"

    requested_step = request.args.get("step", "personal").strip().lower()

    if request.method == "POST":
        action = request.form.get("action", "").strip().lower()

        if action == "save_personal_setup":
            saved, message, values = save_personal_setup_from_form(user_id, request.form, request.files)
            form_values.update(values)
            if saved:
                return redirect(url_for("first_time_setup", step="store"))
            setup_message = message
            setup_message_type = "error"
            requested_step = "personal"

        elif action == "save_store_setup":
            saved, message, values = save_store_setup_from_form(user_id, request.form, request.files)
            form_values.update(values)
            if saved:
                session["show_first_upload_prompt"] = True
                return redirect(url_for("first_time_setup", step="upload"))
            setup_message = message
            setup_message_type = "error"
            requested_step = "store"

        elif action == "save_logo_setup":
            # Older Phase 6 forms used a separate logo step. Keep this as a safe fallback.
            return redirect(url_for("first_time_setup", step="store"))

        elif action == "upload_now":
            session.pop("show_first_upload_prompt", None)
            return redirect(url_for("upload_data"))

        elif action == "upload_later":
            session.pop("show_first_upload_prompt", None)
            return redirect(url_for("dashboard"))

        else:
            setup_message = "The setup request could not be processed. Please try again."
            setup_message_type = "error"
            requested_step = "personal"

    show_upload_prompt = bool(session.get("show_first_upload_prompt"))
    if requested_step == "upload" and show_upload_prompt:
        setup_step = "upload_prompt"
    else:
        if not user_needs_onboarding(user_id) and not show_upload_prompt:
            return redirect(url_for("dashboard"))
        setup_step = requested_step if requested_step in {"personal", "store"} else "personal"

    return render_template(
        "first_time_setup.html",
        setup_step=setup_step,
        setup_message=setup_message,
        setup_message_type=setup_message_type,
        form_values=form_values,
        store_type_options=get_store_type_options(),
        user_role_options=get_user_role_options(),
        employee_role_options=EMPLOYEE_ROLE_OPTIONS,
        report_period_options=get_report_period_options(),
        upload_mode_options=get_upload_mode_options(),
        currency_options=get_currency_options(),
    )


@app.route("/download_sample_template")
@login_required
@role_required("upload_data")
def download_sample_template():
    # Header-only template: no fake sales rows are included.
    template_df = pd.DataFrame(columns=STANDARD_TEMPLATE_COLUMNS)

    buffer = BytesIO(template_df.to_csv(index=False).encode("utf-8"))
    buffer.seek(0)
    return send_file(
        buffer,
        mimetype="text/csv",
        as_attachment=True,
        download_name="stockwise_standard_sales_template.csv",
    )


@app.route("/download_standard_template")
@login_required
@role_required("upload_data")
def download_standard_template():
    return download_sample_template()


def _empty_upload_status() -> dict[str, Any]:
    """Fallback upload status used when the database-backed status cannot be loaded."""
    return {
        "selected_filename": None,
        "processed_filename": None,
        "selected_at": None,
        "processed_at": None,
        "selected_coverage_period": "No selected file yet",
        "processed_coverage_period": "No processed sales data yet",
        "coverage_period": "No processed sales data yet",
        "freshness_label": "No uploaded sales data yet",
        "last_processed_label": "No data yet",
        "last_upload_mode": "new",
        "active_filename": None,
        "missing_date_gaps": [],
        "coverage_overlap": None,
        "recommended_upload_frequency": "Daily uploads are recommended for more reliable forecast and stockout insights.",
    }


def _state_only_upload_status() -> dict[str, Any]:
    """Return upload status without touching MySQL.

    The Render logs showed /dashboard timing out while mysql-connector was
    resetting a pooled connection during conn.close(). Dashboard should not
    perform a database hydration just to show the page, so this helper reads only
    the in-memory state for a safe fallback view.
    """
    status = _empty_upload_status()
    try:
        state = get_app_state()
        selected_df = state.selected_data
        processed_df = state.processed_data
        active_df = processed_df if processed_df is not None else selected_df
        active_name = state.processed_filename if processed_df is not None else state.selected_filename
        processed_at = state.processed_at
        status.update({
            "selected_filename": state.selected_filename,
            "processed_filename": state.processed_filename,
            "selected_at": state.selected_at,
            "processed_at": processed_at,
            "selected_coverage_period": infer_coverage_period(selected_df),
            "processed_coverage_period": infer_coverage_period(processed_df),
            "coverage_period": infer_coverage_period(active_df),
            "freshness_label": infer_upload_freshness(processed_at),
            "last_processed_label": format_datetime(processed_at),
            "last_upload_mode": state.last_upload_mode,
            "active_filename": active_name,
            "missing_date_gaps": find_missing_date_gaps(active_df),
            "coverage_overlap": coverage_overlap(selected_df, processed_df),
        })
    except Exception as exc:
        print(f"[StockWise state-only upload status fallback] {type(exc).__name__}: {exc}", flush=True)
    return status


def _dashboard_recovery_summary() -> dict[str, Any]:
    """Empty dashboard summary that does not query MySQL or model artifacts."""
    chart_data = get_dashboard_chart_data(None)
    return {
        "metrics": None,
        "priority_rows": [],
        "dashboard_priority_rows": [],
        "buyer_behavior_insights": [],
        "chart_type": chart_data.get("chart_type", "line"),
        "chart_labels": chart_data.get("chart_labels", []),
        "chart_datasets": chart_data.get("chart_datasets", []),
        "chart_message": chart_data.get("chart_message"),
        "chart_options": chart_data.get("chart_options", {}),
        "chart_explanation": chart_data.get("chart_explanation") or "No chart data available yet.",
        "chart_focus_label": chart_data.get("chart_focus_label", "Overall Demand"),
        "has_forecast": False,
        "forecast_start_label": None,
        "model_ui_summary": _empty_model_ui_summary(),
    }


@app.route("/dashboard")
@login_required
@role_required("dashboard")
def dashboard():
    # Keep /dashboard responsive on Render. Do not hydrate the latest upload from
    # MySQL here; the log showed Gunicorn killing the worker while this page was
    # closing/resetting a pooled MySQL connection. Other data pages can still load
    # database-backed records, but the landing dashboard should fail open.
    skip_dataset_reload = True
    try:
        dataset = get_app_state().processed_data
    except Exception:
        dataset = None

    dashboard_summary = _dashboard_recovery_summary()
    upload_status = _state_only_upload_status()
    current_role = get_session_role()
    try:
        role_dashboard = get_role_dashboard_context(current_role, dashboard_summary, upload_status)
    except Exception as role_exc:
        print(f"[StockWise dashboard role fallback] {type(role_exc).__name__}: {role_exc}", flush=True)
        role_dashboard = {
            "role": normalize_role(current_role),
            "title": "Dashboard",
            "subtitle": "",
            "has_data": False,
            "empty_prompt": "Upload sales data to start generating store insights.",
            "primary_action_label": "Upload New Data" if role_can_access("upload_data", current_role) else "",
            "primary_action_url": url_for("upload_data") if role_can_access("upload_data", current_role) else "",
            "secondary_actions": [],
            "show_owner_dashboard": normalize_role(current_role) == "Owner",
            "show_manager_dashboard": normalize_role(current_role) == "Store Manager",
            "show_assistant_dashboard": normalize_role(current_role) == "Operational Assistant",
            "show_full_owner_dashboard": normalize_role(current_role) == "Owner",
            "show_operational_chart": False,
            "show_buyer_behavior": False,
            "show_priority_table": False,
            "show_simple_tasks": normalize_role(current_role) == "Operational Assistant",
            "review_rows": [],
            "high_risk_rows": [],
            "low_stock_rows": [],
            "assistant_task_rows": [],
            "manager_focus_items": [],
            "assistant_reminders": [],
            "stable_count": 0,
            "monitored_count": 0,
            "low_stock_count": 0,
            "review_count": 0,
            "high_risk_count": 0,
            "last_processed_label": upload_status.get("last_processed_label", "No data yet"),
            "latest_upload_status_label": upload_status.get("freshness_label", "No uploaded sales data yet"),
            "latest_results_label": "No generated results yet",
        }

    return render_page(
        "dashboard.html",
        data=dataset,
        processed_data=dataset,
        _skip_dataset_reload=skip_dataset_reload,
        dashboard_summary=dashboard_summary,
        role_dashboard=role_dashboard,
        metrics=dashboard_summary["metrics"],
        upload_status=upload_status,
        recent_activity=[],
        model_ui_summary=_empty_model_ui_summary(),
        processed_filename=upload_status.get("processed_filename"),
    )


@app.route("/upload_data", methods=["GET", "POST"])
@login_required
@role_required("upload_data")
def upload_data():
    """Active Upload Sales Data route.

    The upload workflow implementation lives in _upload_data_phase_with_models().
    Keeping this thin route wrapper avoids Flask endpoint overrides and makes the
    active upload path clear for future maintenance.
    """
    return _upload_data_phase_with_models()


@app.route("/insights")
@login_required
@role_required("insights")
def insights():
    dataset = get_processed_dataset()
    active_tab = request.args.get("tab", "forecast").strip().lower()
    if active_tab not in {"forecast", "risk"}:
        active_tab = "forecast"

    try:
        forecast_summary = get_forecast_summary(dataset)
    except Exception:
        forecast_summary = _empty_forecast_summary()

    try:
        risk_summary = get_stock_risk_summary(dataset)
    except Exception:
        risk_summary = _empty_stock_risk_summary()

    return render_page(
        "insights.html",
        active_tab=active_tab,
        metrics=None,
        forecast_summary=forecast_summary,
        risk_summary=risk_summary,
    )


@app.route("/demand_forecast")
@login_required
@role_required("insights")
def demand_forecast():
    return redirect(url_for("insights", tab="forecast"))


@app.route("/stock_risk")
@login_required
@role_required("insights")
def stock_risk():
    return redirect(url_for("insights", tab="risk"))


@app.route("/products")
@login_required
@role_required("products")
def products_redirect():
    return redirect(url_for("product_list"))


@app.route("/product_list")
@login_required
@role_required("products")
def product_list():
    product_summary = get_product_summary(get_processed_dataset())
    current_role = get_session_role()
    product_view_mode = {
        "Owner": "full",
        "Store Manager": "full",
        "Operational Assistant": "basic",
    }.get(current_role, "basic")

    inventory_rows = [row.copy() for row in product_summary["inventory_rows"]]
    detail_map = {name: detail.copy() for name, detail in product_summary["detail_map"].items()}
    stock_statuses = list(product_summary["stock_statuses"])
    product_summary_counts = product_summary["summary_counts"]

    if product_view_mode == "basic":
        simplified_statuses = []
        for row in inventory_rows:
            assistant_status = get_assistant_stock_status(row)
            row["stock_status"] = assistant_status
            if assistant_status not in simplified_statuses:
                simplified_statuses.append(assistant_status)
            product_name = row.get("product_name")
            if product_name in detail_map:
                detail_map[product_name]["stock_status"] = assistant_status
                detail_map[product_name]["main_reason"] = get_assistant_stock_concern(row)
                detail_map[product_name]["why_flagged"] = get_assistant_stock_concern(row)
                detail_map[product_name]["suggested_action"] = get_assistant_suggested_action(row)
                detail_map[product_name]["stock_preparation_guide"] = (
                    "Check the shelf display, review the storage area, and inform the Store Manager if the item needs replenishment."
                    if assistant_status in {"Low", "Monitor"}
                    else "Keep this item on normal shelf checks and report only if the stock changes quickly."
                )
        stock_statuses = [status for status in ["Low", "Monitor", "Stable"] if status in simplified_statuses]
        product_summary_counts = {
            "total_products": len(inventory_rows),
            "low_stock": sum(1 for row in inventory_rows if row.get("stock_status") == "Low"),
            "needs_review": sum(1 for row in inventory_rows if row.get("stock_status") == "Monitor"),
            "monitor_products": sum(1 for row in inventory_rows if row.get("stock_status") == "Monitor"),
            "stable_products": sum(1 for row in inventory_rows if row.get("stock_status") == "Stable"),
        }

    return render_page(
        "product_list.html",
        inventory_rows=inventory_rows,
        categories=product_summary["categories"],
        stock_statuses=stock_statuses,
        risk_levels=product_summary["risk_levels"],
        product_detail_map=detail_map,
        product_summary_counts=product_summary_counts,
        product_view_mode=product_view_mode,
    )


@app.route("/reports")
@login_required
@role_required("reports")
def reports():
    dataset = get_processed_dataset()
    settings_preferences = get_settings_preferences()
    default_report_type = settings_preferences.get("default_report_type", "demand_forecast_summary")
    report_type = request.args.get("report_type", default_report_type).strip() or default_report_type
    period_key = request.args.get("period", settings_preferences.get("default_report_period", "last_30_days")).strip() or "last_30_days"
    legacy_product_filter = request.args.get("product_name", "").strip()
    category_filter = request.args.get("category", "").strip()
    risk_filter = request.args.get("risk", "").strip()
    product_filters = [name.strip() for name in request.args.getlist("products") if name.strip()]
    if legacy_product_filter and not product_filters:
        product_filters = [legacy_product_filter]

    report_context = get_report_context(
        dataset,
        report_type,
        period_key,
        product_filter=legacy_product_filter,
        risk_filter=risk_filter,
        category_filter=category_filter,
        product_filters=product_filters,
    )
    report_summary = get_report_summary(dataset)

    return render_page(
        "reports.html",
        report_context=report_context,
        report_rows=report_context["rows"],
        report_period_options=get_report_period_options(),
        report_type_options=get_report_type_options(),
        product_names=report_summary["product_names"],
        product_category_groups=report_summary.get("product_category_groups", {}),
        report_categories=report_summary.get("categories", get_category_filter_options()),
        report_risk_levels=report_summary["risk_levels"],
    )


@app.route("/reports/export/log", methods=["POST"])
@login_required
@role_required("reports")
def log_report_export_activity():
    payload = request.get_json(silent=True) or {}
    export_format = str(payload.get("format") or "Report").strip().upper() or "Report"
    export_status = str(payload.get("status") or "Success").strip().title() or "Success"
    if export_status not in {"Success", "Failed", "Warning", "Pending"}:
        export_status = "Success"

    action = "Report exported" if export_status == "Success" else "Report export failed"
    add_activity_log(action, "Reports", export_status)
    return jsonify({"success": True, "format": export_format, "status": export_status})


@app.route("/reports/export/xlsx", methods=["POST"])
@login_required
@role_required("reports")
def export_report_xlsx():
    """Build an XLSX report from the currently visible report table rows."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    except Exception:
        return jsonify({"success": False, "message": "XLSX export is not available in this setup."}), 500

    payload = request.get_json(silent=True) or {}
    headers = payload.get("headers") or []
    rows = payload.get("rows") or []

    if not headers or not rows:
        return jsonify({"success": False, "message": "No visible report records are available to export."}), 400

    headers = [str(item or "").strip() for item in headers]
    clean_rows: list[list[str]] = []
    for row in rows:
        if isinstance(row, list):
            clean_rows.append([str(cell or "").strip() for cell in row])

    if not clean_rows:
        return jsonify({"success": False, "message": "No visible report records are available to export."}), 400

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "StockWise Report"

    sheet.append(headers)
    for row in clean_rows:
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="FFE58A")
    thin_border = Border(bottom=Side(style="thin", color="E8E1CB"))
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="333333")
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for column_cells in sheet.columns:
        max_length = 12
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            max_length = max(max_length, min(len(str(cell.value or "")), 42))
        sheet.column_dimensions[column_letter].width = max_length + 2

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"StockWise_Report_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def get_employee_summary_for_settings() -> dict[str, int]:
    employees = get_employees_for_current_store(include_removed=True) if is_owner_user() else []
    return {
        "active": sum(1 for employee in employees if employee.get("status_key") == "active"),
        "deactivated": sum(1 for employee in employees if employee.get("status_key") == "deactivated"),
        "pending": sum(1 for employee in employees if str(employee.get("last_login", "")).lower() in {"not yet", ""}),
        "removed": sum(1 for employee in employees if employee.get("status_key") == "removed"),
        "total": len(employees),
    }


def get_data_management_summary() -> dict[str, str]:
    dataset = get_processed_dataset()
    state = get_app_state()
    latest_dataset = state.processed_filename or get_processed_filename() or "No processed dataset yet"
    if dataset is None or getattr(dataset, "empty", True):
        return {
            "latest_dataset": latest_dataset,
            "sales_records": "0",
            "products_detected": "0",
            "date_coverage": "No date coverage yet",
            "processed_at": "Not processed yet",
            "has_processed_data": False,
        }
    products_detected = 0
    if "product_name" in dataset.columns:
        products_detected = int(dataset["product_name"].dropna().astype(str).str.strip().replace("", pd.NA).dropna().nunique())
    processed_at = state.processed_at
    try:
        processed_label = processed_at.strftime("%b %d, %Y %I:%M %p") if processed_at else "Not processed yet"
    except Exception:
        processed_label = str(processed_at or "Not processed yet")
    return {
        "latest_dataset": latest_dataset,
        "sales_records": f"{len(dataset):,}",
        "products_detected": f"{products_detected:,}",
        "date_coverage": infer_coverage_period(dataset),
        "processed_at": processed_label,
        "has_processed_data": True,
    }


@app.route("/settings", methods=["GET", "POST"])
@login_required
@role_required("settings")
def settings():
    settings_message, settings_message_type = read_settings_notice()

    if request.method == "POST":
        settings_message, settings_message_type = save_settings_from_form(request.form, request.files)

    return render_page(
        "settings.html",
        settings_preferences=get_settings_preferences(),
        settings_message=settings_message,
        settings_message_type=settings_message_type,
        report_period_options=get_report_period_options(),
        report_type_options=get_report_type_options(),
        forecast_range_options=get_forecast_ranges(),
        upload_mode_options=get_upload_mode_options(),
        store_type_options=get_store_type_options(),
        user_role_options=get_user_role_options(),
        currency_options=get_currency_options(),
        product_view_options=get_default_product_view_options(),
        yes_no_options=get_yes_no_options(),
        data_date_format_options=get_data_date_format_options(),
        data_time_format_options=get_data_time_format_options(),
        payday_handling_options=get_payday_handling_options(),
        duplicate_handling_options=get_duplicate_handling_options(),
        column_mapping_fields=DATA_FORMAT_SYSTEM_FIELDS,
        standard_template_url=url_for("download_standard_template"),
        employee_role_options=EMPLOYEE_ROLE_OPTIONS,
        employees=get_employees_for_current_store(include_removed=False) if is_owner_user() else [],
        removed_employees=[employee for employee in get_employees_for_current_store(include_removed=True) if employee.get("status_key") == "removed"] if is_owner_user() else [],
        activity_logs=get_activity_logs_for_current_store(limit=5) if is_owner_user() else [],
        employee_status_options=get_employee_status_options(),
        employee_summary=get_employee_summary_for_settings(),
        data_management_summary=get_data_management_summary(),
        is_owner_account=is_owner_user(),
        settings_allowed_sections=get_settings_allowed_sections(),
    )


def apply_activity_date_range(filters: dict[str, str]) -> dict[str, str]:
    date_range = (filters.get("date_range") or "all").strip() or "all"
    filters["date_range"] = date_range
    today = datetime.now().date()

    if date_range == "today":
        filters["date_from"] = today.strftime("%Y-%m-%d")
        filters["date_to"] = today.strftime("%Y-%m-%d")
    elif date_range == "last_7_days":
        filters["date_from"] = (today - timedelta(days=6)).strftime("%Y-%m-%d")
        filters["date_to"] = today.strftime("%Y-%m-%d")
    elif date_range == "last_30_days":
        filters["date_from"] = (today - timedelta(days=29)).strftime("%Y-%m-%d")
        filters["date_to"] = today.strftime("%Y-%m-%d")
    elif date_range == "this_month":
        filters["date_from"] = today.replace(day=1).strftime("%Y-%m-%d")
        filters["date_to"] = today.strftime("%Y-%m-%d")
    elif date_range == "custom":
        filters["date_from"] = (filters.get("date_from") or "").strip()
        filters["date_to"] = (filters.get("date_to") or "").strip()
    else:
        filters["date_range"] = "all"
        filters["date_from"] = ""
        filters["date_to"] = ""

    if filters.get("date_range") == "custom" and filters.get("date_from") and filters.get("date_to"):
        filters["date_range_label"] = f"Custom: {filters['date_from']} to {filters['date_to']}"
    else:
        filters["date_range_label"] = ""

    return filters


@app.route("/activity_logs")
@login_required
@role_required("settings")
def activity_logs():
    if not is_owner_user():
        add_activity_log("Access denied", "Activity Logs", "Blocked")
        return render_template("access_denied.html", blocked_page="Activity Logs", user_role=get_session_role()), 403
    filters = apply_activity_date_range({
        "employee_id": request.args.get("employee_id", ""),
        "user": request.args.get("user", ""),
        "role": request.args.get("role", ""),
        "module": request.args.get("module", ""),
        "action": request.args.get("action", ""),
        "status": request.args.get("status", ""),
        "date_range": request.args.get("date_range", "all"),
        "date_from": request.args.get("date_from", ""),
        "date_to": request.args.get("date_to", ""),
    })
    filters["user_label"] = get_activity_selected_user_label(filters.get("employee_id", ""), filters.get("user", ""))
    if str(filters.get("employee_id") or "").strip().isdigit() and filters.get("user_label") != "All users":
        filters["user"] = filters["user_label"]

    logs, filter_options = get_full_activity_logs_for_current_store(filters=filters, limit=250)
    return render_page(
        "activity_logs.html",
        activity_logs=logs,
        activity_filters=filters,
        activity_filter_options=filter_options,
        activity_label_options=get_activity_filter_label_options(filter_options),
    )


@app.route("/team_access")
@login_required
@role_required("settings")
def team_access():
    if not is_owner_user():
        add_activity_log("Access denied", "Team & Access", "Blocked")
        return render_template("access_denied.html", blocked_page="Team & Access", user_role=get_session_role()), 403
    settings_message, settings_message_type = read_settings_notice()
    active_employees = get_employees_for_current_store(include_removed=False)
    all_employees = get_employees_for_current_store(include_removed=True)
    removed_employees = [employee for employee in all_employees if employee.get("status_key") == "removed"]
    return render_page(
        "team_access.html",
        settings_message=settings_message,
        settings_message_type=settings_message_type,
        employee_summary=get_employee_summary_for_settings(),
        employees=active_employees,
        removed_employees=removed_employees,
        employee_role_options=EMPLOYEE_ROLE_OPTIONS,
        employee_status_options=get_employee_status_options(),
    )




def clear_processed_data_for_current_store() -> tuple[bool, str]:
    """Safely remove processed sales data from the active workspace without touching accounts or store settings."""
    if not is_owner_user():
        return False, "Only the Owner can clear processed data."

    user_id = get_current_user_id()
    if not user_id:
        return False, "Your session could not be verified. Please log in again."

    try:
        ensure_model_tables()
        store_user_ids = get_store_user_ids(user_id)
        placeholders, params = make_in_clause(store_user_ids)
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        try:
            cursor.execute(
                f"""
                SELECT upload_id
                FROM uploads
                WHERE user_id IN ({placeholders}) AND upload_status = 'processed'
                """,
                params,
            )
            upload_ids = [int(row["upload_id"]) for row in (cursor.fetchall() or []) if row.get("upload_id")]
            if not upload_ids:
                clear_processed_dataset()
                return False, "There is no processed sales data to clear."

            upload_placeholders, upload_params = make_in_clause(upload_ids)

            # Remove generated model outputs tied to the cleared uploads.  Account,
            # store, employee, and activity-log records are intentionally preserved.
            for table_name in ("forecasts", "stockout_predictions", "model_runs"):
                try:
                    cursor.execute(f"DELETE FROM {table_name} WHERE upload_id IN ({upload_placeholders})", upload_params)
                except Exception:
                    pass

            cursor.execute(
                f"""
                UPDATE uploads
                SET upload_status = 'cleared',
                    remarks = CONCAT(COALESCE(remarks, ''), CASE WHEN COALESCE(remarks, '') = '' THEN '' ELSE ' | ' END, 'Cleared by Owner from Data Management')
                WHERE upload_id IN ({upload_placeholders})
                """,
                upload_params,
            )
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            cursor.close()
            conn.close()

        clear_processed_dataset()
        clear_selected_dataset()
        add_activity_log("Processed data cleared", "Data Management", "Success")
        return True, "Processed sales data has been cleared. Accounts, employees, store settings, and activity logs were kept."
    except Exception:
        return False, "Processed sales data could not be cleared right now. Please check the database connection and try again."


@app.route("/data_management")
@login_required
@role_required("settings")
def data_management():
    if not is_owner_user():
        add_activity_log("Access denied", "Data Management", "Blocked")
        return render_template("access_denied.html", blocked_page="Data Management", user_role=get_session_role()), 403
    settings_message, settings_message_type = read_settings_notice()
    return render_page(
        "data_management.html",
        data_management_summary=get_data_management_summary(),
        settings_message=settings_message,
        settings_message_type=settings_message_type,
    )


@app.route("/data_management/clear_processed", methods=["POST"])
@login_required
@role_required("settings")
def clear_processed_data_route():
    if not is_owner_user():
        add_activity_log("Access denied", "Data Management", "Blocked")
        set_settings_notice("Only the Owner can clear processed data.", "error")
        return redirect(url_for("settings"))
    success, message = clear_processed_data_for_current_store()
    set_settings_notice(message, "success" if success else "error")
    return redirect(url_for("data_management"))


def redirect_non_owner_team_action():
    if is_owner_user():
        return None
    add_activity_log("Access denied", "Team & Access", "Blocked")
    set_settings_notice("Only the Owner can manage employees.", "error")
    return redirect(url_for("settings"))


@app.route("/settings/employees/add", methods=["POST"])
@login_required
@role_required("settings")
def add_employee_route():
    owner_redirect = redirect_non_owner_team_action()
    if owner_redirect:
        return owner_redirect
    success, message, _temporary_password = create_employee_account(
        request.form.get("employee_first_name", ""),
        request.form.get("employee_last_name", ""),
        request.form.get("employee_email", ""),
        request.form.get("employee_role", ""),
    )
    set_settings_notice(message, "success" if success else "error")
    return redirect(url_for("team_access"))


@app.route("/settings/employees/<int:employee_id>/role", methods=["POST"])
@login_required
@role_required("settings")
def update_employee_role_route(employee_id: int):
    owner_redirect = redirect_non_owner_team_action()
    if owner_redirect:
        return owner_redirect
    role_value = request.form.get(f"employee_role_{employee_id}", request.form.get("employee_role", ""))
    success, message = update_employee_role(employee_id, role_value)
    set_settings_notice(message, "success" if success else "error")
    return redirect(url_for("team_access"))


@app.route("/settings/employees/<int:employee_id>/status", methods=["POST"])
@login_required
@role_required("settings")
def update_employee_status_route(employee_id: int):
    owner_redirect = redirect_non_owner_team_action()
    if owner_redirect:
        return owner_redirect
    status_value = request.form.get(f"employee_status_{employee_id}", request.form.get("employee_status", "deactivated"))
    success, message = update_employee_status(employee_id, status_value)
    set_settings_notice(message, "success" if success else "error")
    return redirect(url_for("team_access"))


@app.route("/settings/employees/<int:employee_id>/deactivate", methods=["POST"])
@login_required
@role_required("settings")
def deactivate_employee_route(employee_id: int):
    owner_redirect = redirect_non_owner_team_action()
    if owner_redirect:
        return owner_redirect
    success, message = deactivate_employee(employee_id)
    set_settings_notice(message, "success" if success else "error")
    return redirect(url_for("team_access"))


@app.route("/settings/employees/<int:employee_id>/rehire", methods=["POST"])
@login_required
@role_required("settings")
def rehire_employee_route(employee_id: int):
    owner_redirect = redirect_non_owner_team_action()
    if owner_redirect:
        return owner_redirect
    success, message = rehire_employee(employee_id)
    set_settings_notice(message, "success" if success else "error")
    return redirect(url_for("team_access"))


@app.route("/settings/employees/<int:employee_id>/reset_password", methods=["POST"])
@login_required
@role_required("settings")
def reset_employee_password_route(employee_id: int):
    owner_redirect = redirect_non_owner_team_action()
    if owner_redirect:
        return owner_redirect
    success, message = reset_employee_password(employee_id)
    set_settings_notice(message, "success" if success else "error")
    return redirect(url_for("team_access"))


# =========================================
# ACTIVE UPLOAD WORKFLOW IMPLEMENTATION
# =========================================

def _upload_data_phase_with_models():
    last_action = ""
    if request.method == "POST":
        action = request.form.get("action", "").strip()
        last_action = action
        if action == "update_mode":
            upload_mode = set_selected_upload_mode(request.form.get("upload_mode", get_app_state().last_upload_mode))
            upload_message, upload_message_type = refresh_upload_mode_feedback(get_selected_dataset(), upload_mode)
            if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json":
                selected_data = get_selected_dataset()
                preview_base_df = get_processed_dataset() if upload_mode == "append" else None
                _, preprocessing_summary = preprocess_dataset(selected_data, base_df=preview_base_df, upload_mode=upload_mode)
                return jsonify({
                    "ok": True,
                    "upload_mode": upload_mode,
                    "mode_label": preprocessing_summary.get("mode_label", get_upload_mode_label(upload_mode)),
                    "message": upload_message,
                    "message_type": upload_message_type,
                    "has_selected_file": selected_data is not None,
                })
        elif action == "select":
            if request.form.get("onboarding_upload") == "1":
                session.pop("show_first_upload_prompt", None)
            file = request.files.get("file")
            upload_mode = set_selected_upload_mode(request.form.get("upload_mode", get_app_state().last_upload_mode))
            if file is None or file.filename == "":
                set_upload_feedback("No file selected. Please choose a CSV or Excel file.", "error")
            elif not allowed_file(file.filename):
                clear_selected_dataset()
                set_upload_feedback("Invalid file format. Please upload a CSV, XLSX, or XLS file.", "error")
            else:
                try:
                    file_size = get_upload_file_size(file)
                    selected_data, selected_filename = read_uploaded_file(file)
                    store_selected_dataset(selected_data, selected_filename, upload_mode=upload_mode, file_size=file_size, file_type=get_file_type_label(selected_filename))
                    add_activity_log("File upload", "Upload Sales Data", "Success")
                    validation = analyze_upload_dataset(selected_data)
                    has_overlap = coverage_overlap(selected_data, get_processed_dataset())
                    if validation.get("missing_required_columns"):
                        missing_columns = ", ".join(validation.get("missing_required_columns") or [])
                        set_upload_feedback(f"File selected, but these required columns are missing: {missing_columns}. Please update the file or review Data Format Settings column mapping before generating results.", "error")
                        notify_upload_event("Upload failed", "A selected sales file is missing required columns.", "error", success=False)
                    elif validation.get("valid_rows", 0) <= 0:
                        set_upload_feedback("File selected, but no valid sales records were found. Please check the dates, product names, and quantities.", "error")
                        notify_upload_event("Upload failed", "A selected sales file had no valid sales records.", "error", success=False)
                    else:
                        if has_overlap:
                            set_upload_feedback(
                                "Coverage overlap detected. You may append missing records or replace the previous processed sales data.",
                                "warning",
                            )
                            notify_upload_event("Upload needs review", "A selected sales file has coverage overlap and needs review before results are generated.", "warning", success=False)
                        elif validation.get("total_invalid_rows", 0) > 0 or validation.get("duplicate_rows", 0) > 0:
                            set_upload_feedback(
                                "File selected with review notes. Some rows may be cleaned or excluded before analysis.",
                                "warning",
                            )
                        else:
                            set_upload_feedback("File selected successfully. Review the record check summary, then continue through the next steps when ready.", "info")
                except ValueError as exc:
                    clear_selected_dataset()
                    add_activity_log("Failed upload", "Upload Sales Data", "Failed")
                    notify_upload_event("Upload failed", "The selected file could not be read. Please review the file and try again.", "error", success=False)
                    set_upload_feedback(str(exc), "error")
                except Exception:
                    clear_selected_dataset()
                    add_activity_log("Failed upload", "Upload Sales Data", "Failed")
                    notify_upload_event("Upload failed", "The selected file could not be read. Please review the file and try again.", "error", success=False)
                    set_upload_feedback("Unable to read the selected file. Please check the file and try again.", "error")
        elif action == "process":
            selected_data = get_selected_dataset()
            selected_filename = get_selected_filename()
            upload_mode = set_selected_upload_mode(request.form.get("upload_mode", get_app_state().last_upload_mode))
            validation = analyze_upload_dataset(selected_data)
            if selected_data is None or selected_filename is None:
                set_upload_feedback("No file is ready for processing. Please upload a file first.", "error")
            elif not validation["can_process"]:
                if validation.get("missing_required_columns"):
                    missing_columns = ", ".join(validation.get("missing_required_columns") or [])
                    set_upload_feedback(f"The uploaded file is missing required columns: {missing_columns}. Please update the file or review Data Format Settings column mapping before generating results.", "error")
                else:
                    set_upload_feedback("The selected file still needs review before results can be generated. Please check the record summary.", "error")
            else:
                base_df = get_processed_dataset() if upload_mode == "append" else None
                processed_data, preprocessing_summary = preprocess_dataset(selected_data, base_df=base_df, upload_mode=upload_mode)
                if processed_data is None or processed_data.empty:
                    set_upload_feedback("Results could not be generated because no valid sales records remained after cleaning. Please review the uploaded file.", "error")
                else:
                    processed_filename = selected_filename if upload_mode != "append" else f"{selected_filename} (updated)"
                    current_user_id = get_current_user_id()
                    if current_user_id is None:
                        set_upload_feedback("Your session could not be verified. Please log in again.", "error")
                    else:
                        try:
                            upload_id = save_processed_dataset_to_database(user_id=current_user_id, filename=processed_filename, processed_df=processed_data, upload_mode=upload_mode)
                            store_processed_dataset(processed_data, processed_filename, upload_id=upload_id)
                            store_selected_dataset(selected_data, selected_filename, upload_mode=upload_mode)
                            try:
                                run_model_pipeline(current_user_id, upload_id, processed_data)
                                warm_up_generated_page_contexts(processed_data)
                                add_result_notifications(current_user_id, processed_data)
                                add_activity_log("Generate results", "Upload Sales Data", "Success")
                                set_upload_feedback(
                                    "Results generated successfully.",
                                    "success",
                                )
                            except Exception:
                                notify_upload_event("Upload needs review", "Sales records were saved, but results could not be generated yet.", "warning", success=False)
                                add_activity_log("Generate results", "Upload Sales Data", "Failed")
                                clear_model_cache()
                                set_upload_feedback(
                                    "Sales records were saved, but forecast and stockout risk results could not be generated yet. Please review the uploaded records and try again.",
                                    "error",
                                )
                        except Exception:
                            notify_upload_event("Upload failed", "The uploaded records could not be saved or processed. Please review the file and try again.", "error", success=False)
                            add_activity_log("Generate results", "Upload Sales Data", "Failed")
                            set_upload_feedback(
                                "The uploaded records could not be saved or processed right now. Please check the file and try again.",
                                "error",
                            )
        elif action == "clear_selected":
            clear_selected_dataset()
            set_upload_feedback("The selected file has been cleared from the upload workspace.", "info")
        elif action == "disable_processed":
            clear_processed_dataset()
            set_upload_feedback("The processed sales data has been cleared from memory.", "info")
        elif action == "finish_upload":
            clear_selected_dataset()
            set_upload_feedback("Upload workflow reset. The latest processed sales data remains available across the system.", "info")

    selected_data = get_selected_dataset()
    processed_data = get_processed_dataset()
    upload_message, upload_message_type = get_upload_feedback()
    validation_summary = analyze_upload_dataset(selected_data)
    current_upload_mode = get_app_state().last_upload_mode
    if selected_data is None and upload_message is None:
        current_upload_mode = set_selected_upload_mode(get_settings_preferences().get("default_upload_mode", current_upload_mode))
    preview_base_df = processed_data if current_upload_mode == "append" else None
    preprocessing_preview, preprocessing_summary = preprocess_dataset(selected_data, base_df=preview_base_df, upload_mode=current_upload_mode)
    upload_status = get_upload_status()
    wizard_state = get_upload_wizard_state(selected_data, validation_summary, get_processed_filename(), upload_message_type=upload_message_type, last_action=last_action)
    return render_page('upload_data.html', selected_data=selected_data, processed_data=processed_data, selected_filename=get_selected_filename(), selected_file_info=get_selected_file_info(), upload_message=upload_message, upload_message_type=upload_message_type, upload_status=upload_status, validation_summary=validation_summary, preprocessing_summary=preprocessing_summary, preprocessing_preview=preprocessing_preview, required_columns=REQUIRED_UPLOAD_COLUMNS, optional_columns=OPTIONAL_UPLOAD_COLUMNS, preview_source=selected_data if selected_data is not None else processed_data, sample_template_url=url_for('download_sample_template'), upload_wizard=wizard_state)


if __name__ == "__main__":
    app.run(debug=True)
